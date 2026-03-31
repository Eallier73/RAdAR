from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from classification_evaluation import (
    compute_classification_loss_h,
    compute_classification_metrics,
    compute_total_classification_loss,
    walk_forward_predict_classifier,
)
from classification_targets import build_classification_model_frame, serialize_classification_target_metadata
from config import (
    DATASET_PATH,
    DATE_COLUMN,
    DEFAULT_HORIZONS,
    DEFAULT_INITIAL_TRAIN_SIZE,
    DEFAULT_LAGS,
    DEFAULT_RANDOM_STATE,
    DEFAULT_SHEET_NAME,
    FEATURE_MODE_ALL,
    FEATURE_MODE_CHOICES,
    TARGET_MODE_CLF_BANDAS_5CLASES,
    TARGET_MODE_CLF_CHOICES,
)
from data_master import get_base_feature_columns, load_master_dataset
from evaluation import resolve_feature_mode
from experiment_logger import DEFAULT_RUNS_DIR, DEFAULT_WORKBOOK, RadarExperimentTracker
from pipeline_common import (
    add_reference_comparison_args,
    build_selected_features_summary,
    normalize_reference_run_ids,
    parse_horizons,
    parse_lags,
    parse_string_sequence,
    serialize_model_params,
)


CLASSIFICATION_TASK_TYPE = "clasificacion"


def add_common_classification_args(parser: argparse.ArgumentParser, *, default_run_id: str) -> None:
    parser.add_argument("--run-id", default=default_run_id, help="Run_ID a registrar en el grid.")
    parser.add_argument(
        "--target-mode-clf",
        choices=TARGET_MODE_CLF_CHOICES,
        default=TARGET_MODE_CLF_BANDAS_5CLASES,
        help="Modo de target para clasificacion.",
    )
    parser.add_argument(
        "--lags",
        default=",".join(str(lag) for lag in DEFAULT_LAGS),
        help="Lags separados por coma. Ejemplo: 1,2,3,4",
    )
    parser.add_argument(
        "--feature-mode",
        choices=FEATURE_MODE_CHOICES,
        default=FEATURE_MODE_ALL,
        help="Modo de seleccion de features.",
    )
    parser.add_argument(
        "--initial-train-size",
        type=int,
        default=DEFAULT_INITIAL_TRAIN_SIZE,
        help=f"Tamano inicial de entrenamiento. Default: {DEFAULT_INITIAL_TRAIN_SIZE}",
    )
    parser.add_argument(
        "--horizons",
        default=",".join(str(horizon) for horizon in DEFAULT_HORIZONS),
        help="Horizontes a correr, separados por coma.",
    )
    parser.add_argument(
        "--dataset-path",
        type=Path,
        default=DATASET_PATH,
        help=f"Ruta del dataset maestro. Default: {DATASET_PATH}",
    )
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help=f"Hoja del dataset maestro. Default: {DEFAULT_SHEET_NAME}",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Workbook del grid de experimentos. Default: {DEFAULT_WORKBOOK}",
    )
    parser.add_argument(
        "--runs-dir",
        type=Path,
        default=DEFAULT_RUNS_DIR,
        help=f"Directorio donde se guardan runs y artefactos. Default: {DEFAULT_RUNS_DIR}",
    )
    parser.add_argument(
        "--random-seed",
        type=int,
        default=DEFAULT_RANDOM_STATE,
        help=f"Semilla base de reproducibilidad. Default: {DEFAULT_RANDOM_STATE}",
    )


def finalize_common_classification_args(args: argparse.Namespace) -> argparse.Namespace:
    args.lags = parse_lags(args.lags)
    args.horizons = parse_horizons(args.horizons)
    args.dataset_path = args.dataset_path.expanduser().resolve()
    args.workbook = args.workbook.expanduser().resolve()
    args.runs_dir = args.runs_dir.expanduser().resolve()
    if hasattr(args, "extra_reference_run_ids"):
        args.extra_reference_run_ids = parse_string_sequence(args.extra_reference_run_ids)
    return args


def build_classification_run_parameters(
    *,
    args: argparse.Namespace,
    model_name: str,
    feature_columns: list[str],
    model_params: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "task_type": CLASSIFICATION_TASK_TYPE,
        "dataset_path": str(args.dataset_path),
        "sheet_name": args.sheet_name,
        "model": model_name,
        "target_mode_clf": args.target_mode_clf,
        "feature_mode": args.feature_mode,
        "feature_columns": feature_columns,
        "lags": list(args.lags),
        "horizons": list(args.horizons),
        "initial_train_size": args.initial_train_size,
        "model_params": serialize_model_params(model_params),
    }


def build_classification_notas_config(
    *,
    args: argparse.Namespace,
    model_name: str,
    model_params: dict[str, Any] | None,
    horizon: int,
    summary_row: dict[str, Any],
) -> str:
    payload = {
        "task_type": CLASSIFICATION_TASK_TYPE,
        "model": model_name,
        "target_mode_clf": args.target_mode_clf,
        "feature_mode": args.feature_mode,
        "lags": list(args.lags),
        "initial_train_size": args.initial_train_size,
        "horizon": horizon,
        "rows_modeling": summary_row["rows_modeling"],
        "rows_eval": summary_row["rows_eval"],
        "feature_candidates": summary_row["feature_candidates"],
        "selected_feature_count_avg": summary_row["selected_feature_count_avg"],
        "target_build_metadata": summary_row["target_build_metadata"],
        "model_params": serialize_model_params(model_params),
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def load_run_horizon_classification_metrics(workbook_path: Path, run_id: str) -> dict[int, dict[str, float]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook["RESULTADOS_GRID"]
    headers = {worksheet.cell(1, c).value: c for c in range(1, worksheet.max_column + 1)}
    required_headers = [
        "Run_ID",
        "Horizonte_sem",
        "Accuracy_5clases",
        "Balanced_accuracy_5clases",
        "Macro_f1_5clases",
        "Recall_baja_fuerte",
        "Recall_baja_total",
        "Precision_baja_fuerte",
        "Recall_sube_fuerte",
        "Recall_sube_total",
        "Accuracy_3clases",
        "Macro_f1_3clases",
        "Balanced_accuracy_3clases",
        "Loss_clasificacion_h",
    ]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"El workbook no contiene columnas de clasificacion requeridas: {missing}")

    metrics_by_horizon: dict[int, dict[str, float]] = {}
    for row_idx in range(2, worksheet.max_row + 1):
        if worksheet.cell(row_idx, headers["Run_ID"]).value != run_id:
            continue
        horizon = int(worksheet.cell(row_idx, headers["Horizonte_sem"]).value)
        metrics_by_horizon[horizon] = {
            "accuracy_5clases": float(worksheet.cell(row_idx, headers["Accuracy_5clases"]).value),
            "balanced_accuracy_5clases": float(worksheet.cell(row_idx, headers["Balanced_accuracy_5clases"]).value),
            "macro_f1_5clases": float(worksheet.cell(row_idx, headers["Macro_f1_5clases"]).value),
            "recall_baja_fuerte": float(worksheet.cell(row_idx, headers["Recall_baja_fuerte"]).value),
            "recall_baja_total": float(worksheet.cell(row_idx, headers["Recall_baja_total"]).value),
            "precision_baja_fuerte": float(worksheet.cell(row_idx, headers["Precision_baja_fuerte"]).value),
            "recall_sube_fuerte": float(worksheet.cell(row_idx, headers["Recall_sube_fuerte"]).value),
            "recall_sube_total": float(worksheet.cell(row_idx, headers["Recall_sube_total"]).value),
            "accuracy_3clases": float(worksheet.cell(row_idx, headers["Accuracy_3clases"]).value),
            "macro_f1_3clases": float(worksheet.cell(row_idx, headers["Macro_f1_3clases"]).value),
            "balanced_accuracy_3clases": float(worksheet.cell(row_idx, headers["Balanced_accuracy_3clases"]).value),
            "loss_clasificacion_h": float(worksheet.cell(row_idx, headers["Loss_clasificacion_h"]).value),
        }
    if not metrics_by_horizon:
        raise ValueError(f"No se encontraron metricas de clasificacion para {run_id} en {workbook_path}.")
    return metrics_by_horizon


def build_classification_comparison_payload(
    *,
    workbook_path: Path,
    reference_run_id: str,
    clean_run_id: str,
    horizon_results: list[dict[str, Any]],
    l_total_clasificacion: float | None,
) -> dict[str, Any]:
    reference = load_run_horizon_classification_metrics(workbook_path=workbook_path, run_id=reference_run_id)
    clean = {int(item["horizonte_sem"]): item for item in horizon_results}
    comparison_rows = []
    for horizon in sorted(clean):
        reference_row = reference[horizon]
        clean_row = clean[horizon]
        comparison_rows.append(
            {
                "horizonte_sem": horizon,
                "loss_clasificacion_h_original": reference_row["loss_clasificacion_h"],
                "loss_clasificacion_h_clean": float(clean_row["loss_clasificacion_h"]),
                "delta_loss_clasificacion_h": float(clean_row["loss_clasificacion_h"])
                - reference_row["loss_clasificacion_h"],
                "accuracy_5clases_original": reference_row["accuracy_5clases"],
                "accuracy_5clases_clean": float(clean_row["accuracy_5clases"]),
                "delta_accuracy_5clases": float(clean_row["accuracy_5clases"]) - reference_row["accuracy_5clases"],
                "balanced_accuracy_5clases_original": reference_row["balanced_accuracy_5clases"],
                "balanced_accuracy_5clases_clean": float(clean_row["balanced_accuracy_5clases"]),
                "delta_balanced_accuracy_5clases": float(clean_row["balanced_accuracy_5clases"])
                - reference_row["balanced_accuracy_5clases"],
                "macro_f1_5clases_original": reference_row["macro_f1_5clases"],
                "macro_f1_5clases_clean": float(clean_row["macro_f1_5clases"]),
                "delta_macro_f1_5clases": float(clean_row["macro_f1_5clases"]) - reference_row["macro_f1_5clases"],
                "recall_baja_fuerte_original": reference_row["recall_baja_fuerte"],
                "recall_baja_fuerte_clean": float(clean_row["recall_baja_fuerte"]),
                "delta_recall_baja_fuerte": float(clean_row["recall_baja_fuerte"])
                - reference_row["recall_baja_fuerte"],
            }
        )

    return {
        "reference_run_id": reference_run_id,
        "clean_run_id": clean_run_id,
        "l_total_clasificacion_clean": float(l_total_clasificacion) if l_total_clasificacion is not None else None,
        "comparacion_por_horizonte": comparison_rows,
    }


def save_reference_classification_comparisons(
    *,
    run,
    workbook_path: Path,
    clean_run_id: str,
    reference_run_ids: tuple[str, ...],
    horizon_results: list[dict[str, Any]],
    l_total_clasificacion: float | None,
) -> None:
    for reference_run_id in reference_run_ids:
        payload = build_classification_comparison_payload(
            workbook_path=workbook_path,
            reference_run_id=reference_run_id,
            clean_run_id=clean_run_id,
            horizon_results=horizon_results,
            l_total_clasificacion=l_total_clasificacion,
        )
        run.save_json(
            payload,
            f"comparacion_vs_{reference_run_id}.json",
            artifact_type="comparacion",
            notes=f"Comparacion de la corrida de clasificacion contra {reference_run_id}.",
        )


def run_classification_experiment(
    *,
    args: argparse.Namespace,
    script_path: str | Path,
    experiment_id: str,
    family: str,
    model_name: str,
    estimator,
    model_params: dict[str, Any] | None = None,
    comentarios: str = "",
    always_include_columns: list[str] | None = None,
    reference_run_ids: tuple[str, ...] = (),
) -> dict[str, Any]:
    tracker = RadarExperimentTracker(workbook_path=args.workbook, runs_dir=args.runs_dir)
    reference_values = tracker.get_reference_values()
    df = load_master_dataset(dataset_path=args.dataset_path, sheet_name=args.sheet_name)
    base_feature_columns = get_base_feature_columns(df)

    run = tracker.start_run(
        run_id=args.run_id,
        experiment_id=experiment_id,
        family=family,
        model=model_name,
        script_path=script_path,
        parametros=build_classification_run_parameters(
            args=args,
            model_name=model_name,
            feature_columns=base_feature_columns,
            model_params=model_params,
        ),
    )

    predictions_by_horizon: dict[int, pd.DataFrame] = {}
    horizon_results: list[dict[str, Any]] = []
    rows_summary: list[dict[str, Any]] = []
    dataset_periodo = f"{df[DATE_COLUMN].min().date()} a {df[DATE_COLUMN].max().date()}"

    for horizon in args.horizons:
        modeling_df, modeling_features, target_columns = build_classification_model_frame(
            df=df,
            horizon=horizon,
            feature_columns=base_feature_columns,
            lags=args.lags,
            target_mode_clf=args.target_mode_clf,
        )
        predictions = walk_forward_predict_classifier(
            estimator=estimator,
            data=modeling_df,
            feature_columns=modeling_features,
            class_label_column=target_columns["class_label_column"],
            class_id_column=target_columns["class_id_column"],
            class_3_label_column=target_columns["class_3_label_column"],
            delta_column=target_columns["delta_column"],
            initial_train_size=args.initial_train_size,
            feature_mode=args.feature_mode,
            always_include_columns=always_include_columns,
        )
        predictions["horizonte_sem"] = horizon
        predictions["target_mode_clf"] = args.target_mode_clf
        predictions["feature_mode"] = args.feature_mode
        predictions["run_id"] = args.run_id
        predictions["model_name"] = model_name
        predictions_by_horizon[horizon] = predictions

        if args.feature_mode != FEATURE_MODE_ALL:
            selected_features_summary = build_selected_features_summary(predictions)
            run.save_dataframe(
                selected_features_summary,
                f"features_seleccionadas_h{horizon}.csv",
                artifact_type="seleccion_features",
                notes=(
                    "Resumen de combinaciones de features seleccionadas por fold externo. "
                    "El filtro se calcula solo con el train de cada fold externo."
                ),
            )

        summary_row = {
            "horizonte_sem": horizon,
            "rows_modeling": len(modeling_df),
            "rows_eval": len(predictions),
            "feature_candidates": len(modeling_features),
            "selected_feature_count_avg": float(predictions["selected_feature_count"].mean()),
            "selected_feature_count_min": int(predictions["selected_feature_count"].min()),
            "selected_feature_count_max": int(predictions["selected_feature_count"].max()),
            "feature_mode": args.feature_mode,
            "target_mode_clf": args.target_mode_clf,
            "target_build_metadata": serialize_classification_target_metadata(
                horizon=horizon,
                target_mode_clf=args.target_mode_clf,
            ),
        }
        rows_summary.append(summary_row)

        metrics = compute_classification_metrics(predictions)
        loss_h = compute_classification_loss_h(metrics=metrics, horizon=horizon, reference_values=reference_values)
        run.save_json(
            metrics["matriz_confusion_5clases"],
            f"matriz_confusion_h{horizon}.json",
            artifact_type="matriz_confusion",
            notes=f"Matriz de confusion 5 clases para horizonte {horizon}.",
        )
        run.save_json(
            metrics["soporte_por_clase"],
            f"soporte_clases_h{horizon}.json",
            artifact_type="soporte_clases",
            notes=f"Soporte por clase 5 clases para horizonte {horizon}.",
        )

        horizon_results.append(
            {
                "horizonte_sem": horizon,
                "target": args.target_mode_clf,
                "variables_temporales": f"y_t + lags {list(args.lags)}",
                "variables_tematicas": ", ".join(base_feature_columns),
                "transformacion": "sin_escalar",
                "seleccion_variables": args.feature_mode,
                "validacion": "walk-forward_expanding",
                "dataset_periodo": dataset_periodo,
                "notas_config": build_classification_notas_config(
                    args=args,
                    model_name=model_name,
                    model_params=model_params,
                    horizon=horizon,
                    summary_row=summary_row,
                ),
                "estado": "corrido",
                "comentarios": comentarios,
                "task_type": CLASSIFICATION_TASK_TYPE,
                "loss_h": loss_h,
                "loss_clasificacion_h": loss_h,
                **metrics,
            }
        )

    for horizon, predictions in predictions_by_horizon.items():
        run.save_dataframe(
            predictions,
            f"predicciones_h{horizon}.csv",
            artifact_type="predicciones",
            notes=f"Predicciones walk-forward de clasificacion para horizonte {horizon}.",
        )

    run.save_json(
        rows_summary,
        "resumen_modeling_horizontes.json",
        artifact_type="resumen",
        notes="Resumen de tamanos, target de clasificacion y seleccion de features por horizonte.",
    )

    total_classification = compute_total_classification_loss(horizon_results=horizon_results)
    if reference_run_ids:
        save_reference_classification_comparisons(
            run=run,
            workbook_path=tracker.workbook_path,
            clean_run_id=args.run_id,
            reference_run_ids=reference_run_ids,
            horizon_results=horizon_results,
            l_total_clasificacion=total_classification["l_total_clasificacion"],
        )

    run.finalize(
        horizon_results=horizon_results,
        target=args.target_mode_clf,
        variables_temporales=f"y_t + lags {list(args.lags)}",
        variables_tematicas=", ".join(base_feature_columns),
        transformacion="sin_escalar",
        seleccion_variables=args.feature_mode,
        validacion="walk-forward_expanding",
        dataset_periodo=dataset_periodo,
        notas_config=json.dumps(rows_summary, ensure_ascii=False),
        estado="corrido",
        comentarios=(
            f"{comentarios} | L_total_Clasificacion={total_classification['l_total_clasificacion']:.6f}"
            if comentarios and total_classification["l_total_clasificacion"] is not None
            else (
                f"L_total_Clasificacion={total_classification['l_total_clasificacion']:.6f}"
                if total_classification["l_total_clasificacion"] is not None
                else comentarios
            )
        ),
        summary_metrics=total_classification,
        task_type=CLASSIFICATION_TASK_TYPE,
    )

    return {
        "workbook_path": tracker.workbook_path,
        "run_dir": run.run_dir,
        "horizon_results": horizon_results,
        "rows_summary": rows_summary,
        "l_total_clasificacion": total_classification["l_total_clasificacion"],
    }


def prepare_classification_references(
    reference_run_id: str | None,
    extra_reference_run_ids: tuple[str, ...] | list[str] | None = None,
) -> tuple[str, ...]:
    return normalize_reference_run_ids(reference_run_id, extra_reference_run_ids)


__all__ = [
    "CLASSIFICATION_TASK_TYPE",
    "add_common_classification_args",
    "add_reference_comparison_args",
    "finalize_common_classification_args",
    "prepare_classification_references",
    "run_classification_experiment",
]
