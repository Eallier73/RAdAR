#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from experiment_logger import (
    ARTEFACTOS_SHEET,
    DEFAULT_RUNS_DIR,
    DEFAULT_WORKBOOK,
    RESULTADOS_SHEET,
    SUMMARY_SHEET,
    RadarExperimentTracker,
    find_header_map,
)


ROOT_DIR = Path(__file__).resolve().parents[2]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Repara o retrocarga resultados numericos al grid desde artefactos existentes.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Workbook maestro. Default: {DEFAULT_WORKBOOK}",
    )
    parser.add_argument(
        "--runs-dir",
        type=Path,
        default=DEFAULT_RUNS_DIR,
        help=f"Directorio de runs. Default: {DEFAULT_RUNS_DIR}",
    )
    parser.add_argument(
        "--run-id",
        action="append",
        default=[],
        help="Run_ID especifico a reparar. Se puede repetir.",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Reprocesa todos los runs detectados, aunque ya tengan metricas numericas.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo reporta que haria, sin escribir en el Excel.",
    )
    return parser.parse_args()


def iter_run_dirs(runs_dir: Path) -> list[Path]:
    run_dirs: list[Path] = []
    for metadata_path in sorted(runs_dir.glob("*/metadata_run.json")):
        run_dir = metadata_path.parent
        if (run_dir / "metricas_horizonte.json").exists():
            run_dirs.append(run_dir)
    return run_dirs


def load_workbook_state(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True)

    summary_ws = workbook[SUMMARY_SHEET]
    results_ws = workbook[RESULTADOS_SHEET]
    artifacts_ws = workbook[ARTEFACTOS_SHEET]

    summary_headers = find_header_map(summary_ws)
    results_headers = find_header_map(results_ws)
    artifacts_headers = find_header_map(artifacts_ws)

    summary_by_run: dict[str, dict[str, Any]] = {}
    results_by_run: dict[str, list[dict[str, Any]]] = defaultdict(list)
    artifacts_by_run: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row_idx in range(2, summary_ws.max_row + 1):
        run_id = summary_ws.cell(row_idx, summary_headers["Run_ID"]).value
        if run_id in (None, ""):
            continue
        summary_by_run[str(run_id).strip()] = row_to_dict(summary_ws, summary_headers, row_idx)

    for row_idx in range(2, results_ws.max_row + 1):
        run_id = results_ws.cell(row_idx, results_headers["Run_ID"]).value
        if run_id in (None, ""):
            continue
        results_by_run[str(run_id).strip()].append(row_to_dict(results_ws, results_headers, row_idx))

    for row_idx in range(2, artifacts_ws.max_row + 1):
        run_id = artifacts_ws.cell(row_idx, artifacts_headers["Run_ID"]).value
        if run_id in (None, ""):
            continue
        artifacts_by_run[str(run_id).strip()].append(row_to_dict(artifacts_ws, artifacts_headers, row_idx))

    return {
        "summary_by_run": summary_by_run,
        "results_by_run": results_by_run,
        "artifacts_by_run": artifacts_by_run,
    }


def row_to_dict(worksheet, headers: dict[str, int], row_idx: int) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for header, col_idx in headers.items():
        row[header] = worksheet.cell(row_idx, col_idx).value
    return row


def needs_backfill(
    run_id: str,
    horizon_results: list[dict[str, Any]],
    summary_row: dict[str, Any] | None,
    result_rows: list[dict[str, Any]],
) -> tuple[bool, str]:
    if not summary_row:
        return True, "no tiene fila en RUN_SUMMARY"

    required_summary = [
        "Horizons_loaded",
        "Sum_weighted_Loss_h",
        "L_total_Radar",
        "Avg_MAE",
        "Avg_RMSE",
        "Dir_acc_prom",
        "Det_caidas_prom",
    ]
    missing_summary = [field for field in required_summary if summary_row.get(field) is None]
    if missing_summary:
        return True, f"faltan metricas en RUN_SUMMARY: {', '.join(missing_summary)}"

    expected_horizons = sorted(int(item["horizonte_sem"]) for item in horizon_results)
    present_horizons = sorted(
        int(row["Horizonte_sem"])
        for row in result_rows
        if row.get("Horizonte_sem") not in (None, "")
    )
    if present_horizons != expected_horizons:
        return True, (
            f"horizontes incompletos en RESULTADOS_GRID: "
            f"esperados={expected_horizons}, presentes={present_horizons}"
        )

    required_result = [
        "L_num",
        "L_trend",
        "L_risk",
        "L_tol",
        "Loss_h",
        "MAE",
        "RMSE",
        "Direccion_accuracy",
        "Deteccion_caidas",
    ]
    for row in result_rows:
        horizon = row.get("Horizonte_sem")
        missing = [field for field in required_result if row.get(field) is None]
        if missing:
            return True, f"faltan metricas en RESULTADOS_GRID para h={horizon}: {', '.join(missing)}"

    return False, f"{run_id} ya tiene metricas numericas persistidas"


def build_artifacts(
    run_id: str,
    run_dir: Path,
    artifact_rows: list[dict[str, Any]],
) -> list[dict[str, str]]:
    if artifact_rows:
        artifacts: list[dict[str, str]] = []
        for row in artifact_rows:
            relative_path = row.get("Ruta") or ""
            artifact_path = ROOT_DIR / relative_path if relative_path else run_dir
            artifacts.append(
                {
                    "artifact_type": str(row.get("Tipo") or ""),
                    "label": str(row.get("Nombre") or Path(artifact_path).name),
                    "path": str(artifact_path),
                    "notes": str(row.get("Notas") or ""),
                }
            )
        return artifacts

    derived_artifacts: list[dict[str, str]] = []
    for path in sorted(run_dir.rglob("*")):
        if not path.is_file():
            continue
        artifact_type, notes = classify_artifact(path)
        derived_artifacts.append(
            {
                "artifact_type": artifact_type,
                "label": path.name,
                "path": str(path),
                "notes": notes,
            }
        )
    return derived_artifacts


def classify_artifact(path: Path) -> tuple[str, str]:
    if path.name == "parametros_run.json":
        return "parametros", "Parametros del experimento."
    if path.name == "metricas_horizonte.json":
        return "metricas", "Resumen estructurado por horizonte."
    if path.name == "metadata_run.json":
        return "metadata", "Metadata base del run."
    if path.name == "resumen_modeling_horizontes.json":
        return "resumen", "Resumen de tamanos por horizonte."
    if path.name.startswith("predicciones_h") and path.suffix.lower() == ".csv":
        horizon = path.stem.replace("predicciones_h", "")
        return "predicciones", f"Predicciones registradas para horizonte {horizon}."
    if path.parent.name == "scripts" and path.suffix.lower() == ".py":
        return "script_snapshot", "Copia del script ejecutado para reproducibilidad."
    return "archivo", "Artefacto detectado durante backfill."


def build_log_payload(
    run_dir: Path,
    metadata: dict[str, Any],
    metrics_payload: dict[str, Any],
    summary_row: dict[str, Any] | None,
    result_rows: list[dict[str, Any]],
    artifact_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    horizon_results = metrics_payload["horizon_results"]
    first_horizon = horizon_results[0]
    first_result_row = result_rows[0] if result_rows else {}
    summary_row = summary_row or {}

    comentarios = (
        summary_row.get("Comentarios")
        or first_result_row.get("Comentarios")
        or "Backfill automatico desde artefactos existentes."
    )
    estado = first_result_row.get("Estado") or "corrido"

    return {
        "run_id": metadata["run_id"],
        "experiment_id": metadata["experiment_id"],
        "family": metadata["family"],
        "model": metadata["model"],
        "script_path": metadata["script_path"],
        "run_dir": metadata["run_dir"],
        "snapshot_path": metadata["snapshot_path"],
        "parametros_path": metadata.get("parametros_path") or None,
        "resultados_path": run_dir / "metricas_horizonte.json",
        "horizon_results": horizon_results,
        "artifacts": build_artifacts(metadata["run_id"], run_dir, artifact_rows),
        "target": first_horizon.get("target") or first_result_row.get("Target") or "",
        "variables_temporales": (
            first_horizon.get("variables_temporales")
            or first_result_row.get("Variables_temporales")
            or ""
        ),
        "variables_tematicas": (
            first_horizon.get("variables_tematicas")
            or first_result_row.get("Variables_tematicas")
            or ""
        ),
        "transformacion": (
            first_horizon.get("transformacion")
            or first_result_row.get("Transformacion")
            or ""
        ),
        "seleccion_variables": (
            first_horizon.get("seleccion_variables")
            or first_result_row.get("Seleccion_variables")
            or ""
        ),
        "validacion": first_horizon.get("validacion") or first_result_row.get("Validacion") or "",
        "dataset_periodo": (
            first_horizon.get("dataset_periodo")
            or first_result_row.get("Dataset_Periodo")
            or ""
        ),
        "notas_config": (
            first_horizon.get("notas_config")
            or first_result_row.get("Notas_config")
            or ""
        ),
        "estado": estado,
        "comentarios": comentarios,
        "l_coh": summary_row.get("L_coh"),
        "timestamp_run": metrics_payload.get("created_at") or metadata.get("created_at"),
    }


def load_run_payload(run_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    metadata = json.loads((run_dir / "metadata_run.json").read_text(encoding="utf-8"))
    metrics_payload = json.loads((run_dir / "metricas_horizonte.json").read_text(encoding="utf-8"))
    return metadata, metrics_payload


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve()
    runs_dir = args.runs_dir.expanduser().resolve()

    tracker = RadarExperimentTracker(workbook_path=workbook_path, runs_dir=runs_dir)
    workbook_state = load_workbook_state(workbook_path)
    candidate_run_dirs = iter_run_dirs(runs_dir)

    selected_run_ids = {run_id.strip() for run_id in args.run_id if run_id.strip()}
    repaired = 0
    skipped = 0

    for run_dir in candidate_run_dirs:
        metadata, metrics_payload = load_run_payload(run_dir)
        run_id = str(metadata["run_id"]).strip()

        if selected_run_ids and run_id not in selected_run_ids:
            continue

        summary_row = workbook_state["summary_by_run"].get(run_id)
        result_rows = workbook_state["results_by_run"].get(run_id, [])
        artifact_rows = workbook_state["artifacts_by_run"].get(run_id, [])

        should_backfill, reason = needs_backfill(
            run_id=run_id,
            horizon_results=metrics_payload["horizon_results"],
            summary_row=summary_row,
            result_rows=result_rows,
        )

        if not args.all and not should_backfill:
            print(f"SKIP {run_id}: {reason}")
            skipped += 1
            continue

        payload = build_log_payload(
            run_dir=run_dir,
            metadata=metadata,
            metrics_payload=metrics_payload,
            summary_row=summary_row,
            result_rows=result_rows,
            artifact_rows=artifact_rows,
        )

        if args.dry_run:
            action = "REPAIR" if should_backfill else "REFRESH"
            print(f"{action} {run_id}: {reason}")
            repaired += 1
            continue

        tracker.log_run(**payload)
        action = "REPAIRED" if should_backfill else "REFRESHED"
        print(f"{action} {run_id}: {reason}")
        repaired += 1

    print(f"Total runs procesados: {repaired + skipped}")
    print(f"Runs escritos: {repaired}")
    print(f"Runs omitidos: {skipped}")


if __name__ == "__main__":
    main()
