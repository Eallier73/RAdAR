#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from config import DATE_COLUMN


RUN_DIR_PATTERN = re.compile(
    r"^(?P<run_id>.+?)_(?P<timestamp>\d{8}_\d{6})(?P<aborted>_aborted)?$"
)
RUN_ID_PATTERN = re.compile(r"\b(?:E|C)\d+_v\d+(?:_clean)?\b")

CORE_FILES = (
    "metadata_run.json",
    "parametros_run.json",
    "metricas_horizonte.json",
    "resumen_modeling_horizontes.json",
)

COMMON_PARAM_EXCLUDE_KEYS = {
    "dataset_path",
    "feature_columns",
    "feature_mode",
    "horizons",
    "initial_train_size",
    "lags",
    "model",
    "model_params",
    "sheet_name",
    "target_mode",
    "task_type",
    "transform_mode",
}

MODEL_HYPERPARAM_KEYS: tuple[tuple[tuple[str, ...], tuple[str, ...]], ...] = (
    (
        ("ridge",),
        (
            "alpha",
            "alpha_grid",
            "alphas",
            "inner_splits",
            "tuning_metric",
            "tuning_strategy",
            "uses_scaling",
        ),
    ),
    (
        ("bayesian",),
        (
            "alpha_1",
            "alpha_2",
            "alpha_init",
            "compute_score",
            "copy_x",
            "fit_intercept",
            "lambda_1",
            "lambda_2",
            "lambda_init",
            "max_iter",
            "tol",
            "verbose",
            "tuning_strategy",
            "uses_scaling",
        ),
    ),
    (
        ("huber",),
        (
            "alpha",
            "epsilon",
            "max_iter",
            "tol",
            "inner_splits",
            "tuning_metric",
            "tuning_strategy",
            "param_grid",
            "winsor_lower_quantile",
            "winsor_upper_quantile",
            "uses_scaling",
        ),
    ),
    (
        ("random_forest", "extra_trees"),
        (
            "tree_model",
            "bootstrap",
            "n_estimators",
            "max_depth",
            "max_features",
            "min_samples_leaf",
            "min_samples_split",
            "class_weight",
            "random_state",
            "uses_scaling",
        ),
    ),
    (
        ("xgboost",),
        (
            "n_estimators",
            "learning_rate",
            "max_depth",
            "subsample",
            "colsample_bytree",
            "min_child_weight",
            "reg_alpha",
            "reg_lambda",
            "objective",
            "eval_metric",
            "random_state",
            "tuning_metric",
            "tuning_strategy",
            "param_grid",
            "uses_scaling",
        ),
    ),
    (
        ("catboost",),
        (
            "iterations",
            "depth",
            "learning_rate",
            "l2_leaf_reg",
            "subsample",
            "loss_function",
            "eval_metric",
            "random_seed",
            "random_state",
            "inner_splits",
            "tuning_metric",
            "tuning_strategy",
            "param_grid",
            "uses_categorical_features",
            "uses_scaling",
        ),
    ),
    (
        ("sarimax", "arimax"),
        (
            "order",
            "seasonal_order",
            "trend",
            "maxiter",
            "measurement_error",
            "enforce_stationarity",
            "enforce_invertibility",
            "tuning_strategy",
            "uses_scaling",
            "uses_categorical_features",
        ),
    ),
    (
        ("prophet",),
        (
            "changepoint_prior_scale",
            "seasonality_prior_scale",
            "seasonality_mode",
            "weekly_seasonality",
            "yearly_seasonality",
            "daily_seasonality",
        ),
    ),
    (
        ("lightgbm",),
        (
            "n_estimators",
            "num_leaves",
            "max_depth",
            "learning_rate",
            "min_child_samples",
            "subsample",
            "colsample_bytree",
            "reg_alpha",
            "reg_lambda",
            "objective",
            "random_state",
        ),
    ),
)

EXPLAINABILITY_COEF_PATTERNS = ("coef", "coefficient", "coeficient")
EXPLAINABILITY_IMPORTANCE_PATTERNS = ("importance", "importancia", "feature_import")
EXPLAINABILITY_SHAP_PATTERNS = ("shap",)

HORIZON_METRIC_SPECS = (
    {"metric": "mae", "column_suffix": "mae", "sense": "menor_es_mejor", "ascending": True},
    {"metric": "rmse", "column_suffix": "rmse", "sense": "menor_es_mejor", "ascending": True},
    {
        "metric": "direction_accuracy",
        "column_suffix": "direction_accuracy",
        "sense": "mayor_es_mejor",
        "ascending": False,
    },
    {
        "metric": "deteccion_caidas",
        "column_suffix": "deteccion_caidas",
        "sense": "mayor_es_mejor",
        "ascending": False,
    },
    {"metric": "loss_h", "column_suffix": "loss", "sense": "menor_es_mejor", "ascending": True},
)
RANKING_TIEBREAKER_TEXT = "L_total_Radar asc, run_id asc"
PREDICTION_DATE_CANDIDATES = (DATE_COLUMN, "fecha", "fecha_semana", "week_start")
PREDICTION_Y_TRUE_CANDIDATES = ("y_true", "y_true_model", "actual", "y_actual")
PREDICTION_Y_PRED_CANDIDATES = ("y_pred", "y_pred_model", "pred", "prediction")

E9_FINAL_CANDIDATES_BY_HORIZON: dict[int, list[str]] = {
    1: ["E1_v5_clean", "E5_v4_clean", "E3_v2_clean", "E2_v3_clean"],
    2: ["E1_v5_clean", "E5_v4_clean", "E2_v3_clean", "E7_v3_clean"],
    3: ["E1_v5_clean", "E5_v4_clean", "E3_v2_clean", "E7_v3_clean"],
    4: ["E1_v5_clean", "E5_v4_clean", "E3_v2_clean", "E2_v3_clean"],
}

E9_RESERVE_CANDIDATES_BY_HORIZON: dict[int, list[str]] = {
    1: ["E1_v4_clean"],
    2: ["E1_v4_clean", "E8_v2_clean"],
    3: ["E1_v4_clean", "E1_v2_clean"],
    4: ["E1_v4_clean", "E4_v1_clean"],
}

E9_CONSTRUCT_DEFINITIONS: dict[str, str] = {
    "campeon global de familia": "Mejor run vigente dentro de su familia experimental y ancla natural para el ensemble.",
    "referencia parsimoniosa": "Run conservado por simplicidad, estabilidad o interpretabilidad, aunque no sea el campeon global.",
    "especialista de horizonte": "Run con senal particularmente util en uno o mas horizontes concretos y con diversidad util frente al nucleo.",
    "candidato reserva": "Run documentado como alternativa metodologica, pero fuera del bloque principal por tamano o estabilidad.",
    "run dominado": "Run superado por otro muy cercano de su misma familia sin aportar diversidad real adicional.",
    "run no preferible para ensemble principal": "Run excluido del ensemble principal por ruido, redundancia, inestabilidad o comparabilidad deficiente.",
}

FAMILY_STATUS_VIGENTE_ROWS: tuple[dict[str, str], ...] = (
    {
        "family": "E1",
        "estado_vigente": "cerrada",
        "mejor_run": "E1_v5_clean",
        "rol_funcional": "referente numerico puro",
        "decision_metodologica": (
            "Mejor baseline lineal y mejor referente actual para pronostico numerico puro del porcentaje."
        ),
        "siguiente_movimiento": "sin expansion prevista",
    },
    {
        "family": "E2",
        "estado_vigente": "cerrada",
        "mejor_run": "E2_v3_clean",
        "rol_funcional": "referencia robusta",
        "decision_metodologica": "Familia robusta util como contraste historico, pero sin competitividad frente a Ridge.",
        "siguiente_movimiento": "sin expansion prevista",
    },
    {
        "family": "E3",
        "estado_vigente": "cerrada en rama base",
        "mejor_run": "E3_v2_clean",
        "rol_funcional": "referencia no lineal base",
        "decision_metodologica": "Sigue siendo la referencia bagging util, pero ya desplazada por CatBoost.",
        "siguiente_movimiento": "mantener como candidato de diversidad",
    },
    {
        "family": "E4",
        "estado_vigente": "cerrada",
        "mejor_run": "E4_v1_clean",
        "rol_funcional": "boosting historico",
        "decision_metodologica": "No desplazo a E3 ni a Ridge; queda como referencia secundaria.",
        "siguiente_movimiento": "sin expansion prevista",
    },
    {
        "family": "E5",
        "estado_vigente": "abierta madura",
        "mejor_run": "E5_v4_clean",
        "rol_funcional": "mejor no lineal tabular",
        "decision_metodologica": "Campeon no lineal tabular vigente; sigue siendo referencia competitiva central.",
        "siguiente_movimiento": "mantener como candidato fuerte para arquitecturas compuestas",
    },
    {
        "family": "E6",
        "estado_vigente": "debilitada",
        "mejor_run": "E6_v1_clean",
        "rol_funcional": "referencia temporal estructurada debil",
        "decision_metodologica": "ARIMAX no fue competitivo y no justifica continuidad inmediata.",
        "siguiente_movimiento": "solo reserva conceptual para hibridos",
    },
    {
        "family": "E7",
        "estado_vigente": "intermedia",
        "mejor_run": "E7_v3_clean",
        "rol_funcional": "referencia temporal con changepoints",
        "decision_metodologica": "Prophet supero a E6 pero no entro al bloque contendiente principal.",
        "siguiente_movimiento": "mantener como referencia temporal secundaria",
    },
    {
        "family": "E8",
        "estado_vigente": "intermedia",
        "mejor_run": "E8_v2_clean",
        "rol_funcional": "hibrido residual auditable",
        "decision_metodologica": "Mostro validez metodologica, pero no agrego mejora suficiente frente a su base.",
        "siguiente_movimiento": "sin expansion amplia; solo hipotesis residual muy acotada si reaparece",
    },
    {
        "family": "E9",
        "estado_vigente": "pausada util",
        "mejor_run": "E9_v2_clean",
        "rol_funcional": "referente riesgo-direccion-caidas",
        "decision_metodologica": (
            "No reemplaza a E1_v5_clean; aporta valor operativo en riesgo, direccion y deteccion de caidas. "
            "Queda util pero no definitiva."
        ),
        "siguiente_movimiento": "pausa metodologica; posible reactivacion futura",
    },
    {
        "family": "E10",
        "estado_vigente": "cerrada para promocion",
        "mejor_run": "E10_v1_clean",
        "rol_funcional": "antecedente metodologico de seleccion contextual",
        "decision_metodologica": (
            "E10 ya fue probado con E10_v1_clean. La corrida fue metodologicamente limpia y trazable, "
            "pero no supero al selector fijo, a E1_v5_clean ni a E9_v2_clean; la hipotesis de promocion "
            "no quedo confirmada y la rama se cierra para promocion bajo su formulacion actual."
        ),
        "siguiente_movimiento": "conservar como antecedente; no reabrir sin reformulacion fuerte",
    },
    {
        "family": "E11",
        "estado_vigente": "evaluada sin promocion",
        "mejor_run": "E11_v2_clean",
        "rol_funcional": "apertura dual numerica + categorica",
        "decision_metodologica": (
            "E11 ya fue abierta con tres variantes controladas. E11_v1_clean y E11_v2_clean preservan "
            "exactamente el bloque numerico de E1_v5_clean; E11_v2_clean agrega una capa binaria de caidas "
            "con senal moderada; E11_v3_clean no mejora el global. Ninguna variante desplaza a E1_v5_clean "
            "ni a E9_v2_clean y la familia queda abierta solo como evidencia dual no promocionable todavia."
        ),
        "siguiente_movimiento": "sin promocion; congelar apertura y solo reabrir con hipotesis dual realmente nueva",
    },
    {
        "family": "E12",
        "estado_vigente": "evaluada sin promocion",
        "mejor_run": "E12_v3_clean",
        "rol_funcional": "representacion enriquecida por horizonte",
        "decision_metodologica": (
            "E12 abrio la pregunta de si parte de la ventaja operativa de E9 podia absorberse mediante "
            "representacion enriquecida dentro de un Ridge simple y trazable. Ninguna de sus tres variantes "
            "mejora de forma defendible a E1_v5_clean ni a E9_v2_clean; la mejor lectura es un hallazgo "
            "parcial de que el bloque de desacuerdo aporta mas que el bloque de regimen, pero sin promocion."
        ),
        "siguiente_movimiento": "sin expansion inmediata; reabrir solo con hipotesis de H1 mas precisa y defendible",
    },
    {
        "family": "C1",
        "estado_vigente": "evaluada y pausada tempranamente",
        "mejor_run": "",
        "rol_funcional": "clasificacion bagging base",
        "decision_metodologica": (
            "C1 si se ejecuto con tres corridas canonicas, pero las tres colapsaron a clase unica y no dejaron "
            "senal discriminativa util. La rama queda abierta solo como antecedente, no como linea prioritaria."
        ),
        "siguiente_movimiento": "sin expansion inmediata; revisar target clasificacion antes de reactivar",
    },
    {
        "family": "C2",
        "estado_vigente": "evaluada y pausada tempranamente",
        "mejor_run": "C2_v1_clean",
        "rol_funcional": "clasificacion boosting",
        "decision_metodologica": (
            "C2 si se ejecuto con tres corridas canonicas de XGBoost, pero las tres reprodujeron el mismo "
            "colapso a clase unica observado en C1 y no agregaron senal discriminativa util."
        ),
        "siguiente_movimiento": "sin expansion inmediata; revisar target clasificacion antes de reactivar",
    },
    {
        "family": "C3",
        "estado_vigente": "infraestructura preparada",
        "mejor_run": "",
        "rol_funcional": "clasificacion boosting",
        "decision_metodologica": (
            "La familia tiene runner y prompt canonicos, pero no corridas ejecutadas todavia."
        ),
        "siguiente_movimiento": "no prioritaria; ejecutar solo si se redefine la agenda de clasificacion",
    },
    {
        "family": "C4",
        "estado_vigente": "infraestructura preparada",
        "mejor_run": "",
        "rol_funcional": "clasificacion boosting",
        "decision_metodologica": (
            "La familia tiene runner y prompt canonicos, pero no corridas ejecutadas todavia."
        ),
        "siguiente_movimiento": "no prioritaria; ejecutar solo si se redefine la agenda de clasificacion",
    },
)

FAMILY_CANONICAL_CONSTRUCTS: dict[str, str] = {
    "E1": "baseline lineal numerico principal",
    "E2": "familia robusta de contraste historico",
    "E3": "referencia bagging de diversidad",
    "E4": "boosting historico cerrado",
    "E5": "campeon no lineal tabular vigente",
    "E6": "referencia temporal debilitada",
    "E7": "referencia temporal secundaria",
    "E8": "hibrido residual auditable",
    "E9": "rama operativa de riesgo-direccion-caidas",
    "E10": "selector contextual probado y no promovido",
    "E11": "apertura dual controlada no promocionable",
    "E12": "familia de representacion enriquecida evaluada",
    "C1": "clasificacion base evaluada",
    "C2": "clasificacion boosting evaluada",
    "C3": "clasificacion boosting preparada",
    "C4": "clasificacion boosting preparada",
}

OPERATIONAL_BENCHMARK_ROWS: tuple[dict[str, str], ...] = (
    {
        "benchmark_id": "benchmark_numerico_puro_vigente",
        "plano_funcional": "numerico",
        "run_id": "E1_v5_clean",
        "estado_operativo": "vigente",
        "rol_operativo": "benchmark numerico puro",
        "justificacion": (
            "Mejor referente actual para pronostico numerico puro del Radar; permanece como benchmark "
            "operativo controlado mientras no exista una arquitectura futura que lo supere de forma defendible."
        ),
    },
    {
        "benchmark_id": "benchmark_operativo_riesgo_vigente",
        "plano_funcional": "riesgo-direccion-caidas",
        "run_id": "E9_v2_clean",
        "estado_operativo": "vigente",
        "rol_operativo": "benchmark operativo de riesgo-direccion-caidas",
        "justificacion": (
            "Mejor referente actual de direccion y deteccion de caidas; permanece como benchmark operativo "
            "controlado sin reemplazar al benchmark numerico puro."
        ),
    },
)

RADAR_LAYER_ROWS: tuple[dict[str, str], ...] = (
    {
        "capa": "experimental",
        "estado": "activa",
        "alcance": "familias E1-E10 y ramas C1-C4 bajo disciplina temporal completa",
        "objetivo": "seguir evaluando hipotesis nuevas con comparabilidad fuerte y trazabilidad por run",
        "restriccion": "sin leakage, sin tuning con futuro, sin promotion por intuicion",
    },
    {
        "capa": "operativa_controlada",
        "estado": "activa",
        "alcance": "ejecucion estable de benchmarks vigentes E1_v5_clean y E9_v2_clean",
        "objetivo": "comparacion operativa controlada sin reescribir la logica canonica de los runners",
        "restriccion": "no equivale a produccion final unificada; conserva dualidad funcional vigente",
    },
    {
        "capa": "futura_dual",
        "estado": "evaluada sin promocion",
        "alcance": "familia E11 dual numerica + categorica",
        "objetivo": "explorar una arquitectura compuesta que en el futuro pueda reemplazar la salida dual vigente",
        "restriccion": "no promocionar ni vender como resuelta mientras no supere de forma defendible a E1_v5_clean y E9_v2_clean",
    },
)

DUAL_FUNCTIONAL_METRIC_ROWS: tuple[dict[str, Any], ...] = (
    {"horizonte_sem": 1, "medida": "MAE", "E1_v5_clean": 0.121907, "E9_v2_clean": 0.123894, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 1, "medida": "RMSE", "E1_v5_clean": 0.160550, "E9_v2_clean": 0.166878, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 1, "medida": "Direction accuracy", "E1_v5_clean": 0.846154, "E9_v2_clean": 0.857143, "ganador": "E9_v2_clean"},
    {"horizonte_sem": 1, "medida": "Deteccion de caidas", "E1_v5_clean": 0.833333, "E9_v2_clean": 1.000000, "ganador": "E9_v2_clean"},
    {"horizonte_sem": 1, "medida": "Loss_h", "E1_v5_clean": 0.064860, "E9_v2_clean": 0.051557, "ganador": "E9_v2_clean"},
    {"horizonte_sem": 2, "medida": "MAE", "E1_v5_clean": 0.118495, "E9_v2_clean": 0.136777, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 2, "medida": "RMSE", "E1_v5_clean": 0.153470, "E9_v2_clean": 0.178276, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 2, "medida": "Direction accuracy", "E1_v5_clean": 0.760000, "E9_v2_clean": 0.720000, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 2, "medida": "Deteccion de caidas", "E1_v5_clean": 0.700000, "E9_v2_clean": 0.833333, "ganador": "E9_v2_clean"},
    {"horizonte_sem": 2, "medida": "Loss_h", "E1_v5_clean": 0.071501, "E9_v2_clean": 0.063405, "ganador": "E9_v2_clean"},
    {"horizonte_sem": 3, "medida": "MAE", "E1_v5_clean": 0.092137, "E9_v2_clean": 0.129780, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 3, "medida": "RMSE", "E1_v5_clean": 0.123798, "E9_v2_clean": 0.172374, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 3, "medida": "Direction accuracy", "E1_v5_clean": 0.666667, "E9_v2_clean": 0.708333, "ganador": "E9_v2_clean"},
    {"horizonte_sem": 3, "medida": "Deteccion de caidas", "E1_v5_clean": 0.818182, "E9_v2_clean": 1.000000, "ganador": "E9_v2_clean"},
    {"horizonte_sem": 3, "medida": "Loss_h", "E1_v5_clean": 0.068181, "E9_v2_clean": 0.073641, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 4, "medida": "MAE", "E1_v5_clean": 0.077511, "E9_v2_clean": 0.115178, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 4, "medida": "RMSE", "E1_v5_clean": 0.108231, "E9_v2_clean": 0.157989, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 4, "medida": "Direction accuracy", "E1_v5_clean": 0.702703, "E9_v2_clean": 0.658824, "ganador": "E1_v5_clean"},
    {"horizonte_sem": 4, "medida": "Deteccion de caidas", "E1_v5_clean": 0.777778, "E9_v2_clean": 0.833333, "ganador": "E9_v2_clean"},
    {"horizonte_sem": 4, "medida": "Loss_h", "E1_v5_clean": 0.038900, "E9_v2_clean": 0.038906, "ganador": "E1_v5_clean"},
)

DUAL_POLICY_SUMMARY_ROWS: tuple[dict[str, str], ...] = (
    {
        "capa": "salida_numerica_principal",
        "politica_operativa": "Siempre E1_v5_clean",
        "detalle": "Forecast numerico oficial del sistema",
    },
    {
        "capa": "deteccion_de_caidas",
        "politica_operativa": "Siempre E9_v2_clean",
        "detalle": "Alerta oficial de caida por horizonte",
    },
    {
        "capa": "direction_accuracy_H1",
        "politica_operativa": "E9_v2_clean",
        "detalle": "Politica fija direccional por horizonte",
    },
    {
        "capa": "direction_accuracy_H2",
        "politica_operativa": "E1_v5_clean",
        "detalle": "Politica fija direccional por horizonte",
    },
    {
        "capa": "direction_accuracy_H3",
        "politica_operativa": "E9_v2_clean",
        "detalle": "Politica fija direccional por horizonte",
    },
    {
        "capa": "direction_accuracy_H4",
        "politica_operativa": "E1_v5_clean",
        "detalle": "Politica fija direccional por horizonte",
    },
)

RUN_CANONICAL_CONSTRUCTS: dict[str, dict[str, str]] = {
    "C1_v1_clean": {
        "constructo": "baseline prudente de clasificacion",
        "justificacion": "Primera apertura real de clasificacion; sirve como antecedente metodologico, pero no dejo senal util por colapso del target.",
    },
    "C1_v2_clean": {
        "constructo": "control parsimonioso de clasificacion",
        "justificacion": "Variante mas contenida dentro de C1; confirma que el problema era el target y no solo la complejidad del clasificador.",
    },
    "C1_v3_clean": {
        "constructo": "control flexible de clasificacion",
        "justificacion": "Variante mas flexible de C1; tampoco rompe el colapso a clase unica y cierra la rama temprana.",
    },
    "C2_v1_clean": {
        "constructo": "baseline prudente de clasificacion boosting",
        "justificacion": "Primera apertura real de XGBoost multiclase; replica el mismo colapso a clase unica observado en C1.",
    },
    "C2_v2_clean": {
        "constructo": "control parsimonioso de boosting de clasificacion",
        "justificacion": "Variante mas regularizada dentro de C2; confirma que el problema sigue siendo el target y no la capacidad del booster.",
    },
    "C2_v3_clean": {
        "constructo": "control flexible de boosting de clasificacion",
        "justificacion": "Variante algo mas flexible de C2; tampoco rompe el colapso y cierra la familia tempranamente.",
    },
    "E10_v1_clean": {
        "constructo": "selector contextual duro cerrado para promocion",
        "justificacion": "Primera corrida canonica de gating contextual; metodologicamente limpia, pero no competitiva y cerrada para promocion bajo la formulacion actual.",
    },
    "E11_v1_clean": {
        "constructo": "apertura dual paralela",
        "justificacion": "Primera apertura dual controlada; conserva exactamente el bloque numerico de E1_v5_clean, pero la capa ternaria de direccion sigue casi colapsada a estabilidad.",
    },
    "E11_v2_clean": {
        "constructo": "mejor apertura dual con detector de caidas",
        "justificacion": "Mantiene el benchmark numerico de E1_v5_clean y agrega una capa binaria de caidas con senal real, aunque todavia insuficiente para desplazar a E9_v2_clean.",
    },
    "E11_v3_clean": {
        "constructo": "control residual dual no ganador",
        "justificacion": "Prueba residual dual metodologicamente limpia; agrega complejidad y no mejora el global frente a E1_v5_clean.",
    },
    "E12_v1_clean": {
        "constructo": "baseline de representacion minima no ganador",
        "justificacion": (
            "Primera apertura de representacion enriquecida con bloque parsimonioso; no captura la mejora "
            "esperada en H1 y queda por detras de E1_v5_clean y E9_v2_clean."
        ),
    },
    "E12_v2_clean": {
        "constructo": "representacion ampliada de diversidad no ganadora",
        "justificacion": (
            "Amplia el bloque de representacion con mas diversidad entre familias base; agrega complejidad "
            "sin mejora defendible frente a los benchmarks vigentes."
        ),
    },
    "E12_v3_clean": {
        "constructo": "control de desacuerdo puro con hallazgo parcial",
        "justificacion": (
            "Aisla senales de desacuerdo entre bases sin bloque de regimen; es la mejor variante interna de "
            "E12, pero sigue sin superar a E1_v5_clean ni a E9_v2_clean."
        ),
    },
    "E1_1_v1_bayesian_base": {
        "constructo": "micro-rama tactica bayesiana",
        "justificacion": "Verificacion corta para probar margen lineal adicional; no altero el cierre principal de E1.",
    },
    "E1_v1": {
        "constructo": "baseline historico retrointegrado",
        "justificacion": "Run historico recuperado para trazabilidad y comparacion retrospectiva; no gobierna la lectura vigente.",
    },
    "E1_v2": {
        "constructo": "baseline historico documentado",
        "justificacion": "Version historica previa al cierre limpio; sirve como antecedente de la evolucion de Ridge.",
    },
    "E1_v2_clean": {
        "constructo": "baseline Ridge limpio historico",
        "justificacion": "Primera referencia limpia reconstruida de Ridge; importante como antecedente comparativo, pero superada por E1_v4 y E1_v5.",
    },
    "E1_v3_clean": {
        "constructo": "hipotesis delta descartada",
        "justificacion": "Control sobre target delta; su deterioro ayudo a cerrar la via de Ridge en nivel como direccion preferible.",
    },
    "E1_v4_clean": {
        "constructo": "referencia parsimoniosa vigente",
        "justificacion": "Run equilibrado y mas simple dentro de E1; conserva valor canonico aunque no sea el campeon global.",
    },
    "E1_v5_clean": {
        "constructo": "campeon numerico puro vigente",
        "justificacion": "Mejor referente actual para pronostico numerico puro del Radar.",
    },
    "E2_v1_clean": {
        "constructo": "baseline robusto limpio",
        "justificacion": "Primera apertura Huber bajo el marco limpio; fija el punto de partida de la familia robusta.",
    },
    "E2_v2_clean": {
        "constructo": "control de convergencia",
        "justificacion": "Sirve para descartar que el desempeno de Huber se explicara por optimizacion incompleta.",
    },
    "E2_v3_clean": {
        "constructo": "campeon interno de familia cerrada",
        "justificacion": "Mejor Huber interno; conserva valor historico como referencia robusta, pero no reabre la familia.",
    },
    "E3_v1_clean": {
        "constructo": "baseline no lineal bagging",
        "justificacion": "Primera apertura no lineal util de bagging; demuestra que la no linealidad si agrega senal.",
    },
    "E3_v2_clean": {
        "constructo": "campeon bagging vigente",
        "justificacion": "Mejor referencia bagging cerrada; sigue aportando diversidad real frente a lineales y boosting.",
    },
    "E3_v3_clean": {
        "constructo": "variante extra trees descartada",
        "justificacion": "Prueba de flexibilidad adicional que no justifico expansion dentro de E3.",
    },
    "E4_v1_clean": {
        "constructo": "campeon boosting historico",
        "justificacion": "Mejor run interno de XGBoost; se conserva como referencia secundaria, no como linea vigente.",
    },
    "E4_v2_clean": {
        "constructo": "variante parsimoniosa descartada",
        "justificacion": "Control corr vs all dentro de boosting; no mejoro al baseline de la familia.",
    },
    "E4_v3_clean": {
        "constructo": "variante delta descartada",
        "justificacion": "Control target delta dentro de boosting; deterioro fuerte que ayudo a cerrar E4.",
    },
    "E5_v1_clean": {
        "constructo": "baseline catboost",
        "justificacion": "Primera apertura de CatBoost bajo el marco limpio; punto de partida de la familia tabular dominante.",
    },
    "E5_v2_clean": {
        "constructo": "control corr_vs_all catboost",
        "justificacion": "Prueba focal del universo exogeno en CatBoost; quedo dominada por E5_v4.",
    },
    "E5_v3_clean": {
        "constructo": "control de profundidad catboost",
        "justificacion": "Ajuste acotado de complejidad en CatBoost; no mejoro al campeon interno.",
    },
    "E5_v4_clean": {
        "constructo": "campeon no lineal tabular vigente",
        "justificacion": "Mejor run no lineal tabular del proyecto y referencia competitiva central.",
    },
    "E5_v5_clean": {
        "constructo": "tuning focal dominado",
        "justificacion": "Segunda ola de tuning acotado; util para cierre interno, pero superada por E5_v4.",
    },
    "E6_v1_clean": {
        "constructo": "campeon interno debilitado",
        "justificacion": "Mejor ARIMAX interno, aunque claramente no competitivo; conserva valor solo como referencia temporal debil.",
    },
    "E6_v2_clean": {
        "constructo": "control amplitud exogena descartado",
        "justificacion": "Prueba de amplitud exogena en ARIMAX; empeoro y debilito aun mas a la familia.",
    },
    "E7_v1_clean": {
        "constructo": "baseline temporal parsimonioso",
        "justificacion": "Apertura limpia de Prophet con exogenas parsimoniosas; prueba que la familia mejora a E6.",
    },
    "E7_v2_clean": {
        "constructo": "variante all colapsada",
        "justificacion": "Control de amplitud exogena en Prophet; colapso numerico que refuerza la parsimonia.",
    },
    "E7_v3_clean": {
        "constructo": "campeon temporal secundario",
        "justificacion": "Mejor Prophet interno; aporta sesgo temporal estructurado sin entrar al bloque principal.",
    },
    "E8_v1_clean": {
        "constructo": "hibrido residual inicial",
        "justificacion": "Primera apertura residual limpia; valida la arquitectura, pero no la vuelve competitiva.",
    },
    "E8_v2_clean": {
        "constructo": "campeon hibrido residual",
        "justificacion": "Mejor residual interno y referencia de la familia hibrida, aunque no supera a sus mejores bases.",
    },
    "E8_v3_clean": {
        "constructo": "variante prophet residual colapsada",
        "justificacion": "Prueba temporal residual sobre CatBoost que no funciono y cerro esa direccion.",
    },
    "E9_smoke_tmp": {
        "constructo": "smoke tecnico no canonico",
        "justificacion": "Corrida tecnica de integracion para E9; no debe leerse como evidencia comparativa final.",
    },
    "E9_v1_clean": {
        "constructo": "baseline stacking clasico ridge",
        "justificacion": "Primera corrida operativa de stacking controlado; deja una mejora parcial localizada pero no gana el global.",
    },
    "E9_v2_clean": {
        "constructo": "campeon riesgo-direccion-caidas",
        "justificacion": "Mejor referente actual de riesgo, direccion y deteccion de caidas dentro del Radar.",
    },
}

E9_PRIORITY_RUN_NOTES: dict[str, dict[str, str]] = {
    "E1_v5_clean": {
        "constructo": "campeon global de familia",
        "justificacion": "Mejor run global vigente y ancla lineal principal del stacking controlado.",
    },
    "E1_v4_clean": {
        "constructo": "referencia parsimoniosa",
        "justificacion": "Ridge parsimonioso y equilibrado; se conserva como control auditable y reserva, pero no entra por redundancia alta con E1_v5_clean.",
    },
    "E1_v2_clean": {
        "constructo": "candidato reserva",
        "justificacion": "Reserva puntual para H3 por buen loss_h3, pero se excluye del bloque principal por duplicidad interna dentro de Ridge.",
    },
    "E2_v3_clean": {
        "constructo": "especialista de horizonte",
        "justificacion": "Huber robusto con senal real en H2 y H4; aporta diversidad util frente a Ridge y CatBoost.",
    },
    "E3_v2_clean": {
        "constructo": "campeon global de familia",
        "justificacion": "Mejor bagging vigente; mantiene diversidad real frente a lineales y boosting.",
    },
    "E4_v1_clean": {
        "constructo": "candidato reserva",
        "justificacion": "Campeon interno de XGBoost, pero queda fuera del ensemble principal por redundancia con CatBoost y bagging; se conserva solo como reserva tardia.",
    },
    "E5_v4_clean": {
        "constructo": "campeon global de familia",
        "justificacion": "Mejor no lineal tabular vigente; entra como eje principal de E9 en todos los horizontes.",
    },
    "E7_v3_clean": {
        "constructo": "especialista de horizonte",
        "justificacion": "Prophet parsimonioso mejora claramente a E6 y aporta sesgo temporal estructurado util en H2 y H3.",
    },
    "E8_v2_clean": {
        "constructo": "candidato reserva",
        "justificacion": "Mejor hibrido residual interno; se conserva solo como reserva puntual por H2, no como base principal.",
    },
    "E6_v1_clean": {
        "constructo": "run no preferible para ensemble principal",
        "justificacion": "ARIMAX quedo debilitado y no competitivo; no agrega una ventaja suficiente frente a E7 ni frente al bloque principal.",
    },
    "E6_v2_clean": {
        "constructo": "run no preferible para ensemble principal",
        "justificacion": "Variante all de ARIMAX empeoro a la base parsimoniosa y no debe entrar al ensemble principal.",
    },
}


@dataclass
class PredictionAuditResult:
    coverage_row: dict[str, Any]
    standardized_df: pd.DataFrame | None


@dataclass
class HyperparamsNormalizationResult:
    hyperparams_json: str
    hyperparams_hash: str
    hyperparams_firma: str
    hyperparams_resumen: str
    reconstruction_status: str
    reconstruction_note: str


@dataclass
class ExplainabilityAuditResult:
    tiene_coeficientes: bool
    tiene_importancias: bool
    tiene_shap_o_equivalente: bool
    explicabilidad_transversal_homogenea: bool
    artifact_explicabilidad_path: str
    observacion_explicabilidad: str


@dataclass
class RunDirectoryRecord:
    path: Path
    inferred_run_id: str
    timestamp: str | None
    is_aborted: bool
    metadata: dict[str, Any] | None
    parametros: dict[str, Any] | None
    metricas: dict[str, Any] | None
    resumen: list[dict[str, Any]] | None
    comparison_files: list[Path]
    prediction_files: list[Path]
    selected_feature_files: list[Path]
    issues: list[str]

    @property
    def effective_run_id(self) -> str:
        if self.metadata and self.metadata.get("run_id"):
            return str(self.metadata["run_id"])
        return self.inferred_run_id

    @property
    def model(self) -> str | None:
        if self.metadata and self.metadata.get("model"):
            return str(self.metadata["model"])
        if self.parametros and self.parametros.get("model"):
            return str(self.parametros["model"])
        return None

    @property
    def experiment_id(self) -> str | None:
        if self.metadata and self.metadata.get("experiment_id"):
            return str(self.metadata["experiment_id"])
        run_id = self.effective_run_id
        if "_" in run_id:
            return run_id.split("_", 1)[0]
        return None

    @property
    def artifact_status(self) -> str:
        core_present = sum(int((self.path / name).exists()) for name in CORE_FILES)
        if core_present == len(CORE_FILES) and len(self.prediction_files) >= 4:
            return "completo"
        if core_present >= 2 or self.prediction_files or self.selected_feature_files:
            return "parcial"
        return "inconsistente"

    @property
    def completeness_score(self) -> tuple[int, int, int, int, int, int]:
        return (
            sum(int((self.path / name).exists()) for name in CORE_FILES),
            len(self.prediction_files),
            len(self.comparison_files),
            len(self.selected_feature_files),
            int(self.metadata is not None),
            0 if self.is_aborted else 1,
        )


@dataclass
class MasterAuditBuildResult:
    csv_path: Path
    xlsx_path: Path
    json_path: Path
    markdown_path: Path
    master_runs_total: int
    inventory_directories_total: int
    e9_curated_xlsx_path: Path | None = None
    e9_markdown_path: Path | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Audita runs reales de Radar y construye una tabla maestra comparativa por horizonte.",
    )
    parser.add_argument(
        "--runs-dir",
        type=Path,
        default=Path("/home/emilio/Documentos/RAdAR/Experimentos/runs"),
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("/home/emilio/Documentos/RAdAR/Experimentos/grid_experimentos_radar.xlsx"),
    )
    parser.add_argument(
        "--prompts-dir",
        type=Path,
        default=Path("/home/emilio/Documentos/RAdAR/Experimentos/Prompts_Agente"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("/home/emilio/Documentos/RAdAR/Experimentos"),
    )
    return parser.parse_args()


def load_json_if_exists(path: Path) -> dict[str, Any] | list[dict[str, Any]] | None:
    if not path.exists():
        return None
    return json.loads(path.read_text())


def parse_run_directory_name(name: str) -> tuple[str, str | None, bool]:
    match = RUN_DIR_PATTERN.match(name)
    if not match:
        return name, None, False
    return match.group("run_id"), match.group("timestamp"), bool(match.group("aborted"))


def scan_run_directories(runs_dir: Path) -> list[RunDirectoryRecord]:
    records: list[RunDirectoryRecord] = []
    for path in sorted(runs_dir.iterdir()):
        if not path.is_dir():
            continue
        inferred_run_id, timestamp, is_aborted = parse_run_directory_name(path.name)
        comparison_files = sorted(path.glob("comparacion_vs_*.json"))
        prediction_files = sorted(path.glob("predicciones_h*.csv"))
        selected_feature_files = sorted(path.glob("features_seleccionadas_h*.csv"))
        metadata = load_json_if_exists(path / "metadata_run.json")
        parametros = load_json_if_exists(path / "parametros_run.json")
        metricas = load_json_if_exists(path / "metricas_horizonte.json")
        resumen = load_json_if_exists(path / "resumen_modeling_horizontes.json")
        issues: list[str] = []
        if metadata is None:
            issues.append("faltante_metadata")
        if parametros is None:
            issues.append("faltante_parametros")
        if metricas is None:
            issues.append("faltante_metricas")
        if resumen is None:
            issues.append("faltante_resumen")
        if len(prediction_files) < 4:
            issues.append("predicciones_incompletas")
        records.append(
            RunDirectoryRecord(
                path=path,
                inferred_run_id=inferred_run_id,
                timestamp=timestamp,
                is_aborted=is_aborted,
                metadata=metadata if isinstance(metadata, dict) else None,
                parametros=parametros if isinstance(parametros, dict) else None,
                metricas=metricas if isinstance(metricas, dict) else None,
                resumen=resumen if isinstance(resumen, list) else None,
                comparison_files=comparison_files,
                prediction_files=prediction_files,
                selected_feature_files=selected_feature_files,
                issues=issues,
            )
        )
    return records


def select_canonical_records(
    records: list[RunDirectoryRecord],
) -> tuple[dict[str, RunDirectoryRecord], dict[str, list[RunDirectoryRecord]]]:
    grouped: dict[str, list[RunDirectoryRecord]] = {}
    for record in records:
        grouped.setdefault(record.effective_run_id, []).append(record)

    canonical: dict[str, RunDirectoryRecord] = {}
    for run_id, candidates in grouped.items():
        canonical[run_id] = sorted(
            candidates,
            key=lambda item: (
                item.completeness_score,
                item.timestamp or "",
                item.path.name,
            ),
            reverse=True,
        )[0]
    return canonical, grouped


def normalize_feature_mode(value: Any) -> str | None:
    if value is None:
        return None
    raw = str(value).strip().lower()
    mapping = {
        "todas": "all",
        "todas_las_variables": "all",
        "all": "all",
        "corr": "corr",
        "lasso": "lasso",
    }
    return mapping.get(raw, raw or None)


def normalize_transform_mode(value: Any) -> str | None:
    if value is None:
        return None
    raw = str(value).strip().lower()
    mapping = {
        "estandarizada": "standard",
        "standard": "standard",
        "standard_scaler": "standard",
        "robust": "robust",
        "robust_scaler": "robust",
        "winsor": "winsor",
        "sin_escalar": "none",
        "": "none",
        "none": "none",
    }
    return mapping.get(raw, raw or None)


def normalize_family(run_id: str, family: Any, model: Any) -> str | None:
    raw_family = str(family).strip().lower() if family is not None else None
    model_name = str(model).strip().lower() if model is not None else ""
    if run_id.startswith("E1_"):
        return "lineal_regularizado"
    if run_id.startswith("E2_"):
        return "robusto"
    if run_id.startswith("E3_"):
        return "arboles_boosting"
    if run_id.startswith("C1_"):
        return "clasificacion_random_forest"
    if run_id.startswith("C2_"):
        return "clasificacion_xgboost"
    if run_id.startswith("C3_"):
        return "clasificacion_catboost"
    if run_id.startswith("C4_"):
        return "clasificacion_lightgbm"
    if raw_family:
        if "lineal" in raw_family and "ridge" in model_name:
            return "lineal_regularizado"
        return raw_family
    return raw_family


def maybe_cell_value(ws, headers: dict[str, int], row: int, header: str) -> Any:
    if header not in headers:
        return None
    return ws.cell(row, headers[header]).value


def parse_feature_count_prom(
    resumen: list[dict[str, Any]] | None,
    metricas: dict[str, Any] | None,
) -> float | None:
    if resumen:
        values = [
            float(item["selected_feature_count_avg"])
            for item in resumen
            if item.get("selected_feature_count_avg") is not None
        ]
        if values:
            return float(sum(values) / len(values))
    if metricas:
        horizons = metricas.get("horizon_results", [])
        counts = []
        for item in horizons:
            notas = str(item.get("notas_config", ""))
            match = re.search(r"(\d+(?:\.\d+)?)\s+features", notas)
            if match:
                counts.append(float(match.group(1)))
        if counts:
            return float(sum(counts) / len(counts))
    return None


def build_workbook_lookup(workbook_path: Path) -> tuple[dict[str, dict[str, Any]], dict[str, dict[int, dict[str, Any]]]]:
    workbook = load_workbook(workbook_path, data_only=True)
    summary_lookup: dict[str, dict[str, Any]] = {}
    ws = workbook["RUN_SUMMARY"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for row in range(2, ws.max_row + 1):
        run_id = ws.cell(row, headers["Run_ID"]).value
        if not run_id:
            continue
        summary_lookup[str(run_id)] = {
            "Task_type": maybe_cell_value(ws, headers, row, "Task_type"),
            "L_total_Radar": ws.cell(row, headers["L_total_Radar"]).value,
            "L_total_Clasificacion": maybe_cell_value(ws, headers, row, "L_total_Clasificacion"),
            "Avg_MAE": ws.cell(row, headers["Avg_MAE"]).value,
            "Avg_RMSE": ws.cell(row, headers["Avg_RMSE"]).value,
            "Dir_acc_prom": ws.cell(row, headers["Dir_acc_prom"]).value,
            "Det_caidas_prom": ws.cell(row, headers["Det_caidas_prom"]).value,
            "Avg_Accuracy_5clases": maybe_cell_value(ws, headers, row, "Avg_Accuracy_5clases"),
            "Avg_Balanced_accuracy_5clases": maybe_cell_value(ws, headers, row, "Avg_Balanced_accuracy_5clases"),
            "Avg_Macro_f1_5clases": maybe_cell_value(ws, headers, row, "Avg_Macro_f1_5clases"),
            "Avg_Weighted_f1_5clases": maybe_cell_value(ws, headers, row, "Avg_Weighted_f1_5clases"),
            "Avg_Recall_baja_fuerte": maybe_cell_value(ws, headers, row, "Avg_Recall_baja_fuerte"),
            "Avg_Recall_baja_total": maybe_cell_value(ws, headers, row, "Avg_Recall_baja_total"),
            "Avg_Precision_baja_fuerte": maybe_cell_value(ws, headers, row, "Avg_Precision_baja_fuerte"),
            "Avg_Recall_sube_fuerte": maybe_cell_value(ws, headers, row, "Avg_Recall_sube_fuerte"),
            "Avg_Recall_sube_total": maybe_cell_value(ws, headers, row, "Avg_Recall_sube_total"),
            "Avg_Accuracy_3clases": maybe_cell_value(ws, headers, row, "Avg_Accuracy_3clases"),
            "Avg_Macro_f1_3clases": maybe_cell_value(ws, headers, row, "Avg_Macro_f1_3clases"),
            "Avg_Balanced_accuracy_3clases": maybe_cell_value(ws, headers, row, "Avg_Balanced_accuracy_3clases"),
            "Comentarios": ws.cell(row, headers["Comentarios"]).value,
        }

    results_lookup: dict[str, dict[int, dict[str, Any]]] = {}
    ws = workbook["RESULTADOS_GRID"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for row in range(2, ws.max_row + 1):
        run_id = ws.cell(row, headers["Run_ID"]).value
        if not run_id:
            continue
        horizon = ws.cell(row, headers["Horizonte_sem"]).value
        if horizon is None:
            continue
        results_lookup.setdefault(str(run_id), {})[int(horizon)] = {
            "Task_type": maybe_cell_value(ws, headers, row, "Task_type"),
            "Loss_h": ws.cell(row, headers["Loss_h"]).value,
            "MAE": ws.cell(row, headers["MAE"]).value,
            "RMSE": ws.cell(row, headers["RMSE"]).value,
            "Direccion_accuracy": ws.cell(row, headers["Direccion_accuracy"]).value,
            "Deteccion_caidas": ws.cell(row, headers["Deteccion_caidas"]).value,
            "Accuracy_5clases": maybe_cell_value(ws, headers, row, "Accuracy_5clases"),
            "Balanced_accuracy_5clases": maybe_cell_value(ws, headers, row, "Balanced_accuracy_5clases"),
            "Macro_f1_5clases": maybe_cell_value(ws, headers, row, "Macro_f1_5clases"),
            "Weighted_f1_5clases": maybe_cell_value(ws, headers, row, "Weighted_f1_5clases"),
            "Recall_baja_fuerte": maybe_cell_value(ws, headers, row, "Recall_baja_fuerte"),
            "Recall_baja_total": maybe_cell_value(ws, headers, row, "Recall_baja_total"),
            "Precision_baja_fuerte": maybe_cell_value(ws, headers, row, "Precision_baja_fuerte"),
            "Recall_sube_fuerte": maybe_cell_value(ws, headers, row, "Recall_sube_fuerte"),
            "Recall_sube_total": maybe_cell_value(ws, headers, row, "Recall_sube_total"),
            "Accuracy_3clases": maybe_cell_value(ws, headers, row, "Accuracy_3clases"),
            "Macro_f1_3clases": maybe_cell_value(ws, headers, row, "Macro_f1_3clases"),
            "Balanced_accuracy_3clases": maybe_cell_value(ws, headers, row, "Balanced_accuracy_3clases"),
            "Loss_clasificacion_h": maybe_cell_value(ws, headers, row, "Loss_clasificacion_h"),
        }
    return summary_lookup, results_lookup


def horizon_metrics_map(record: RunDirectoryRecord) -> dict[int, dict[str, Any]]:
    if not record.metricas:
        return {}
    result: dict[int, dict[str, Any]] = {}
    for item in record.metricas.get("horizon_results", []):
        horizon = int(item["horizonte_sem"])
        result[horizon] = item
    return result


def extract_target_mode(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> str | None:
    if record.parametros and record.parametros.get("target_mode"):
        return str(record.parametros["target_mode"])
    for item in horizon_map.values():
        if item.get("target"):
            return str(item["target"])
    return None


def extract_task_type(record: RunDirectoryRecord, summary: dict[str, Any]) -> str:
    if record.parametros and record.parametros.get("task_type"):
        return str(record.parametros["task_type"])
    if record.metadata and record.metadata.get("task_type"):
        return str(record.metadata["task_type"])
    if summary.get("Task_type"):
        return str(summary["Task_type"])
    run_id = record.effective_run_id
    if run_id.startswith("C"):
        return "clasificacion"
    return "regresion"


def extract_feature_mode(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> str | None:
    if record.parametros and record.parametros.get("feature_mode") is not None:
        return normalize_feature_mode(record.parametros["feature_mode"])
    for item in horizon_map.values():
        if item.get("seleccion_variables") is not None:
            return normalize_feature_mode(item["seleccion_variables"])
    return None


def extract_transform_mode(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> str | None:
    if record.parametros and "transform_mode" in record.parametros:
        return normalize_transform_mode(record.parametros.get("transform_mode"))
    for item in horizon_map.values():
        if item.get("transformacion") is not None:
            return normalize_transform_mode(item["transformacion"])
    return None


def extract_lags(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> str | None:
    if record.parametros and record.parametros.get("lags"):
        return ",".join(str(value) for value in record.parametros["lags"])
    for item in horizon_map.values():
        text = str(item.get("variables_temporales", ""))
        match = re.search(r"lags\s*\[([0-9,\s]+)\]", text)
        if match:
            return ",".join(token.strip() for token in match.group(1).split(","))
        if "1..4" in text:
            return "1,2,3,4"
    return None


def extract_horizons(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> str | None:
    if record.parametros and record.parametros.get("horizons"):
        return ",".join(str(value) for value in record.parametros["horizons"])
    if horizon_map:
        return ",".join(str(value) for value in sorted(horizon_map))
    return None


def parse_notes_config_payload(raw_value: Any) -> dict[str, Any]:
    if raw_value is None:
        return {}
    if isinstance(raw_value, dict):
        return raw_value
    if isinstance(raw_value, str):
        token = raw_value.strip()
        if not token:
            return {}
        try:
            payload = json.loads(token)
        except json.JSONDecodeError:
            return {}
        return payload if isinstance(payload, dict) else {}
    return {}


def extract_first_notes_config(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> dict[str, Any]:
    for horizon in sorted(horizon_map):
        payload = parse_notes_config_payload(horizon_map[horizon].get("notas_config"))
        if payload:
            return payload
    return {}


def extract_initial_train_size(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> int | None:
    if record.parametros and record.parametros.get("initial_train_size") is not None:
        return int(record.parametros["initial_train_size"])
    notes_payload = extract_first_notes_config(record, horizon_map)
    if notes_payload.get("initial_train_size") is not None:
        return int(notes_payload["initial_train_size"])
    return None


def infer_tuning_interno(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> bool | None:
    model_params = {}
    if record.parametros and isinstance(record.parametros.get("model_params"), dict):
        model_params = record.parametros["model_params"]
    tuning_strategy = str(model_params.get("tuning_strategy", "")).strip().lower()
    if tuning_strategy:
        if "sin_tuning" in tuning_strategy or "baseline_fijo" in tuning_strategy:
            return False
        if any(token in tuning_strategy for token in ("tscv", "temporal", "interno", "grid")):
            return True

    model_name = (record.model or "").strip().lower()
    if "tscv" in model_name:
        return True
    if any(token in model_name for token in ("random_forest", "extra_trees", "xgboost")):
        return False

    notes_payload = extract_first_notes_config(record, horizon_map)
    model_params_notes = notes_payload.get("model_params", {})
    if isinstance(model_params_notes, dict):
        tuning_strategy_notes = str(model_params_notes.get("tuning_strategy", "")).strip().lower()
        if tuning_strategy_notes:
            if "sin_tuning" in tuning_strategy_notes or "baseline_fijo" in tuning_strategy_notes:
                return False
            if any(token in tuning_strategy_notes for token in ("tscv", "temporal", "interno", "grid")):
                return True
    return None


def normalize_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    raw = str(value).strip().lower()
    if raw in {"true", "1", "si", "sí", "yes"}:
        return True
    if raw in {"false", "0", "no"}:
        return False
    return None


def normalize_created_at(record: RunDirectoryRecord) -> str | None:
    if record.metadata and record.metadata.get("created_at"):
        return str(record.metadata["created_at"])
    if record.timestamp:
        token = record.timestamp
        return f"{token[0:4]}-{token[4:6]}-{token[6:8]} {token[9:11]}:{token[11:13]}:{token[13:15]}"
    return None


def extract_notes_config_text(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> str:
    for horizon in sorted(horizon_map):
        raw = horizon_map[horizon].get("notas_config")
        if raw not in (None, ""):
            return str(raw)
    if record.parametros:
        return json.dumps(record.parametros, ensure_ascii=False, sort_keys=True)
    return ""


def resolve_script_family(record: RunDirectoryRecord) -> str | None:
    if record.metadata and record.metadata.get("script_path"):
        return Path(str(record.metadata["script_path"])).name
    return None


def resolve_path_if_exists(path: Path) -> str:
    return str(path.resolve()) if path.exists() else ""


def get_prediction_file(record: RunDirectoryRecord, horizon: int) -> Path:
    return record.path / f"predicciones_h{horizon}.csv"


def extract_script_name(record: RunDirectoryRecord) -> str | None:
    script_path = (record.metadata or {}).get("script_path")
    if not script_path:
        return None
    return Path(str(script_path)).name


def extract_script_path(record: RunDirectoryRecord) -> str | None:
    script_path = (record.metadata or {}).get("script_path")
    return str(script_path) if script_path else None


def extract_timestamp_run(record: RunDirectoryRecord) -> str | None:
    if record.timestamp:
        return record.timestamp
    created_at = normalize_created_at(record)
    if created_at:
        return created_at.replace("-", "").replace(":", "").replace(" ", "_")
    return None


def extract_version_canonica(run_id: str) -> str:
    if "_" not in run_id:
        return run_id
    return run_id.split("_", 1)[1]


def build_status_canonico(record: RunDirectoryRecord) -> str:
    if record.artifact_status == "completo":
        return "canonico_completo"
    if record.artifact_status == "parcial":
        return "canonico_parcial"
    return "canonico_inconsistente"


def build_canonical_exclusion_reason(record: RunDirectoryRecord) -> str:
    if not record.issues:
        return ""
    return ";".join(record.issues)


def resumen_by_horizon(record: RunDirectoryRecord) -> dict[int, dict[str, Any]]:
    if not record.resumen:
        return {}
    result: dict[int, dict[str, Any]] = {}
    for item in record.resumen:
        horizonte = item.get("horizonte_sem")
        if horizonte is None:
            continue
        result[int(horizonte)] = item
    return result


def extract_validation_scheme(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> str | None:
    for horizon in sorted(horizon_map):
        value = horizon_map[horizon].get("validacion")
        if value not in (None, ""):
            return str(value)
    notes_payload = extract_first_notes_config(record, horizon_map)
    if notes_payload.get("validation_scheme") not in (None, ""):
        return str(notes_payload["validation_scheme"])
    model_name = (record.model or "").strip().lower()
    if "sarimax" in model_name or "arimax" in model_name:
        return "walk-forward_expanding"
    return None


def extract_tuning_metric(record: RunDirectoryRecord, horizon_map: dict[int, dict[str, Any]]) -> str | None:
    params = record.parametros or {}
    model_params = params.get("model_params", {})
    if isinstance(model_params, dict) and model_params.get("tuning_metric") not in (None, ""):
        return str(model_params["tuning_metric"])
    notes_payload = extract_first_notes_config(record, horizon_map)
    model_params_notes = notes_payload.get("model_params", {})
    if isinstance(model_params_notes, dict) and model_params_notes.get("tuning_metric") not in (None, ""):
        return str(model_params_notes["tuning_metric"])
    return None


def extract_feature_counts_by_horizon(
    record: RunDirectoryRecord,
    horizon_map: dict[int, dict[str, Any]],
) -> dict[int, float | None]:
    resumen_map = resumen_by_horizon(record)
    counts: dict[int, float | None] = {1: None, 2: None, 3: None, 4: None}
    for horizon in (1, 2, 3, 4):
        resumen_item = resumen_map.get(horizon, {})
        if resumen_item.get("selected_feature_count_avg") is not None:
            counts[horizon] = float(resumen_item["selected_feature_count_avg"])
            continue
        notes_payload = parse_notes_config_payload(horizon_map.get(horizon, {}).get("notas_config"))
        if notes_payload.get("selected_feature_count_avg") is not None:
            counts[horizon] = float(notes_payload["selected_feature_count_avg"])
    return counts


def extract_source_dataset_period(horizon_map: dict[int, dict[str, Any]]) -> str | None:
    for horizon in sorted(horizon_map):
        dataset_period = horizon_map[horizon].get("dataset_periodo")
        if dataset_period not in (None, ""):
            return str(dataset_period)
    return None


def extract_dataset_path(record: RunDirectoryRecord) -> str | None:
    params = record.parametros or {}
    dataset_path = params.get("dataset_path")
    return str(dataset_path) if dataset_path else None


def extract_feature_columns(record: RunDirectoryRecord) -> list[str]:
    params = record.parametros or {}
    feature_columns = params.get("feature_columns")
    if isinstance(feature_columns, list):
        return [str(value) for value in feature_columns]
    return []


def extract_raw_model_params(record: RunDirectoryRecord) -> dict[str, Any]:
    params = record.parametros or {}
    raw: dict[str, Any] = {}
    model_params = params.get("model_params")
    if isinstance(model_params, dict):
        raw.update(model_params)
    for key, value in params.items():
        if key in COMMON_PARAM_EXCLUDE_KEYS:
            continue
        if key == "model_params":
            continue
        raw[key] = value
    return raw


def summarize_numeric_list(values: list[Any]) -> dict[str, Any]:
    numeric_values = [float(value) for value in values]
    return {
        "size": len(numeric_values),
        "min": min(numeric_values),
        "max": max(numeric_values),
    }


def normalize_param_value(key: str, value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, tuple):
        return [normalize_param_value(key, item) for item in value]
    if isinstance(value, list):
        if not value:
            return []
        if key in {"alpha_grid", "alphas"} and all(isinstance(item, (int, float)) for item in value):
            return summarize_numeric_list(value)
        if all(isinstance(item, (int, float)) for item in value):
            if len(value) <= 8:
                return [float(item) if isinstance(item, float) else item for item in value]
            return summarize_numeric_list(value)
        if len(value) <= 6:
            return [normalize_param_value(key, item) for item in value]
        return {"size": len(value)}
    if isinstance(value, dict):
        if key == "param_grid":
            return {
                inner_key: normalize_param_value(inner_key, inner_value)
                for inner_key, inner_value in sorted(value.items())
            }
        return {
            str(inner_key): normalize_param_value(str(inner_key), inner_value)
            for inner_key, inner_value in sorted(value.items())
            if inner_value not in (None, "", [], {})
        }
    return str(value)


def select_model_hyperparam_keys(model_name: str, raw_params: dict[str, Any]) -> list[str]:
    model_lower = model_name.strip().lower()
    for tokens, keys in MODEL_HYPERPARAM_KEYS:
        if any(token in model_lower for token in tokens):
            selected = [key for key in keys if key in raw_params]
            unknown = [key for key in sorted(raw_params) if key not in selected]
            return selected + unknown
    return sorted(raw_params)


def compact_value_for_signature(value: Any) -> str:
    if isinstance(value, list):
        return ",".join(str(item) for item in value)
    if isinstance(value, dict):
        if set(value.keys()) == {"size", "min", "max"}:
            return f"{value['size']}[{value['min']}..{value['max']}]"
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return str(value)


def build_hyperparams_summary(normalized: dict[str, Any]) -> str:
    priority_keys = (
        "target_mode",
        "feature_mode",
        "lags",
        "horizons",
        "transform_mode",
        "order",
        "trend",
        "iterations",
        "depth",
        "learning_rate",
        "l2_leaf_reg",
        "n_estimators",
        "max_depth",
        "max_features",
        "subsample",
        "colsample_bytree",
        "reg_alpha",
        "reg_lambda",
        "inner_splits",
        "tuning_strategy",
    )
    parts: list[str] = []
    used_keys: set[str] = set()
    for key in priority_keys:
        value = normalized.get(key)
        if value in (None, "", [], {}):
            continue
        parts.append(f"{key}={compact_value_for_signature(value)}")
        used_keys.add(key)
    for key in sorted(normalized):
        if key in used_keys:
            continue
        value = normalized[key]
        if value in (None, "", [], {}):
            continue
        parts.append(f"{key}={compact_value_for_signature(value)}")
    return " | ".join(parts[:8])


def normalize_hyperparams(
    *,
    record: RunDirectoryRecord,
    target_mode: str | None,
    feature_mode: str | None,
    transform_mode: str | None,
    lags: str | None,
    horizons: str | None,
    initial_train_size: int | None,
    validation_scheme: str | None,
) -> HyperparamsNormalizationResult:
    raw_params = extract_raw_model_params(record)
    model_name = record.model or ""
    relevant_keys = select_model_hyperparam_keys(model_name, raw_params)

    normalized: dict[str, Any] = {
        "target_mode": target_mode,
        "feature_mode": feature_mode,
        "transform_mode": transform_mode,
        "lags": [int(token) for token in str(lags).split(",") if str(token).strip().isdigit()] if lags else None,
        "horizons": [int(token) for token in str(horizons).split(",") if str(token).strip().isdigit()] if horizons else None,
        "initial_train_size": initial_train_size,
        "validation_scheme": validation_scheme,
    }
    for key in relevant_keys:
        value = raw_params.get(key)
        if value in (None, "", [], {}):
            continue
        normalized[key] = normalize_param_value(key, value)

    normalized = {key: value for key, value in normalized.items() if value not in (None, "", [], {})}
    hyperparams_json = json.dumps(normalized, ensure_ascii=False, sort_keys=True)
    hyperparams_hash = hashlib.sha1(hyperparams_json.encode("utf-8")).hexdigest()[:12] if normalized else ""
    hyperparams_resumen = build_hyperparams_summary(normalized)
    hyperparams_firma = f"{record.model}|{hyperparams_resumen}" if hyperparams_resumen else str(record.model or "")
    hyperparams_firma = hyperparams_firma[:240]

    structural_ok = all(value not in (None, "", [], {}) for value in (record.model, target_mode, lags, horizons))
    model_specific_keys = {
        key
        for key in normalized
        if key
        not in {
            "target_mode",
            "feature_mode",
            "transform_mode",
            "lags",
            "horizons",
            "initial_train_size",
            "validation_scheme",
        }
    }
    if not model_specific_keys:
        status = "no_recuperable"
        note = "sin_parametros_modelo_suficientes"
    elif structural_ok:
        status = "completo"
        note = "reconstruccion_desde_parametros_y_metadata"
    else:
        status = "parcial"
        note = "faltan_campos_estructurales_pero_el_modelo_se_reconstruyo_parcialmente"

    return HyperparamsNormalizationResult(
        hyperparams_json=hyperparams_json,
        hyperparams_hash=hyperparams_hash,
        hyperparams_firma=hyperparams_firma,
        hyperparams_resumen=hyperparams_resumen,
        reconstruction_status=status,
        reconstruction_note=note,
    )


def detect_explainability_artifacts(record: RunDirectoryRecord) -> ExplainabilityAuditResult:
    coef_paths: list[Path] = []
    importance_paths: list[Path] = []
    shap_paths: list[Path] = []
    for artifact in record.path.rglob("*"):
        if not artifact.is_file():
            continue
        name = artifact.name.lower()
        if any(token in name for token in EXPLAINABILITY_COEF_PATTERNS):
            coef_paths.append(artifact)
        if any(token in name for token in EXPLAINABILITY_IMPORTANCE_PATTERNS):
            importance_paths.append(artifact)
        if any(token in name for token in EXPLAINABILITY_SHAP_PATTERNS):
            shap_paths.append(artifact)

    primary_path = ""
    if coef_paths:
        primary_path = str(coef_paths[0].resolve())
    elif importance_paths:
        primary_path = str(importance_paths[0].resolve())
    elif shap_paths:
        primary_path = str(shap_paths[0].resolve())

    tiene_coeficientes = bool(coef_paths)
    tiene_importancias = bool(importance_paths)
    tiene_shap = bool(shap_paths)
    explicabilidad_homogenea = bool(primary_path)

    if explicabilidad_homogenea:
        observacion = "Artefacto explicativo localizado, pero la homogeneidad transversal debe validarse por familia."
    elif record.selected_feature_files:
        observacion = "Solo hay features_seleccionadas por horizonte; no hay coeficientes/importancias comparables."
    else:
        observacion = "Sin artefactos explicativos comparables exportados."

    return ExplainabilityAuditResult(
        tiene_coeficientes=tiene_coeficientes,
        tiene_importancias=tiene_importancias,
        tiene_shap_o_equivalente=tiene_shap,
        explicabilidad_transversal_homogenea=explicabilidad_homogenea,
        artifact_explicabilidad_path=primary_path,
        observacion_explicabilidad=observacion,
    )


def compute_best_horizon_by_loss(row: dict[str, Any], *, task_type: str) -> str | None:
    metric_suffix = "loss_clasificacion" if task_type == "clasificacion" else "loss"
    values = {
        f"H{horizon}": row.get(f"H{horizon}_{metric_suffix}")
        for horizon in (1, 2, 3, 4)
        if row.get(f"H{horizon}_{metric_suffix}") is not None
    }
    if not values:
        return None
    return min(values.items(), key=lambda item: float(item[1]))[0]


def build_fortalezas_operativas(row: dict[str, Any], *, task_type: str) -> str:
    if task_type == "clasificacion":
        values = {
            f"H{horizon}": row.get(f"H{horizon}_recall_baja_fuerte")
            for horizon in (1, 2, 3, 4)
            if row.get(f"H{horizon}_recall_baja_fuerte") is not None
        }
        if not values:
            return ""
        best_h, best_value = max(values.items(), key=lambda item: float(item[1]))
        return f"mejor_recall_baja_fuerte={best_h}:{best_value:.6f}"

    dir_values = {
        f"H{horizon}": row.get(f"H{horizon}_direction_accuracy")
        for horizon in (1, 2, 3, 4)
        if row.get(f"H{horizon}_direction_accuracy") is not None
    }
    risk_values = {
        f"H{horizon}": row.get(f"H{horizon}_deteccion_caidas")
        for horizon in (1, 2, 3, 4)
        if row.get(f"H{horizon}_deteccion_caidas") is not None
    }
    best_dir = max(dir_values.items(), key=lambda item: float(item[1])) if dir_values else None
    best_risk = max(risk_values.items(), key=lambda item: float(item[1])) if risk_values else None
    parts: list[str] = []
    if best_dir is not None:
        parts.append(f"mejor_dir={best_dir[0]}:{best_dir[1]:.6f}")
    if best_risk is not None:
        parts.append(f"mejor_caidas={best_risk[0]}:{best_risk[1]:.6f}")
    return " | ".join(parts)


def build_notas_config_clave(
    *,
    validation_scheme: str | None,
    tuning_interno: bool | None,
    tuning_metric: str | None,
    hyperparams_summary: str,
) -> str:
    parts: list[str] = []
    if validation_scheme:
        parts.append(f"validacion={validation_scheme}")
    if tuning_interno is not None:
        parts.append("tuning=si" if tuning_interno else "tuning=no")
    if tuning_metric:
        parts.append(f"metric_tuning={tuning_metric}")
    if hyperparams_summary:
        parts.append(hyperparams_summary)
    return " | ".join(parts)


def build_master_rows(
    canonical: dict[str, RunDirectoryRecord],
    summary_lookup: dict[str, dict[str, Any]],
    results_lookup: dict[str, dict[int, dict[str, Any]]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for run_id in sorted(canonical):
        record = canonical[run_id]
        horizon_map = horizon_metrics_map(record)
        summary = summary_lookup.get(run_id, {})
        workbook_horizons = results_lookup.get(run_id, {})
        feature_count_prom = parse_feature_count_prom(record.resumen, record.metricas)
        task_type = extract_task_type(record, summary)
        row: dict[str, Any] = {
            "run_id": run_id,
            "task_type": task_type,
            "family": normalize_family(run_id, (record.metadata or {}).get("family"), record.model),
            "model": record.model,
            "target_mode": extract_target_mode(record, horizon_map),
            "feature_mode": extract_feature_mode(record, horizon_map),
            "transform_mode": extract_transform_mode(record, horizon_map),
            "lags": extract_lags(record, horizon_map),
            "feature_count_prom": feature_count_prom,
            "L_total_Radar": summary.get("L_total_Radar"),
            "L_total_Clasificacion": summary.get("L_total_Clasificacion"),
            "observacion_breve": summary.get("Comentarios") or (
                "Run consolidado desde artefactos en disco."
                if record.artifact_status == "completo"
                else "Run con artefactos incompletos."
            ),
            "status_run": record.artifact_status,
        }
        for horizon in (1, 2, 3, 4):
            horizon_row = horizon_map.get(horizon, {})
            workbook_row = workbook_horizons.get(horizon, {})
            row[f"H{horizon}_mae"] = horizon_row.get("mae", workbook_row.get("MAE"))
            row[f"H{horizon}_rmse"] = horizon_row.get("rmse", workbook_row.get("RMSE"))
            row[f"H{horizon}_direction_accuracy"] = horizon_row.get(
                "direccion_accuracy", workbook_row.get("Direccion_accuracy")
            )
            row[f"H{horizon}_deteccion_caidas"] = horizon_row.get(
                "deteccion_caidas", workbook_row.get("Deteccion_caidas")
            )
            row[f"H{horizon}_loss"] = horizon_row.get("loss_h", workbook_row.get("Loss_h"))
            row[f"H{horizon}_accuracy_5clases"] = horizon_row.get(
                "accuracy_5clases", workbook_row.get("Accuracy_5clases")
            )
            row[f"H{horizon}_balanced_accuracy_5clases"] = horizon_row.get(
                "balanced_accuracy_5clases", workbook_row.get("Balanced_accuracy_5clases")
            )
            row[f"H{horizon}_macro_f1_5clases"] = horizon_row.get(
                "macro_f1_5clases", workbook_row.get("Macro_f1_5clases")
            )
            row[f"H{horizon}_weighted_f1_5clases"] = horizon_row.get(
                "weighted_f1_5clases", workbook_row.get("Weighted_f1_5clases")
            )
            row[f"H{horizon}_recall_baja_fuerte"] = horizon_row.get(
                "recall_baja_fuerte", workbook_row.get("Recall_baja_fuerte")
            )
            row[f"H{horizon}_recall_baja_total"] = horizon_row.get(
                "recall_baja_total", workbook_row.get("Recall_baja_total")
            )
            row[f"H{horizon}_precision_baja_fuerte"] = horizon_row.get(
                "precision_baja_fuerte", workbook_row.get("Precision_baja_fuerte")
            )
            row[f"H{horizon}_recall_sube_fuerte"] = horizon_row.get(
                "recall_sube_fuerte", workbook_row.get("Recall_sube_fuerte")
            )
            row[f"H{horizon}_recall_sube_total"] = horizon_row.get(
                "recall_sube_total", workbook_row.get("Recall_sube_total")
            )
            row[f"H{horizon}_accuracy_3clases"] = horizon_row.get(
                "accuracy_3clases", workbook_row.get("Accuracy_3clases")
            )
            row[f"H{horizon}_macro_f1_3clases"] = horizon_row.get(
                "macro_f1_3clases", workbook_row.get("Macro_f1_3clases")
            )
            row[f"H{horizon}_balanced_accuracy_3clases"] = horizon_row.get(
                "balanced_accuracy_3clases", workbook_row.get("Balanced_accuracy_3clases")
            )
            row[f"H{horizon}_loss_clasificacion"] = horizon_row.get(
                "loss_clasificacion_h", workbook_row.get("Loss_clasificacion_h")
            )

        row["mae_promedio"] = summary.get("Avg_MAE")
        if row["mae_promedio"] is None:
            mae_values = [row[f"H{h}_mae"] for h in (1, 2, 3, 4) if row[f"H{h}_mae"] is not None]
            row["mae_promedio"] = float(sum(mae_values) / len(mae_values)) if mae_values else None

        row["rmse_promedio"] = summary.get("Avg_RMSE")
        if row["rmse_promedio"] is None:
            rmse_values = [row[f"H{h}_rmse"] for h in (1, 2, 3, 4) if row[f"H{h}_rmse"] is not None]
            row["rmse_promedio"] = float(sum(rmse_values) / len(rmse_values)) if rmse_values else None

        row["direction_accuracy_promedio"] = summary.get("Dir_acc_prom")
        if row["direction_accuracy_promedio"] is None:
            dir_values = [
                row[f"H{h}_direction_accuracy"]
                for h in (1, 2, 3, 4)
                if row[f"H{h}_direction_accuracy"] is not None
            ]
            row["direction_accuracy_promedio"] = (
                float(sum(dir_values) / len(dir_values)) if dir_values else None
            )

        row["deteccion_caidas_promedio"] = summary.get("Det_caidas_prom")
        if row["deteccion_caidas_promedio"] is None:
            risk_values = [
                row[f"H{h}_deteccion_caidas"]
                for h in (1, 2, 3, 4)
                if row[f"H{h}_deteccion_caidas"] is not None
            ]
            row["deteccion_caidas_promedio"] = (
                float(sum(risk_values) / len(risk_values)) if risk_values else None
            )

        row["accuracy_5clases_promedio"] = summary.get("Avg_Accuracy_5clases")
        row["balanced_accuracy_5clases_promedio"] = summary.get("Avg_Balanced_accuracy_5clases")
        row["macro_f1_5clases_promedio"] = summary.get("Avg_Macro_f1_5clases")
        row["weighted_f1_5clases_promedio"] = summary.get("Avg_Weighted_f1_5clases")
        row["recall_baja_fuerte_promedio"] = summary.get("Avg_Recall_baja_fuerte")
        row["recall_baja_total_promedio"] = summary.get("Avg_Recall_baja_total")
        row["precision_baja_fuerte_promedio"] = summary.get("Avg_Precision_baja_fuerte")
        row["recall_sube_fuerte_promedio"] = summary.get("Avg_Recall_sube_fuerte")
        row["recall_sube_total_promedio"] = summary.get("Avg_Recall_sube_total")
        row["accuracy_3clases_promedio"] = summary.get("Avg_Accuracy_3clases")
        row["macro_f1_3clases_promedio"] = summary.get("Avg_Macro_f1_3clases")
        row["balanced_accuracy_3clases_promedio"] = summary.get("Avg_Balanced_accuracy_3clases")

        if task_type == "clasificacion":
            required_columns = [
                "L_total_Clasificacion",
                "H1_accuracy_5clases",
                "H1_balanced_accuracy_5clases",
                "H1_macro_f1_5clases",
                "H1_recall_baja_fuerte",
                "H1_loss_clasificacion",
                "H2_accuracy_5clases",
                "H2_balanced_accuracy_5clases",
                "H2_macro_f1_5clases",
                "H2_recall_baja_fuerte",
                "H2_loss_clasificacion",
                "H3_accuracy_5clases",
                "H3_balanced_accuracy_5clases",
                "H3_macro_f1_5clases",
                "H3_recall_baja_fuerte",
                "H3_loss_clasificacion",
                "H4_accuracy_5clases",
                "H4_balanced_accuracy_5clases",
                "H4_macro_f1_5clases",
                "H4_recall_baja_fuerte",
                "H4_loss_clasificacion",
            ]
        else:
            required_columns = [
                "L_total_Radar",
                "H1_mae",
                "H1_rmse",
                "H1_direction_accuracy",
                "H1_deteccion_caidas",
                "H1_loss",
                "H2_mae",
                "H2_rmse",
                "H2_direction_accuracy",
                "H2_deteccion_caidas",
                "H2_loss",
                "H3_mae",
                "H3_rmse",
                "H3_direction_accuracy",
                "H3_deteccion_caidas",
                "H3_loss",
                "H4_mae",
                "H4_rmse",
                "H4_direction_accuracy",
                "H4_deteccion_caidas",
                "H4_loss",
            ]
        missing_required = [column for column in required_columns if row.get(column) is None]
        if missing_required:
            row["status_run"] = "parcial" if record.artifact_status == "completo" else record.artifact_status
            if row["observacion_breve"]:
                row["observacion_breve"] = (
                    f"{row['observacion_breve']} | faltantes_metricos={','.join(missing_required)}"
                )
        rows.append(row)
    return rows


def build_inventory_rows(
    records: list[RunDirectoryRecord],
    canonical: dict[str, RunDirectoryRecord],
    grouped: dict[str, list[RunDirectoryRecord]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in sorted(records, key=lambda item: item.path.name):
        run_id = record.effective_run_id
        is_canonical = canonical.get(run_id) == record
        duplicate_count = len(grouped.get(run_id, []))
        comment_parts = []
        if duplicate_count > 1:
            comment_parts.append(f"duplicado_detectado={duplicate_count}")
            if is_canonical:
                comment_parts.append("directorio_valido_resuelto")
            else:
                comment_parts.append("intento_descartado_por_menor_completitud")
        if record.is_aborted:
            comment_parts.append("directorio_abortado")
        if record.issues:
            comment_parts.append(",".join(record.issues))
        rows.append(
            {
                "run_id": run_id,
                "family": normalize_family(run_id, (record.metadata or {}).get("family"), record.model),
                "model": record.model,
                "path_run": str(record.path.resolve()),
                "metadata_run": str((record.path / "metadata_run.json").resolve())
                if (record.path / "metadata_run.json").exists()
                else "",
                "metricas_horizonte": str((record.path / "metricas_horizonte.json").resolve())
                if (record.path / "metricas_horizonte.json").exists()
                else "",
                "parametros_run": str((record.path / "parametros_run.json").resolve())
                if (record.path / "parametros_run.json").exists()
                else "",
                "predicciones": len(record.prediction_files),
                "comparaciones": len(record.comparison_files),
                "features_seleccionadas": len(record.selected_feature_files),
                "resumen_horizontes": int((record.path / "resumen_modeling_horizontes.json").exists()),
                "status_artifactos": record.artifact_status,
                "comentario": "; ".join(comment_parts),
                "es_directorio_canonico": is_canonical,
            }
        )
    return rows


def build_runs_catalog_rows(
    canonical: dict[str, RunDirectoryRecord],
    summary_lookup: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for run_id in sorted(canonical):
        record = canonical[run_id]
        horizon_map = horizon_metrics_map(record)
        resumen_map = resumen_by_horizon(record)
        summary = summary_lookup.get(run_id, {})
        task_type = extract_task_type(record, summary)
        target_mode = extract_target_mode(record, horizon_map)
        feature_mode = extract_feature_mode(record, horizon_map)
        transform_mode = extract_transform_mode(record, horizon_map)
        lags = extract_lags(record, horizon_map)
        horizons = extract_horizons(record, horizon_map)
        initial_train_size = extract_initial_train_size(record, horizon_map)
        validation_scheme = extract_validation_scheme(record, horizon_map)
        tuning_interno = infer_tuning_interno(record, horizon_map)
        tuning_metric = extract_tuning_metric(record, horizon_map)
        feature_counts = extract_feature_counts_by_horizon(record, horizon_map)
        feature_count_promedio = parse_feature_count_prom(record.resumen, record.metricas)
        hyperparams = normalize_hyperparams(
            record=record,
            target_mode=target_mode,
            feature_mode=feature_mode,
            transform_mode=transform_mode,
            lags=lags,
            horizons=horizons,
            initial_train_size=initial_train_size,
            validation_scheme=validation_scheme,
        )
        explainability = detect_explainability_artifacts(record)
        dataset_period = extract_source_dataset_period(horizon_map)
        dataset_path = extract_dataset_path(record)
        feature_columns = extract_feature_columns(record)
        observacion_breve = summary.get("Comentarios") or (
            "Run consolidado desde artefactos en disco."
            if record.artifact_status == "completo"
            else "Run con artefactos incompletos."
        )
        metadata_path = record.path / "metadata_run.json"
        parametros_path = record.path / "parametros_run.json"
        metricas_path = record.path / "metricas_horizonte.json"
        resumen_path = record.path / "resumen_modeling_horizontes.json"
        row: dict[str, Any] = {
            "run_id": run_id,
            "task_type": task_type,
            "family": normalize_family(run_id, (record.metadata or {}).get("family"), record.model),
            "model": record.model,
            "script_family": resolve_script_family(record),
            "script_nombre": extract_script_name(record),
            "script_ruta": extract_script_path(record),
            "target_mode": target_mode,
            "feature_mode": feature_mode,
            "transform_mode": transform_mode,
            "lags": lags,
            "horizons": horizons,
            "initial_train_size": initial_train_size,
            "validation_scheme": validation_scheme,
            "tuning_interno": tuning_interno,
            "tuning_metric": tuning_metric,
            "fecha_run": normalize_created_at(record),
            "timestamp_run": extract_timestamp_run(record),
            "run_dir": str(record.path.resolve()),
            "version_canonica": extract_version_canonica(run_id),
            "es_run_maestro": True,
            "status_canonico": build_status_canonico(record),
            "motivo_exclusion_si_aplica": build_canonical_exclusion_reason(record),
            "status_run": record.artifact_status,
            "L_total_Radar": summary.get("L_total_Radar"),
            "L_total_Clasificacion": summary.get("L_total_Clasificacion"),
            "path_run": str(record.path.resolve()),
            "metadata_run_path": resolve_path_if_exists(metadata_path),
            "parametros_run_path": resolve_path_if_exists(parametros_path),
            "metricas_horizonte_path": resolve_path_if_exists(metricas_path),
            "resumen_horizontes_path": resolve_path_if_exists(resumen_path),
            "dataset_path": dataset_path or "",
            "source_dataset_period": dataset_period or "",
            "notas_config": extract_notes_config_text(record, horizon_map),
            "notas_config_clave": build_notas_config_clave(
                validation_scheme=validation_scheme,
                tuning_interno=tuning_interno,
                tuning_metric=tuning_metric,
                hyperparams_summary=hyperparams.hyperparams_resumen,
            ),
            "hyperparams_json": hyperparams.hyperparams_json,
            "hyperparams_hash": hyperparams.hyperparams_hash,
            "hyperparams_firma": hyperparams.hyperparams_firma,
            "hyperparams_resumen": hyperparams.hyperparams_resumen,
            "reconstruccion_hiperparams_status": hyperparams.reconstruction_status,
            "reconstruccion_hiperparams_nota": hyperparams.reconstruction_note,
            "feature_count_promedio": feature_count_promedio,
            "feature_count_h1": feature_counts.get(1),
            "feature_count_h2": feature_counts.get(2),
            "feature_count_h3": feature_counts.get(3),
            "feature_count_h4": feature_counts.get(4),
            "features_artifact_available": bool(record.selected_feature_files),
            "seleccion_variables_tipo": feature_mode,
            "feature_columns_count": len(feature_columns),
            "exogenas_o_no": "con_exogenas" if feature_columns else "sin_exogenas",
            "usa_target_delta": target_mode == "delta",
            "usa_target_nivel": target_mode == "nivel",
            "mae_promedio": summary.get("Avg_MAE"),
            "rmse_promedio": summary.get("Avg_RMSE"),
            "direction_accuracy_promedio": summary.get("Dir_acc_prom"),
            "deteccion_caidas_promedio": summary.get("Det_caidas_prom"),
            "loss_h1": horizon_map.get(1, {}).get("loss_h"),
            "loss_h2": horizon_map.get(2, {}).get("loss_h"),
            "loss_h3": horizon_map.get(3, {}).get("loss_h"),
            "loss_h4": horizon_map.get(4, {}).get("loss_h"),
            "mejor_horizonte_por_loss": "",
            "fortalezas_operativas": "",
            "observacion_breve": observacion_breve,
            "tiene_predicciones_h1": False,
            "tiene_predicciones_h2": False,
            "tiene_predicciones_h3": False,
            "tiene_predicciones_h4": False,
            "filas_pred_h1": 0,
            "filas_pred_h2": 0,
            "filas_pred_h3": 0,
            "filas_pred_h4": 0,
            "mergeable_h1": False,
            "mergeable_h2": False,
            "mergeable_h3": False,
            "mergeable_h4": False,
            "stacking_eligible_global": False,
            "stacking_eligible_h1": False,
            "stacking_eligible_h2": False,
            "stacking_eligible_h3": False,
            "stacking_eligible_h4": False,
            "es_candidato_meta_modelo": False,
            "motivo_no_elegibilidad": "",
            "motivo_exclusion_meta_modelo": "",
            "tiene_coeficientes": explainability.tiene_coeficientes,
            "tiene_importancias": explainability.tiene_importancias,
            "tiene_shap_o_equivalente": explainability.tiene_shap_o_equivalente,
            "explicabilidad_transversal_homogenea": explainability.explicabilidad_transversal_homogenea,
            "artifact_explicabilidad_path": explainability.artifact_explicabilidad_path,
            "observacion_explicabilidad": explainability.observacion_explicabilidad,
        }
        for horizon in (1, 2, 3, 4):
            pred_path = get_prediction_file(record, horizon)
            row[f"predicciones_h{horizon}_path"] = resolve_path_if_exists(pred_path)
            row[f"predicciones_h{horizon}_existe"] = pred_path.exists()
            if row.get(f"loss_h{horizon}") is None and horizon_map.get(horizon, {}).get("loss_h") is not None:
                row[f"loss_h{horizon}"] = horizon_map[horizon]["loss_h"]
            if row.get(f"feature_count_h{horizon}") is None and resumen_map.get(horizon, {}).get("selected_feature_count_avg") is not None:
                row[f"feature_count_h{horizon}"] = float(resumen_map[horizon]["selected_feature_count_avg"])
        row["mejor_horizonte_por_loss"] = compute_best_horizon_by_loss(row, task_type=task_type)
        row["fortalezas_operativas"] = build_fortalezas_operativas(row, task_type=task_type)
        rows.append(row)
    return rows


def build_metricas_por_horizonte_long(
    canonical: dict[str, RunDirectoryRecord],
    master_df: pd.DataFrame,
    results_lookup: dict[str, dict[int, dict[str, Any]]],
) -> pd.DataFrame:
    master_lookup = {
        row["run_id"]: row
        for row in master_df.to_dict(orient="records")
    }
    rows: list[dict[str, Any]] = []
    for run_id in sorted(canonical):
        record = canonical[run_id]
        horizon_map = horizon_metrics_map(record)
        workbook_horizons = results_lookup.get(run_id, {})
        master_row = master_lookup.get(run_id, {})
        for horizon in (1, 2, 3, 4):
            horizon_row = horizon_map.get(horizon, {})
            workbook_row = workbook_horizons.get(horizon, {})
            rows.append(
                {
                    "run_id": run_id,
                    "task_type": master_row.get("task_type"),
                    "horizonte": horizon,
                    "horizonte_label": f"H{horizon}",
                    "family": master_row.get("family"),
                    "model": master_row.get("model"),
                    "target_mode": master_row.get("target_mode"),
                    "feature_mode": master_row.get("feature_mode"),
                    "lags": master_row.get("lags"),
                    "mae": horizon_row.get("mae", workbook_row.get("MAE")),
                    "rmse": horizon_row.get("rmse", workbook_row.get("RMSE")),
                    "direction_accuracy": horizon_row.get(
                        "direccion_accuracy", workbook_row.get("Direccion_accuracy")
                    ),
                    "deteccion_caidas": horizon_row.get(
                        "deteccion_caidas", workbook_row.get("Deteccion_caidas")
                    ),
                    "l_num": horizon_row.get("l_num"),
                    "l_trend": horizon_row.get("l_trend"),
                    "l_risk": horizon_row.get("l_risk"),
                    "l_tol": horizon_row.get("l_tol"),
                    "loss_h": horizon_row.get("loss_h", workbook_row.get("Loss_h")),
                    "accuracy_5clases": horizon_row.get("accuracy_5clases", workbook_row.get("Accuracy_5clases")),
                    "balanced_accuracy_5clases": horizon_row.get(
                        "balanced_accuracy_5clases", workbook_row.get("Balanced_accuracy_5clases")
                    ),
                    "macro_f1_5clases": horizon_row.get("macro_f1_5clases", workbook_row.get("Macro_f1_5clases")),
                    "weighted_f1_5clases": horizon_row.get(
                        "weighted_f1_5clases", workbook_row.get("Weighted_f1_5clases")
                    ),
                    "recall_baja_fuerte": horizon_row.get(
                        "recall_baja_fuerte", workbook_row.get("Recall_baja_fuerte")
                    ),
                    "recall_baja_total": horizon_row.get(
                        "recall_baja_total", workbook_row.get("Recall_baja_total")
                    ),
                    "precision_baja_fuerte": horizon_row.get(
                        "precision_baja_fuerte", workbook_row.get("Precision_baja_fuerte")
                    ),
                    "recall_sube_fuerte": horizon_row.get(
                        "recall_sube_fuerte", workbook_row.get("Recall_sube_fuerte")
                    ),
                    "recall_sube_total": horizon_row.get(
                        "recall_sube_total", workbook_row.get("Recall_sube_total")
                    ),
                    "accuracy_3clases": horizon_row.get("accuracy_3clases", workbook_row.get("Accuracy_3clases")),
                    "macro_f1_3clases": horizon_row.get("macro_f1_3clases", workbook_row.get("Macro_f1_3clases")),
                    "balanced_accuracy_3clases": horizon_row.get(
                        "balanced_accuracy_3clases", workbook_row.get("Balanced_accuracy_3clases")
                    ),
                    "loss_clasificacion_h": horizon_row.get(
                        "loss_clasificacion_h", workbook_row.get("Loss_clasificacion_h")
                    ),
                    "status_run": master_row.get("status_run"),
                    "L_total_Radar": master_row.get("L_total_Radar"),
                    "L_total_Clasificacion": master_row.get("L_total_Clasificacion"),
                }
            )

    long_df = pd.DataFrame(rows)
    rank_specs = (
        ("mae", True, "rank_mae_por_horizonte"),
        ("rmse", True, "rank_rmse_por_horizonte"),
        ("direction_accuracy", False, "rank_direction_accuracy_por_horizonte"),
        ("deteccion_caidas", False, "rank_deteccion_caidas_por_horizonte"),
        ("loss_h", True, "rank_loss_h_por_horizonte"),
    )
    for _, _, rank_col in rank_specs:
        long_df[rank_col] = pd.NA

    for horizon in (1, 2, 3, 4):
        horizon_mask = long_df["horizonte"] == horizon
        horizon_subset = long_df[horizon_mask]
        for metric_col, ascending, rank_col in rank_specs:
            metric_subset = horizon_subset.dropna(subset=[metric_col]).copy()
            if metric_subset.empty:
                continue
            metric_subset = sort_metric_subset(
                metric_subset,
                metric_col=metric_col,
                ascending=ascending,
            )
            rank_map = {
                row["run_id"]: rank
                for rank, (_, row) in enumerate(metric_subset.iterrows(), start=1)
            }
            long_df.loc[horizon_mask, rank_col] = long_df.loc[horizon_mask, "run_id"].map(rank_map)

    return long_df.sort_values(["horizonte", "rank_loss_h_por_horizonte", "run_id"], kind="mergesort").reset_index(drop=True)


def pick_column(columns: list[str], candidates: tuple[str, ...]) -> str | None:
    lower_map = {column.lower(): column for column in columns}
    for candidate in candidates:
        if candidate.lower() in lower_map:
            return lower_map[candidate.lower()]
    return None


def audit_prediction_file(run_id: str, horizon: int, pred_path: Path) -> PredictionAuditResult:
    base_row = {
        "run_id": run_id,
        "horizonte": horizon,
        "horizonte_label": f"H{horizon}",
        "pred_path": str(pred_path.resolve()),
        "existe_archivo": pred_path.exists(),
        "fecha_min_pred": None,
        "fecha_max_pred": None,
        "n_predicciones": 0,
        "n_fechas_unicas": 0,
        "sin_duplicados_fecha": False,
        "columnas_detectadas": "",
        "date_column": "",
        "y_true_column": "",
        "y_pred_column": "",
        "tiene_fecha": False,
        "tiene_y_true": False,
        "tiene_y_pred": False,
        "columnas_minimas_ok": False,
        "orden_temporal_ok": False,
        "compatible_para_merge": False,
        "comentario": "",
    }
    if not pred_path.exists():
        base_row["comentario"] = "archivo_no_encontrado"
        return PredictionAuditResult(base_row, None)

    try:
        raw_df = pd.read_csv(pred_path)
    except Exception as exc:  # pragma: no cover - defensive path
        base_row["comentario"] = f"error_lectura={exc}"
        return PredictionAuditResult(base_row, None)

    columns = list(raw_df.columns)
    base_row["columnas_detectadas"] = ",".join(columns)
    base_row["n_predicciones"] = int(len(raw_df))

    date_col = pick_column(columns, PREDICTION_DATE_CANDIDATES)
    y_true_col = pick_column(columns, PREDICTION_Y_TRUE_CANDIDATES)
    y_pred_col = pick_column(columns, PREDICTION_Y_PRED_CANDIDATES)
    base_row["date_column"] = date_col or ""
    base_row["y_true_column"] = y_true_col or ""
    base_row["y_pred_column"] = y_pred_col or ""
    base_row["tiene_fecha"] = bool(date_col)
    base_row["tiene_y_true"] = bool(y_true_col)
    base_row["tiene_y_pred"] = bool(y_pred_col)
    base_row["columnas_minimas_ok"] = bool(date_col and y_true_col and y_pred_col)

    if not base_row["columnas_minimas_ok"]:
        missing = []
        if not date_col:
            missing.append("fecha")
        if not y_true_col:
            missing.append("y_true")
        if not y_pred_col:
            missing.append("y_pred")
        base_row["comentario"] = f"columnas_faltantes={','.join(missing)}"
        return PredictionAuditResult(base_row, None)

    parsed_dates = pd.to_datetime(raw_df[date_col], errors="coerce")
    standardized_df = pd.DataFrame(
        {
            "fecha": parsed_dates,
            "y_true": pd.to_numeric(raw_df[y_true_col], errors="coerce"),
            "y_pred": pd.to_numeric(raw_df[y_pred_col], errors="coerce"),
        }
    )
    standardized_df = standardized_df.dropna(subset=["fecha", "y_true", "y_pred"]).copy()

    if standardized_df.empty:
        base_row["comentario"] = "sin_filas_validas_para_merge"
        return PredictionAuditResult(base_row, None)

    standardized_df["fecha"] = pd.to_datetime(standardized_df["fecha"])
    base_row["n_predicciones"] = int(len(standardized_df))
    base_row["n_fechas_unicas"] = int(standardized_df["fecha"].nunique())
    base_row["sin_duplicados_fecha"] = bool(base_row["n_predicciones"] == base_row["n_fechas_unicas"])
    base_row["orden_temporal_ok"] = bool(standardized_df["fecha"].is_monotonic_increasing)
    base_row["fecha_min_pred"] = standardized_df["fecha"].min().date().isoformat()
    base_row["fecha_max_pred"] = standardized_df["fecha"].max().date().isoformat()
    base_row["compatible_para_merge"] = bool(
        base_row["columnas_minimas_ok"]
        and base_row["n_predicciones"] > 0
        and base_row["sin_duplicados_fecha"]
    )
    base_row["comentario"] = (
        "ok"
        if base_row["compatible_para_merge"]
        else "inconsistencia_merge"
    )

    standardized_df = standardized_df.sort_values("fecha", kind="mergesort").reset_index(drop=True)
    standardized_df["fecha"] = standardized_df["fecha"].dt.date.astype(str)
    return PredictionAuditResult(base_row, standardized_df)


def build_prediction_coverage(
    canonical: dict[str, RunDirectoryRecord],
) -> tuple[pd.DataFrame, dict[tuple[str, int], pd.DataFrame]]:
    rows: list[dict[str, Any]] = []
    standardized_lookup: dict[tuple[str, int], pd.DataFrame] = {}
    for run_id in sorted(canonical):
        record = canonical[run_id]
        for horizon in (1, 2, 3, 4):
            pred_path = get_prediction_file(record, horizon)
            audit_result = audit_prediction_file(run_id, horizon, pred_path)
            rows.append(audit_result.coverage_row)
            if audit_result.standardized_df is not None:
                standardized_lookup[(run_id, horizon)] = audit_result.standardized_df
    coverage_df = pd.DataFrame(rows).sort_values(["run_id", "horizonte"]).reset_index(drop=True)
    return coverage_df, standardized_lookup


def build_stacking_readiness(
    runs_catalog_df: pd.DataFrame,
    coverage_df: pd.DataFrame,
    stacking_eligibility_df: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    coverage_lookup = coverage_df.groupby("run_id", sort=True)
    eligibility_lookup = stacking_eligibility_df.groupby("run_id", sort=True)

    for _, catalog_row in runs_catalog_df.iterrows():
        run_id = catalog_row["run_id"]
        coverage_subset = (
            coverage_lookup.get_group(run_id) if run_id in coverage_lookup.groups else pd.DataFrame()
        )
        eligibility_subset = (
            eligibility_lookup.get_group(run_id) if run_id in eligibility_lookup.groups else pd.DataFrame()
        )
        available_horizons = coverage_subset[coverage_subset["existe_archivo"]]["horizonte"].tolist()
        compatible_horizons = coverage_subset[coverage_subset["compatible_para_merge"]]["horizonte"].tolist()
        eligible_horizons = eligibility_subset[
            eligibility_subset["stacking_eligible_horizonte"] == True
        ]["horizonte"].tolist()
        predicciones_completas = set(available_horizons) == {1, 2, 3, 4}
        columnas_minimas_ok = bool(not coverage_subset.empty and coverage_subset["columnas_minimas_ok"].all())
        tiene_fecha = bool(not coverage_subset.empty and coverage_subset["tiene_fecha"].all())
        tiene_y_true = bool(not coverage_subset.empty and coverage_subset["tiene_y_true"].all())
        tiene_y_pred = bool(not coverage_subset.empty and coverage_subset["tiene_y_pred"].all())
        orden_temporal_ok = bool(not coverage_subset.empty and coverage_subset["orden_temporal_ok"].all())
        sin_duplicados = bool(not coverage_subset.empty and coverage_subset["sin_duplicados_fecha"].all())
        compatible_para_stacking = bool(set(eligible_horizons) == {1, 2, 3, 4})

        motivos_globales: list[str] = []
        for horizon in (1, 2, 3, 4):
            horizon_subset = eligibility_subset[eligibility_subset["horizonte"] == horizon]
            if horizon_subset.empty:
                motivos_globales.append(f"H{horizon}:sin_auditoria")
                continue
            horizon_reason = str(horizon_subset.iloc[0]["motivo_no_elegibilidad_horizonte"] or "")
            if horizon_reason:
                motivos_globales.append(f"H{horizon}:{horizon_reason}")

        compatible_count = len(eligible_horizons)
        if compatible_para_stacking:
            prioridad = "alta"
        elif compatible_count >= 2:
            prioridad = "media"
        elif compatible_count >= 1:
            prioridad = "baja"
        else:
            prioridad = "excluido"

        row = {
            "run_id": run_id,
            "family": catalog_row["family"],
            "model": catalog_row["model"],
            "identidad_canonica_clara": bool(catalog_row.get("es_run_maestro") and catalog_row.get("version_canonica")),
            "reconstruccion_hiperparams_status": catalog_row.get("reconstruccion_hiperparams_status"),
            "horizontes_disponibles": ",".join(str(h) for h in available_horizons),
            "horizontes_mergeables": ",".join(str(h) for h in compatible_horizons),
            "horizontes_elegibles_stacking": ",".join(str(h) for h in eligible_horizons),
            "predicciones_completas_1a4": predicciones_completas,
            "columnas_minimas_ok": columnas_minimas_ok,
            "tiene_fecha": tiene_fecha,
            "tiene_y_true": tiene_y_true,
            "tiene_y_pred": tiene_y_pred,
            "orden_temporal_ok": orden_temporal_ok,
            "sin_duplicados_fecha": sin_duplicados,
            "mergeable_h1": 1 in set(compatible_horizons),
            "mergeable_h2": 2 in set(compatible_horizons),
            "mergeable_h3": 3 in set(compatible_horizons),
            "mergeable_h4": 4 in set(compatible_horizons),
            "stacking_eligible_h1": 1 in set(eligible_horizons),
            "stacking_eligible_h2": 2 in set(eligible_horizons),
            "stacking_eligible_h3": 3 in set(eligible_horizons),
            "stacking_eligible_h4": 4 in set(eligible_horizons),
            "cantidad_obs_h1": int(coverage_subset.loc[coverage_subset["horizonte"] == 1, "n_predicciones"].max()) if 1 in set(coverage_subset["horizonte"]) else 0,
            "cantidad_obs_h2": int(coverage_subset.loc[coverage_subset["horizonte"] == 2, "n_predicciones"].max()) if 2 in set(coverage_subset["horizonte"]) else 0,
            "cantidad_obs_h3": int(coverage_subset.loc[coverage_subset["horizonte"] == 3, "n_predicciones"].max()) if 3 in set(coverage_subset["horizonte"]) else 0,
            "cantidad_obs_h4": int(coverage_subset.loc[coverage_subset["horizonte"] == 4, "n_predicciones"].max()) if 4 in set(coverage_subset["horizonte"]) else 0,
            "cobertura_total_predicciones": int(coverage_subset["n_predicciones"].sum()) if not coverage_subset.empty else 0,
            "compatible_para_stacking": compatible_para_stacking,
            "prioridad_para_meta_modelo": prioridad,
            "motivo_no_compatible": "; ".join([reason for reason in motivos_globales if reason]),
        }
        rows.append(row)

    return (
        pd.DataFrame(rows)
        .sort_values(["compatible_para_stacking", "run_id"], ascending=[False, True])
        .reset_index(drop=True)
    )


def build_stacking_eligibility_by_horizon(
    runs_catalog_df: pd.DataFrame,
    coverage_df: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    coverage_map = {
        (row["run_id"], int(row["horizonte"])): row
        for row in coverage_df.to_dict(orient="records")
    }
    for _, catalog_row in runs_catalog_df.iterrows():
        for horizon in (1, 2, 3, 4):
            coverage_row = coverage_map.get((catalog_row["run_id"], horizon), {})
            motives: list[str] = []
            identidad_canonica_clara = bool(
                catalog_row.get("es_run_maestro") and catalog_row.get("version_canonica")
            )
            configuracion_recuperable = catalog_row.get("reconstruccion_hiperparams_status") in {"completo", "parcial"}
            metrica_trazable = pd.notna(catalog_row.get(f"loss_h{horizon}"))
            task_type_ok = catalog_row.get("task_type", "regresion") == "regresion"
            status_ok = catalog_row.get("status_run") == "completo"
            existe_archivo = bool(coverage_row.get("existe_archivo", False))
            mergeable = bool(coverage_row.get("compatible_para_merge", False))
            columnas_minimas_ok = bool(coverage_row.get("columnas_minimas_ok", False))
            if not identidad_canonica_clara:
                motives.append("identidad_canonica_no_clara")
            if not task_type_ok:
                motives.append(f"task_type={catalog_row.get('task_type')}")
            if not status_ok:
                motives.append(f"status_run={catalog_row.get('status_run')}")
            if not configuracion_recuperable:
                motives.append("configuracion_no_recuperable")
            if not metrica_trazable:
                motives.append("sin_loss_h")
            if not existe_archivo:
                motives.append("sin_predicciones")
            if existe_archivo and not columnas_minimas_ok:
                motives.append("columnas_minimas_invalidas")
            if existe_archivo and not mergeable:
                motives.append("predicciones_no_mergeables")
            stacking_eligible = bool(
                identidad_canonica_clara
                and task_type_ok
                and status_ok
                and configuracion_recuperable
                and metrica_trazable
                and existe_archivo
                and columnas_minimas_ok
                and mergeable
            )
            rows.append(
                {
                    "run_id": catalog_row["run_id"],
                    "family": catalog_row["family"],
                    "model": catalog_row["model"],
                    "task_type": catalog_row.get("task_type"),
                    "horizonte": horizon,
                    "horizonte_label": f"H{horizon}",
                    "identidad_canonica_clara": identidad_canonica_clara,
                    "status_run": catalog_row.get("status_run"),
                    "reconstruccion_hiperparams_status": catalog_row.get("reconstruccion_hiperparams_status"),
                    "configuracion_recuperable": configuracion_recuperable,
                    "tiene_metrica_h": metrica_trazable,
                    "pred_path": coverage_row.get("pred_path", ""),
                    "tiene_predicciones_h": existe_archivo,
                    "mergeable_h": mergeable,
                    "filas_pred_h": int(coverage_row.get("n_predicciones", 0) or 0),
                    "stacking_eligible_horizonte": stacking_eligible,
                    "motivo_no_elegibilidad_horizonte": ";".join(motives),
                }
            )

    return (
        pd.DataFrame(rows)
        .sort_values(["horizonte", "stacking_eligible_horizonte", "run_id"], ascending=[True, False, True])
        .reset_index(drop=True)
    )


def build_stacking_base_sheet(
    *,
    horizon: int,
    stacking_eligibility_df: pd.DataFrame,
    standardized_lookup: dict[tuple[str, int], pd.DataFrame],
) -> pd.DataFrame:
    eligible = stacking_eligibility_df[
        (stacking_eligibility_df["horizonte"] == horizon)
        & (stacking_eligibility_df["stacking_eligible_horizonte"] == True)
    ].copy()
    eligible = eligible.sort_values("run_id").reset_index(drop=True)

    if eligible.empty:
        return pd.DataFrame(
            columns=[
                "fecha",
                "y_true",
                "n_modelos_disponibles_fila",
                "fila_completa_todos_modelos",
                "cobertura_modelos_fila",
            ]
        )

    merged_df: pd.DataFrame | None = None
    y_true_columns: list[str] = []
    pred_columns: list[str] = []

    for _, row in eligible.iterrows():
        run_id = row["run_id"]
        standardized_df = standardized_lookup.get((run_id, horizon))
        if standardized_df is None or standardized_df.empty:
            continue
        model_df = standardized_df.rename(
            columns={
                "y_true": f"__y_true_{run_id}",
                "y_pred": run_id,
            }
        )
        y_true_columns.append(f"__y_true_{run_id}")
        pred_columns.append(run_id)
        if merged_df is None:
            merged_df = model_df
        else:
            merged_df = merged_df.merge(model_df, on="fecha", how="outer", sort=True)

    if merged_df is None or merged_df.empty:
        return pd.DataFrame(columns=["fecha", "y_true", "n_modelos_disponibles_fila"])

    merged_df = merged_df.sort_values("fecha", kind="mergesort").reset_index(drop=True)
    if y_true_columns:
        merged_df["y_true"] = merged_df[y_true_columns].bfill(axis=1).iloc[:, 0]
        merged_df = merged_df.drop(columns=y_true_columns)
    else:
        merged_df["y_true"] = pd.NA
    ordered_columns = ["fecha", "y_true", *pred_columns]
    merged_df = merged_df[ordered_columns]
    merged_df["n_modelos_disponibles_fila"] = merged_df[pred_columns].notna().sum(axis=1) if pred_columns else 0
    total_models = len(pred_columns)
    merged_df["fila_completa_todos_modelos"] = (
        merged_df["n_modelos_disponibles_fila"] == total_models if total_models else False
    )
    merged_df["cobertura_modelos_fila"] = (
        merged_df["n_modelos_disponibles_fila"] / total_models if total_models else 0.0
    )
    return merged_df


def build_stacking_base_summary(
    *,
    stacking_eligibility_df: pd.DataFrame,
    stacking_base_sheets: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for horizon in (1, 2, 3, 4):
        eligibility_subset = stacking_eligibility_df[stacking_eligibility_df["horizonte"] == horizon].copy()
        included = eligibility_subset[eligibility_subset["stacking_eligible_horizonte"] == True].sort_values("run_id")
        excluded = eligibility_subset[eligibility_subset["stacking_eligible_horizonte"] != True].sort_values("run_id")
        base_df = stacking_base_sheets.get(f"stacking_base_h{horizon}", pd.DataFrame())
        rows.append(
            {
                "horizonte": f"H{horizon}",
                "run_ids_incluidos": ",".join(included["run_id"].tolist()),
                "n_runs_incluidos": int(len(included)),
                "run_ids_excluidos": ",".join(excluded["run_id"].tolist()),
                "exclusiones_detalle": " | ".join(
                    f"{row['run_id']}:{row['motivo_no_elegibilidad_horizonte']}"
                    for _, row in excluded.iterrows()
                    if row["motivo_no_elegibilidad_horizonte"]
                ),
                "filas_base": int(len(base_df)),
                "filas_completas_todos_modelos": int(base_df["fila_completa_todos_modelos"].sum())
                if "fila_completa_todos_modelos" in base_df.columns
                else 0,
                "filas_incompletas": int(
                    len(base_df) - int(base_df["fila_completa_todos_modelos"].sum())
                )
                if "fila_completa_todos_modelos" in base_df.columns
                else int(len(base_df)),
                "cobertura_promedio_modelos_fila": float(base_df["cobertura_modelos_fila"].mean())
                if "cobertura_modelos_fila" in base_df.columns and not base_df.empty
                else 0.0,
            }
        )
    return pd.DataFrame(rows)


def extract_experiment_family(run_id: str) -> str:
    match = re.match(r"^([A-Z]\d+)", str(run_id))
    return match.group(1) if match else str(run_id)


def invert_horizon_candidate_mapping(mapping: dict[int, list[str]]) -> dict[str, list[int]]:
    inverted: dict[str, list[int]] = {}
    for horizon, run_ids in mapping.items():
        for run_id in run_ids:
            inverted.setdefault(run_id, []).append(int(horizon))
    return {run_id: sorted(horizons) for run_id, horizons in inverted.items()}


def build_regression_family_champions(runs_catalog_df: pd.DataFrame) -> dict[str, str]:
    regression_df = runs_catalog_df[
        (runs_catalog_df["task_type"].fillna("regresion") == "regresion")
        & (runs_catalog_df["status_run"] == "completo")
    ].copy()
    if regression_df.empty:
        return {}
    regression_df["familia_experimental"] = regression_df["run_id"].map(extract_experiment_family)
    champions = (
        regression_df.sort_values(["familia_experimental", "L_total_Radar", "run_id"], kind="mergesort")
        .drop_duplicates("familia_experimental", keep="first")
    )
    return dict(zip(champions["familia_experimental"], champions["run_id"]))


def infer_e9_default_note(
    *,
    run_id: str,
    row: pd.Series,
    family_champions: dict[str, str],
) -> tuple[str, str]:
    if row.get("task_type", "regresion") != "regresion":
        return (
            "run no preferible para ensemble principal",
            "E9 se restringe a regresion; las corridas de clasificacion quedan fuera por comparabilidad.",
        )

    eligible_any = any(bool(row.get(f"stacking_eligible_h{horizon}", False)) for horizon in (1, 2, 3, 4))
    if not eligible_any:
        reason = str(row.get("motivo_no_elegibilidad", "") or "sin elegibilidad estructural para stacking")
        return (
            "run no preferible para ensemble principal",
            f"Queda fuera del universo E9 porque no cumple elegibilidad estructural suficiente: {reason}.",
        )

    familia_experimental = extract_experiment_family(run_id)
    family_champion = family_champions.get(familia_experimental)
    if family_champion and family_champion != run_id:
        return (
            "run dominado",
            f"Dominado dentro de {familia_experimental} por {family_champion}; no agrega diversidad real suficiente frente al campeon de su familia.",
        )

    return (
        "run no preferible para ensemble principal",
        "No entra al ensemble principal porque su senal no compensa la redundancia, la complejidad o la debilidad relativa frente al nucleo seleccionado.",
    )


def build_e9_construct_dictionary(runs_catalog_df: pd.DataFrame) -> pd.DataFrame:
    final_horizon_map = invert_horizon_candidate_mapping(E9_FINAL_CANDIDATES_BY_HORIZON)
    reserve_horizon_map = invert_horizon_candidate_mapping(E9_RESERVE_CANDIDATES_BY_HORIZON)
    family_champions = build_regression_family_champions(runs_catalog_df)

    rows: list[dict[str, Any]] = []
    for _, row in runs_catalog_df.sort_values("run_id", kind="mergesort").iterrows():
        run_id = str(row["run_id"])
        familia_experimental = extract_experiment_family(run_id)
        final_horizons = final_horizon_map.get(run_id, [])
        reserve_horizons = reserve_horizon_map.get(run_id, [])
        manual_note = E9_PRIORITY_RUN_NOTES.get(run_id)
        if manual_note:
            constructo = manual_note["constructo"]
            justificacion = manual_note["justificacion"]
        else:
            constructo, justificacion = infer_e9_default_note(
                run_id=run_id,
                row=row,
                family_champions=family_champions,
            )

        entra = bool(final_horizons)
        reserva = bool(reserve_horizons)
        decision_global = "entra" if entra else "reserva" if reserva else "sale"
        motivo_exclusion = "" if (entra or reserva) else justificacion

        rows.append(
            {
                "run_id": run_id,
                "familia": familia_experimental,
                "categoria_modelo": row.get("family"),
                "modelo": row.get("model"),
                "estatus_canonico": row.get("status_canonico"),
                "stacking_eligible_global": bool(row.get("stacking_eligible_global", False)),
                "horizontes_entrada": ",".join(f"H{h}" for h in final_horizons),
                "horizontes_reserva": ",".join(f"H{h}" for h in reserve_horizons),
                "constructo_en_E9": constructo,
                "justificacion_breve": justificacion,
                "entra_a_E9": "si" if entra else "no",
                "reserva": "si" if reserva else "no",
                "decision_global_E9": decision_global,
                "motivo_exclusion": motivo_exclusion,
            }
        )

    return pd.DataFrame(rows)


def compute_e9_coverage_ratios(runs_catalog_df: pd.DataFrame) -> dict[int, int]:
    regression_df = runs_catalog_df[runs_catalog_df["task_type"].fillna("regresion") == "regresion"].copy()
    result: dict[int, int] = {}
    for horizon in (1, 2, 3, 4):
        result[horizon] = int(regression_df[f"filas_pred_h{horizon}"].fillna(0).max()) if not regression_df.empty else 0
    return result


def build_e9_metricas_candidatos(
    runs_catalog_df: pd.DataFrame,
    construct_df: pd.DataFrame,
) -> pd.DataFrame:
    construct_lookup = construct_df.set_index("run_id").to_dict(orient="index")
    coverage_max = compute_e9_coverage_ratios(runs_catalog_df)
    regression_df = runs_catalog_df[runs_catalog_df["task_type"].fillna("regresion") == "regresion"].copy()

    rows: list[dict[str, Any]] = []
    for _, row in regression_df.iterrows():
        run_id = str(row["run_id"])
        construct_info = construct_lookup.get(run_id, {})
        coverage_ratios = []
        for horizon in (1, 2, 3, 4):
            max_rows = coverage_max.get(horizon, 0)
            current_rows = int(row.get(f"filas_pred_h{horizon}", 0) or 0)
            if max_rows > 0:
                coverage_ratios.append(current_rows / max_rows)

        rows.append(
            {
                "run_id": run_id,
                "familia": extract_experiment_family(run_id),
                "categoria_modelo": row.get("family"),
                "modelo": row.get("model"),
                "target_mode": row.get("target_mode"),
                "feature_mode": row.get("feature_mode"),
                "transform_mode": row.get("transform_mode"),
                "lags": row.get("lags"),
                "mae_promedio": row.get("mae_promedio"),
                "rmse_promedio": row.get("rmse_promedio"),
                "direction_accuracy_promedio": row.get("direction_accuracy_promedio"),
                "deteccion_caidas_promedio": row.get("deteccion_caidas_promedio"),
                "L_total_Radar": row.get("L_total_Radar"),
                "loss_h1": row.get("loss_h1"),
                "loss_h2": row.get("loss_h2"),
                "loss_h3": row.get("loss_h3"),
                "loss_h4": row.get("loss_h4"),
                "cobertura_media": float(sum(coverage_ratios) / len(coverage_ratios)) if coverage_ratios else 0.0,
                "horizontes_completos": int(
                    sum(int(bool(row.get(f"stacking_eligible_h{horizon}", False))) for horizon in (1, 2, 3, 4))
                ),
                "elegible_E9": "si"
                if construct_info.get("decision_global_E9") in {"entra", "reserva"}
                else "no",
                "constructo_en_E9": construct_info.get("constructo_en_E9", ""),
                "decision_global_E9": construct_info.get("decision_global_E9", "sale"),
                "observacion_breve": row.get("observacion_breve"),
            }
        )

    result = pd.DataFrame(rows)
    if result.empty:
        return result
    decision_rank = {"entra": 0, "reserva": 1, "sale": 2}
    result["__decision_rank"] = result["decision_global_E9"].map(decision_rank).fillna(9)
    result = result.sort_values(
        ["__decision_rank", "L_total_Radar", "run_id"],
        ascending=[True, True, True],
        kind="mergesort",
    ).drop(columns="__decision_rank")
    return result.reset_index(drop=True)


def build_e9_curated_base_sheet(
    *,
    horizon: int,
    stacking_base_sheet: pd.DataFrame,
) -> pd.DataFrame:
    selected_runs = E9_FINAL_CANDIDATES_BY_HORIZON[horizon]
    available_runs = [run_id for run_id in selected_runs if run_id in stacking_base_sheet.columns]
    if stacking_base_sheet.empty or not available_runs:
        return pd.DataFrame(columns=["fecha", "y_true", "fila_completa", "n_modelos_disponibles"])

    base_df = stacking_base_sheet[["fecha", "y_true", *available_runs]].copy()
    base_df["n_modelos_disponibles"] = base_df[available_runs].notna().sum(axis=1)
    base_df["fila_completa"] = base_df["n_modelos_disponibles"] == len(available_runs)
    base_df["cobertura_modelos_fila"] = base_df["n_modelos_disponibles"] / len(available_runs)
    return base_df


def describe_redundancy_against_finals(
    *,
    raw_base_sheet: pd.DataFrame,
    horizon: int,
    run_id: str,
) -> tuple[float | None, str, str]:
    final_runs = [candidate for candidate in E9_FINAL_CANDIDATES_BY_HORIZON[horizon] if candidate in raw_base_sheet.columns]
    if run_id not in raw_base_sheet.columns:
        return None, "", "sin_columna_de_prediccion_en_base_raw"

    comparison_runs = [candidate for candidate in final_runs if candidate != run_id]
    if not comparison_runs:
        return None, "", "sin_otro_principal_para_comparar"

    corr_series = raw_base_sheet[[run_id, *comparison_runs]].corr(min_periods=10)[run_id].drop(run_id)
    corr_series = corr_series.dropna()
    if corr_series.empty:
        return None, "", "sin_correlacion_estimable_con_principales"

    closest_run = str(corr_series.abs().idxmax())
    closest_value = float(corr_series.loc[closest_run])
    if abs(closest_value) >= 0.995:
        note = f"predicciones casi duplicadas de {closest_run} (corr={closest_value:.3f})"
    elif abs(closest_value) >= 0.90:
        note = f"redundancia alta con {closest_run} (corr={closest_value:.3f})"
    elif abs(closest_value) >= 0.75:
        note = f"solapamiento moderado con {closest_run} (corr={closest_value:.3f})"
    else:
        note = f"diversidad util frente a {closest_run} (corr={closest_value:.3f})"
    return closest_value, closest_run, note


def build_e9_diagnostic_sheet(
    *,
    horizon: int,
    master_df: pd.DataFrame,
    construct_df: pd.DataFrame,
    stacking_base_sheet: pd.DataFrame,
    e9_base_sheet: pd.DataFrame,
) -> pd.DataFrame:
    horizon_label = f"H{horizon}"
    regression_df = master_df[
        (master_df["task_type"].fillna("regresion") == "regresion")
        & (master_df[f"stacking_eligible_h{horizon}"] == True)
    ].copy()
    construct_lookup = construct_df.set_index("run_id").to_dict(orient="index")
    max_rows = int(regression_df[f"filas_pred_h{horizon}"].fillna(0).max()) if not regression_df.empty else 0
    final_runs = set(E9_FINAL_CANDIDATES_BY_HORIZON[horizon])
    reserve_runs = set(E9_RESERVE_CANDIDATES_BY_HORIZON[horizon])

    rows: list[dict[str, Any]] = []
    for _, row in regression_df.iterrows():
        run_id = str(row["run_id"])
        construct_info = construct_lookup.get(run_id, {})
        reserve_horizons_text = str(construct_info.get("horizontes_reserva", "") or "")
        if run_id in final_runs:
            decision = "entra"
        elif run_id in reserve_runs:
            decision = "reserva"
        else:
            decision = "sale"

        corr_value, closest_run, redundancy_note = describe_redundancy_against_finals(
            raw_base_sheet=stacking_base_sheet,
            horizon=horizon,
            run_id=run_id,
        )

        rows.append(
            {
                "run_id": run_id,
                "familia": extract_experiment_family(run_id),
                "categoria_modelo": row.get("family"),
                "modelo": row.get("model"),
                "constructo_en_E9": construct_info.get("constructo_en_E9", ""),
                "decision_sugerida": decision,
                "loss_h": row.get(f"loss_h{horizon}"),
                "mae_h": row.get(f"H{horizon}_mae"),
                "rmse_h": row.get(f"H{horizon}_rmse"),
                "direction_accuracy_h": row.get(f"H{horizon}_direction_accuracy"),
                "deteccion_caidas_h": row.get(f"H{horizon}_deteccion_caidas"),
                "filas_pred_h": int(row.get(f"filas_pred_h{horizon}", 0) or 0),
                "cobertura_relativa_h": (
                    float((row.get(f"filas_pred_h{horizon}", 0) or 0) / max_rows) if max_rows else 0.0
                ),
                "run_principal_mas_parecido": closest_run,
                "correlacion_abs_max_con_principales": abs(corr_value) if corr_value is not None else None,
                "observacion_redundancia": redundancy_note,
                "filas_base_final": int(len(e9_base_sheet)),
                "filas_completas_base_final": int(e9_base_sheet["fila_completa"].sum())
                if "fila_completa" in e9_base_sheet.columns
                else 0,
                "motivo_decision": (
                    construct_info.get("justificacion_breve")
                    if decision in {"entra", "reserva"}
                    else (
                        f"Se conserva solo como reserva en {reserve_horizons_text}; no entra en {horizon_label}."
                        if construct_info.get("decision_global_E9") == "reserva" and reserve_horizons_text
                        else construct_info.get("motivo_exclusion", "")
                    )
                ),
            }
        )

    result = pd.DataFrame(rows)
    if result.empty:
        return result
    decision_rank = {"entra": 0, "reserva": 1, "sale": 2}
    result["__decision_rank"] = result["decision_sugerida"].map(decision_rank).fillna(9)
    result = result.sort_values(
        ["__decision_rank", "loss_h", "run_id"],
        ascending=[True, True, True],
        kind="mergesort",
    ).drop(columns="__decision_rank")
    return result.reset_index(drop=True)


def build_e9_bases_summary(
    e9_base_sheets: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for horizon in (1, 2, 3, 4):
        base_df = e9_base_sheets.get(f"E9_base_h{horizon}", pd.DataFrame())
        candidate_runs = E9_FINAL_CANDIDATES_BY_HORIZON[horizon]
        reserve_runs = E9_RESERVE_CANDIDATES_BY_HORIZON[horizon]
        model_columns = [column for column in base_df.columns if column in candidate_runs]
        coverage_per_model = [
            float(base_df[column].notna().mean()) for column in model_columns
        ] if model_columns else []

        rows.append(
            {
                "horizonte": f"H{horizon}",
                "candidatos_finales": ",".join(candidate_runs),
                "candidatos_reserva": ",".join(reserve_runs),
                "n_candidatos_finales": len(candidate_runs),
                "n_candidatos_reserva": len(reserve_runs),
                "filas_totales": int(len(base_df)),
                "filas_completas": int(base_df["fila_completa"].sum()) if "fila_completa" in base_df.columns else 0,
                "filas_incompletas": int(len(base_df) - int(base_df["fila_completa"].sum()))
                if "fila_completa" in base_df.columns
                else int(len(base_df)),
                "cobertura_media_fila": float(base_df["cobertura_modelos_fila"].mean())
                if "cobertura_modelos_fila" in base_df.columns and not base_df.empty
                else 0.0,
                "cobertura_media_por_modelo": float(sum(coverage_per_model) / len(coverage_per_model))
                if coverage_per_model
                else 0.0,
                "base_lista_para_E9": bool(
                    len(candidate_runs) >= 3
                    and "fila_completa" in base_df.columns
                    and int(base_df["fila_completa"].sum()) > 0
                ),
                "nota_metodologica": "Base curada pequena, con diversidad util y reglas ex ante; lista como insumo de stacking controlado por horizonte.",
            }
        )
    return pd.DataFrame(rows)


def build_family_status_vigente_df() -> pd.DataFrame:
    return pd.DataFrame(FAMILY_STATUS_VIGENTE_ROWS)


def infer_family_code_from_run_id(run_id: str) -> str:
    if run_id.startswith("E1_1_"):
        return "E1.1"
    match = re.match(r"^(E\d+|C\d+)", run_id)
    return match.group(1) if match else ""


def build_family_constructs_df() -> pd.DataFrame:
    family_status_df = build_family_status_vigente_df().copy()
    family_status_df["constructo_familia"] = family_status_df["family"].map(FAMILY_CANONICAL_CONSTRUCTS).fillna(
        family_status_df["rol_funcional"]
    )
    return family_status_df[
        [
            "family",
            "constructo_familia",
            "estado_vigente",
            "mejor_run",
            "rol_funcional",
            "decision_metodologica",
            "siguiente_movimiento",
        ]
    ].copy()


def build_run_constructs_df(master_df: pd.DataFrame, family_status_df: pd.DataFrame) -> pd.DataFrame:
    family_lookup = family_status_df.set_index("family").to_dict(orient="index")
    rows: list[dict[str, Any]] = []
    for _, row in master_df.sort_values("run_id").iterrows():
        run_id = str(row.get("run_id", ""))
        family_code = infer_family_code_from_run_id(run_id)
        family_info = family_lookup.get(family_code, {})
        manual = RUN_CANONICAL_CONSTRUCTS.get(run_id)
        if manual:
            constructo = manual["constructo"]
            justificacion = manual["justificacion"]
        elif str(row.get("status_canonico", "")) != "canonico_completo":
            constructo = "run tecnico no canonico"
            justificacion = "Run retenido solo por trazabilidad; no debe leerse como evidencia comparativa principal."
        else:
            constructo = "run documentado"
            justificacion = "Run registrado con trazabilidad suficiente, pero sin un constructo especifico adicional."
        rows.append(
            {
                "run_id": run_id,
                "family_code": family_code,
                "family_group": row.get("family"),
                "model": row.get("model"),
                "status_canonico": row.get("status_canonico"),
                "L_total_Radar": row.get("L_total_Radar"),
                "estado_familia_vigente": family_info.get("estado_vigente", ""),
                "constructo_run": constructo,
                "rol_funcional_familia": family_info.get("rol_funcional", ""),
                "justificacion_breve": justificacion,
            }
        )
    return pd.DataFrame(rows)


def build_explainability_status_df(runs_catalog_df: pd.DataFrame) -> pd.DataFrame:
    total_runs = int(len(runs_catalog_df))
    with_selected_features = int(runs_catalog_df["features_artifact_available"].fillna(False).sum())
    with_coefficients = int(runs_catalog_df["tiene_coeficientes"].fillna(False).sum())
    with_importances = int(runs_catalog_df["tiene_importancias"].fillna(False).sum())
    with_shap = int(runs_catalog_df["tiene_shap_o_equivalente"].fillna(False).sum())
    comparable_homogeneous = int(
        runs_catalog_df["explicabilidad_transversal_homogenea"].fillna(False).sum()
    )

    rows = [
        {
            "dimension": "clasificacion_principal",
            "valor": "parcial intra-familia",
            "justificacion": (
                "Existen artefactos parciales de seleccion de variables por horizonte en una fraccion de runs, "
                "pero no hay coeficientes, importancias o SHAP exportados de forma homogénea y comparable entre familias."
            ),
        },
        {
            "dimension": "runs_con_features_seleccionadas",
            "valor": with_selected_features,
            "justificacion": (
                f"{with_selected_features} de {total_runs} runs maestros tienen archivos "
                "`features_seleccionadas_h*.csv`; esto sirve para lectura intra-run o intra-familia, no para comparacion transversal canonica."
            ),
        },
        {
            "dimension": "runs_con_coeficientes_exportados",
            "valor": with_coefficients,
            "justificacion": "No se localizaron artefactos de coeficientes exportados de forma sistematica.",
        },
        {
            "dimension": "runs_con_importancias_exportadas",
            "valor": with_importances,
            "justificacion": "No se localizaron artefactos de importancias exportadas de forma sistematica.",
        },
        {
            "dimension": "runs_con_shap_exportado",
            "valor": with_shap,
            "justificacion": "No se localizaron artefactos SHAP o equivalentes reutilizables de forma sistematica.",
        },
        {
            "dimension": "runs_marcados_como_homogeneos",
            "valor": comparable_homogeneous,
            "justificacion": (
                "Ningun run puede considerarse parte de una capa explicativa transversal canonica solo con la evidencia actual."
            ),
        },
        {
            "dimension": "lectura_metodologica_vigente",
            "valor": "performance_y_riesgo_siguen_siendo_la_base_de_decision",
            "justificacion": (
                "La eleccion de modelos sigue descansando principalmente en performance, direccion y deteccion de caidas; "
                "la interpretabilidad actual es fragmentaria y no resuelve comparabilidad inter-familias."
            ),
        },
        {
            "dimension": "faltante_para_capa_transversal_canonica",
            "valor": "exportar_artefactos_homogeneos_por_familia_y_horizonte",
            "justificacion": (
                "Haria falta exportar de forma sistematica, por horizonte y con taxonomia comun, coeficientes o importancias "
                "comparables entre lineales, arboles, boosting y familias temporales."
            ),
        },
    ]
    return pd.DataFrame(rows)


def build_operational_benchmarks_df(master_df: pd.DataFrame) -> pd.DataFrame:
    master_lookup = master_df.set_index("run_id").to_dict(orient="index")
    rows: list[dict[str, Any]] = []
    for spec in OPERATIONAL_BENCHMARK_ROWS:
        run_id = spec["run_id"]
        run_info = master_lookup.get(run_id, {})
        run_dir = str(run_info.get("run_dir") or "")
        parametros_fuente = f"{run_dir}/parametros_run.json" if run_dir else ""
        rows.append(
            {
                "benchmark_id": spec["benchmark_id"],
                "plano_funcional": spec["plano_funcional"],
                "run_id": run_id,
                "family": run_info.get("family"),
                "model": run_info.get("model"),
                "estado_operativo": spec["estado_operativo"],
                "rol_operativo": spec["rol_operativo"],
                "L_total_Radar": run_info.get("L_total_Radar"),
                "mae_promedio": run_info.get("mae_promedio"),
                "rmse_promedio": run_info.get("rmse_promedio"),
                "direction_accuracy_promedio": run_info.get("direction_accuracy_promedio"),
                "deteccion_caidas_promedio": run_info.get("deteccion_caidas_promedio"),
                "script_nombre": run_info.get("script_nombre"),
                "script_ruta": run_info.get("script_ruta"),
                "runner_script_canonico": run_info.get("script_ruta"),
                "run_dir": run_info.get("run_dir"),
                "parametros_fuente": parametros_fuente,
                "status_canonico": run_info.get("status_canonico"),
                "justificacion": spec["justificacion"],
            }
        )
    return pd.DataFrame(rows)


def build_radar_layers_df() -> pd.DataFrame:
    return pd.DataFrame(RADAR_LAYER_ROWS)


def build_dual_functional_metrics_df() -> pd.DataFrame:
    return pd.DataFrame(DUAL_FUNCTIONAL_METRIC_ROWS)


def build_dual_policy_summary_df() -> pd.DataFrame:
    return pd.DataFrame(DUAL_POLICY_SUMMARY_ROWS)


def build_e9_curacion_resumen_df(
    *,
    source_xlsx_path: Path,
    curated_xlsx_path: Path,
) -> pd.DataFrame:
    rows: list[dict[str, str]] = [
        {
            "seccion": "objetivo",
            "campo": "descripcion",
            "valor": "Curar una base pequena, trazable y metodologicamente defendible para E9 stacking clasico controlado por horizonte.",
        },
        {
            "seccion": "fuente",
            "campo": "archivo_fuente",
            "valor": str(source_xlsx_path),
        },
        {
            "seccion": "salida",
            "campo": "archivo_curado",
            "valor": str(curated_xlsx_path),
        },
        {
            "seccion": "reglas",
            "campo": "criterio_principal",
            "valor": "Priorizar runs canonicos, OOF mergeables, diversidad real entre familias y base pequena por horizonte.",
        },
        {
            "seccion": "reglas",
            "campo": "regla_exclusion",
            "valor": "Excluir corridas no canonicas, dominadas, colapsadas, redundantes o sin valor adicional claro para el ensemble principal.",
        },
        {
            "seccion": "reglas",
            "campo": "tamano_objetivo",
            "valor": "Ideal 3-5 modelos por horizonte; maximo razonable 6.",
        },
        {
            "seccion": "advertencias",
            "campo": "no_es_E10",
            "valor": "E9 en esta etapa no es gating contextual ni mezcla oportunista ex post; es stacking clasico controlado por horizonte.",
        },
        {
            "seccion": "uso_esperado",
            "campo": "siguiente_paso",
            "valor": "Usar E9_base_h1..h4 como insumo curado para adaptar y correr run_e9_stacking.py sin leakage.",
        },
    ]

    for constructo, definition in E9_CONSTRUCT_DEFINITIONS.items():
        rows.append(
            {
                "seccion": "constructos",
                "campo": constructo,
                "valor": definition,
            }
        )

    common_core = sorted(
        set(E9_FINAL_CANDIDATES_BY_HORIZON[1])
        & set(E9_FINAL_CANDIDATES_BY_HORIZON[2])
        & set(E9_FINAL_CANDIDATES_BY_HORIZON[3])
        & set(E9_FINAL_CANDIDATES_BY_HORIZON[4])
    )
    rows.append(
        {
            "seccion": "reglas",
            "campo": "nucleo_comun",
            "valor": ",".join(common_core),
        }
    )

    for horizon in (1, 2, 3, 4):
        rows.extend(
            [
                {
                    "seccion": f"H{horizon}",
                    "campo": "candidatos_finales",
                    "valor": ",".join(E9_FINAL_CANDIDATES_BY_HORIZON[horizon]),
                },
                {
                    "seccion": f"H{horizon}",
                    "campo": "candidatos_reserva",
                    "valor": ",".join(E9_RESERVE_CANDIDATES_BY_HORIZON[horizon]),
                },
            ]
        )

    return pd.DataFrame(rows)


def build_e9_markdown_report(
    *,
    source_xlsx_path: Path,
    curated_xlsx_path: Path,
    construct_df: pd.DataFrame,
    e9_bases_summary_df: pd.DataFrame,
) -> str:
    final_runs_by_horizon = {
        f"H{horizon}": ", ".join(E9_FINAL_CANDIDATES_BY_HORIZON[horizon]) for horizon in (1, 2, 3, 4)
    }
    reserve_runs = (
        construct_df[construct_df["decision_global_E9"] == "reserva"][
            ["run_id", "constructo_en_E9", "justificacion_breve"]
        ]
        .sort_values("run_id", kind="mergesort")
        .reset_index(drop=True)
    )
    excluded_runs = (
        construct_df[construct_df["decision_global_E9"] == "sale"][
            ["run_id", "constructo_en_E9", "motivo_exclusion"]
        ]
        .sort_values("run_id", kind="mergesort")
        .reset_index(drop=True)
    )
    complete_rows_lookup = {
        row["horizonte"]: int(row["filas_completas"])
        for row in e9_bases_summary_df.to_dict(orient="records")
    }

    return f"""# Preparacion de tabla E9 stacking controlado

## 1. Archivo base

- Fuente original: `{source_xlsx_path}`
- Salida curada: `{curated_xlsx_path}`

## 2. Hojas creadas o modificadas

- `E9_curacion_resumen`
- `E9_diccionario_constructos`
- `E9_metricas_candidatos`
- `E9_diagnostico_h1`
- `E9_diagnostico_h2`
- `E9_diagnostico_h3`
- `E9_diagnostico_h4`
- `E9_base_h1`
- `E9_base_h2`
- `E9_base_h3`
- `E9_base_h4`
- `E9_bases_resumen`

## 3. Criterio de curacion

- Universo base: tabla maestra completa, sin borrar materia prima.
- Universo elegible: solo runs canonicos de regresion con predicciones OOF mergeables y trazabilidad suficiente.
- Base final E9: 4 candidatos por horizonte, con nucleo comun `E1_v5_clean` + `E5_v4_clean` y rotacion controlada de diversidad (`E3_v2_clean`, `E2_v3_clean`, `E7_v3_clean`).
- Regla de exclusion: sacar variantes dominadas, familias debilitadas, corridas colapsadas y duplicados sin diversidad real.

## 4. Constructos usados

{render_dataframe_markdown(construct_df[['run_id', 'constructo_en_E9', 'decision_global_E9', 'justificacion_breve']], 'Sin constructos.')}

## 5. Runs que entran a E9 por horizonte

- H1: {final_runs_by_horizon['H1']}
- H2: {final_runs_by_horizon['H2']}
- H3: {final_runs_by_horizon['H3']}
- H4: {final_runs_by_horizon['H4']}

## 6. Runs que quedan como reserva

{render_dataframe_markdown(reserve_runs, 'Sin reservas documentadas.')}

## 7. Runs excluidos y por que

{render_dataframe_markdown(excluded_runs, 'Sin exclusiones documentadas.')}

## 8. Filas completas por horizonte

- H1: {complete_rows_lookup.get('H1', 0)}
- H2: {complete_rows_lookup.get('H2', 0)}
- H3: {complete_rows_lookup.get('H3', 0)}
- H4: {complete_rows_lookup.get('H4', 0)}

## 9. Estado de la base para run_e9_stacking.py

- La base curada ya fue integrada operativamente a `run_e9_stacking.py`.
- `E9_base_h1..h4` ya funciona como insumo oficial del stacking clasico controlado por horizonte.
- La tabla ya no es un pendiente de infraestructura: queda como capa curada de seleccion ex ante y trazabilidad para futuras iteraciones de `E9`.

## 10. Riesgos metodologicos residuales

- Sigue existiendo redundancia natural entre variantes de una misma familia; por eso se mantuvo un corte conservador.
- `E8_v2_clean` aporta una senal puntual, pero no lo suficiente para entrar al bloque principal.
- `E7_v3_clean` solo entra donde aporta sesgo temporal distintivo real (`H2` y `H3`).
- El stacking futuro debe seguir usando solo predicciones OOF y un meta-modelo regularizado sobrio.
"""


def enrich_runs_catalog_with_stacking(
    runs_catalog_df: pd.DataFrame,
    coverage_df: pd.DataFrame,
    stacking_eligibility_df: pd.DataFrame,
    stacking_readiness_df: pd.DataFrame,
) -> pd.DataFrame:
    enriched = runs_catalog_df.copy()
    coverage_map = {
        (row["run_id"], int(row["horizonte"])): row
        for row in coverage_df.to_dict(orient="records")
    }
    eligibility_map = {
        (row["run_id"], int(row["horizonte"])): row
        for row in stacking_eligibility_df.to_dict(orient="records")
    }
    readiness_lookup = stacking_readiness_df.set_index("run_id").to_dict(orient="index")

    for horizon in (1, 2, 3, 4):
        enriched[f"tiene_predicciones_h{horizon}"] = enriched["run_id"].map(
            lambda run_id: bool(coverage_map.get((run_id, horizon), {}).get("existe_archivo", False))
        )
        enriched[f"filas_pred_h{horizon}"] = enriched["run_id"].map(
            lambda run_id: int(coverage_map.get((run_id, horizon), {}).get("n_predicciones", 0) or 0)
        )
        enriched[f"mergeable_h{horizon}"] = enriched["run_id"].map(
            lambda run_id: bool(coverage_map.get((run_id, horizon), {}).get("compatible_para_merge", False))
        )
        enriched[f"stacking_eligible_h{horizon}"] = enriched["run_id"].map(
            lambda run_id: bool(eligibility_map.get((run_id, horizon), {}).get("stacking_eligible_horizonte", False))
        )

    enriched["stacking_eligible_global"] = enriched["run_id"].map(
        lambda run_id: bool(readiness_lookup.get(run_id, {}).get("compatible_para_stacking", False))
    )
    enriched["es_candidato_meta_modelo"] = enriched["stacking_eligible_global"]
    enriched["motivo_no_elegibilidad"] = enriched["run_id"].map(
        lambda run_id: str(readiness_lookup.get(run_id, {}).get("motivo_no_compatible", "") or "")
    )
    enriched["motivo_exclusion_meta_modelo"] = enriched["motivo_no_elegibilidad"]
    return enriched


def enrich_master_with_catalog(master_df: pd.DataFrame, runs_catalog_df: pd.DataFrame) -> pd.DataFrame:
    catalog_columns = [
        "run_id",
        "script_nombre",
        "script_ruta",
        "timestamp_run",
        "run_dir",
        "version_canonica",
        "es_run_maestro",
        "status_canonico",
        "motivo_exclusion_si_aplica",
        "validation_scheme",
        "initial_train_size",
        "tuning_interno",
        "tuning_metric",
        "hyperparams_json",
        "hyperparams_hash",
        "hyperparams_firma",
        "hyperparams_resumen",
        "reconstruccion_hiperparams_status",
        "reconstruccion_hiperparams_nota",
        "feature_count_promedio",
        "feature_count_h1",
        "feature_count_h2",
        "feature_count_h3",
        "feature_count_h4",
        "features_artifact_available",
        "seleccion_variables_tipo",
        "source_dataset_period",
        "dataset_path",
        "feature_columns_count",
        "exogenas_o_no",
        "usa_target_delta",
        "usa_target_nivel",
        "loss_h1",
        "loss_h2",
        "loss_h3",
        "loss_h4",
        "mejor_horizonte_por_loss",
        "fortalezas_operativas",
        "tiene_predicciones_h1",
        "tiene_predicciones_h2",
        "tiene_predicciones_h3",
        "tiene_predicciones_h4",
        "filas_pred_h1",
        "filas_pred_h2",
        "filas_pred_h3",
        "filas_pred_h4",
        "mergeable_h1",
        "mergeable_h2",
        "mergeable_h3",
        "mergeable_h4",
        "stacking_eligible_global",
        "stacking_eligible_h1",
        "stacking_eligible_h2",
        "stacking_eligible_h3",
        "stacking_eligible_h4",
        "motivo_no_elegibilidad",
        "tiene_coeficientes",
        "tiene_importancias",
        "tiene_shap_o_equivalente",
        "explicabilidad_transversal_homogenea",
        "artifact_explicabilidad_path",
        "observacion_explicabilidad",
        "notas_config_clave",
    ]
    catalog_subset = runs_catalog_df[catalog_columns].copy()
    return master_df.merge(catalog_subset, on="run_id", how="left")


def build_rankings(master_df: pd.DataFrame) -> pd.DataFrame:
    ranking_source = master_df[master_df["task_type"].fillna("regresion") == "regresion"].copy()
    rows: list[dict[str, Any]] = []
    ranking_specs = [
        ("global", "L_total_Radar"),
        ("H1", "H1_loss"),
        ("H2", "H2_loss"),
        ("H3", "H3_loss"),
        ("H4", "H4_loss"),
    ]
    for criterio, column in ranking_specs:
        subset = ranking_source.dropna(subset=[column]).sort_values(column, ascending=True).reset_index(drop=True)
        for index, (_, row) in enumerate(subset.iterrows(), start=1):
            rows.append(
                {
                    "criterio": criterio,
                    "ranking": index,
                    "run_id": row["run_id"],
                    "family": row["family"],
                    "model": row["model"],
                    "score": row[column],
                    "feature_mode": row["feature_mode"],
                    "lags": row["lags"],
                    "status_run": row["status_run"],
                }
            )
    return pd.DataFrame(rows)


def sort_metric_subset(subset: pd.DataFrame, *, metric_col: str, ascending: bool) -> pd.DataFrame:
    sortable = subset.copy()
    sortable["__l_total_sort"] = sortable["L_total_Radar"].fillna(float("inf"))
    sorted_subset = sortable.sort_values(
        by=[metric_col, "__l_total_sort", "run_id"],
        ascending=[ascending, True, True],
        kind="mergesort",
    ).reset_index(drop=True)
    return sorted_subset.drop(columns="__l_total_sort")


def build_metric_rankings_long(master_df: pd.DataFrame) -> pd.DataFrame:
    ranking_source = master_df[master_df["task_type"].fillna("regresion") == "regresion"].copy()
    rows: list[dict[str, Any]] = []
    for horizon in (1, 2, 3, 4):
        for spec in HORIZON_METRIC_SPECS:
            metric_col = f"H{horizon}_{spec['column_suffix']}"
            subset = ranking_source.dropna(subset=[metric_col]).copy()
            if subset.empty:
                continue
            subset = sort_metric_subset(
                subset,
                metric_col=metric_col,
                ascending=bool(spec["ascending"]),
            )
            previous_value = None
            for rank, (_, row) in enumerate(subset.iterrows(), start=1):
                value = row[metric_col]
                rows.append(
                    {
                        "horizonte": f"H{horizon}",
                        "metrica": spec["metric"],
                        "sentido_optimizacion": spec["sense"],
                        "criterio_desempate": RANKING_TIEBREAKER_TEXT,
                        "rank": rank,
                        "run_id": row["run_id"],
                        "family": row["family"],
                        "model": row["model"],
                        "valor_metrica": value,
                        "status_run": row["status_run"],
                        "L_total_Radar": row["L_total_Radar"],
                        "observacion_breve": row["observacion_breve"],
                        "empate_exacto_con_anterior": bool(previous_value is not None and value == previous_value),
                    }
                )
                previous_value = value
    return pd.DataFrame(rows)


def build_metric_winners_compact(metric_rankings_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if metric_rankings_df.empty:
        return pd.DataFrame(rows)

    grouped = metric_rankings_df.groupby(["horizonte", "metrica"], sort=True)
    for (horizonte, metrica), subset in grouped:
        subset = subset.sort_values("rank", kind="mergesort").reset_index(drop=True)
        first = subset.iloc[0] if len(subset) >= 1 else None
        second = subset.iloc[1] if len(subset) >= 2 else None
        third = subset.iloc[2] if len(subset) >= 3 else None
        rows.append(
            {
                "horizonte": horizonte,
                "metrica": metrica,
                "sentido_optimizacion": first["sentido_optimizacion"] if first is not None else "",
                "criterio_desempate": RANKING_TIEBREAKER_TEXT,
                "mejor_run_id": first["run_id"] if first is not None else "",
                "mejor_family": first["family"] if first is not None else "",
                "mejor_model": first["model"] if first is not None else "",
                "mejor_status_run": first["status_run"] if first is not None else "",
                "mejor_valor": first["valor_metrica"] if first is not None else None,
                "mejor_L_total_Radar": first["L_total_Radar"] if first is not None else None,
                "segundo_run_id": second["run_id"] if second is not None else "",
                "segundo_family": second["family"] if second is not None else "",
                "segundo_model": second["model"] if second is not None else "",
                "segundo_valor": second["valor_metrica"] if second is not None else None,
                "tercer_run_id": third["run_id"] if third is not None else "",
                "tercer_family": third["family"] if third is not None else "",
                "tercer_model": third["model"] if third is not None else "",
                "tercer_valor": third["valor_metrica"] if third is not None else None,
                "empate_exacto_top1_top2": bool(
                    first is not None
                    and second is not None
                    and first["valor_metrica"] == second["valor_metrica"]
                ),
            }
        )
    return pd.DataFrame(rows)


def build_dimension_rankings(master_df: pd.DataFrame) -> pd.DataFrame:
    ranking_source = master_df[master_df["task_type"].fillna("regresion") == "regresion"].copy()
    rows: list[dict[str, Any]] = []
    for horizon in (1, 2, 3, 4):
        mae_col = f"H{horizon}_mae"
        rmse_col = f"H{horizon}_rmse"
        dir_col = f"H{horizon}_direction_accuracy"
        risk_col = f"H{horizon}_deteccion_caidas"
        loss_col = f"H{horizon}_loss"

        best_mae = sort_metric_subset(
            ranking_source.dropna(subset=[mae_col]),
            metric_col=mae_col,
            ascending=True,
        ).iloc[0]
        best_rmse = sort_metric_subset(
            ranking_source.dropna(subset=[rmse_col]),
            metric_col=rmse_col,
            ascending=True,
        ).iloc[0]
        best_direction = sort_metric_subset(
            ranking_source.dropna(subset=[dir_col]),
            metric_col=dir_col,
            ascending=False,
        ).iloc[0]
        best_risk = sort_metric_subset(
            ranking_source.dropna(subset=[risk_col]),
            metric_col=risk_col,
            ascending=False,
        ).iloc[0]
        best_loss = sort_metric_subset(
            ranking_source.dropna(subset=[loss_col]),
            metric_col=loss_col,
            ascending=True,
        ).iloc[0]

        rows.append(
            {
                "horizonte": f"H{horizon}",
                "mejor_mae_run_id": best_mae["run_id"],
                "mejor_mae_family": best_mae["family"],
                "mejor_mae_model": best_mae["model"],
                "mejor_mae": best_mae[mae_col],
                "mejor_rmse_run_id": best_rmse["run_id"],
                "mejor_rmse_family": best_rmse["family"],
                "mejor_rmse_model": best_rmse["model"],
                "mejor_rmse": best_rmse[rmse_col],
                "mejor_direction_accuracy_run_id": best_direction["run_id"],
                "mejor_direction_accuracy_family": best_direction["family"],
                "mejor_direction_accuracy_model": best_direction["model"],
                "mejor_direction_accuracy": best_direction[dir_col],
                "mejor_deteccion_caidas_run_id": best_risk["run_id"],
                "mejor_deteccion_caidas_family": best_risk["family"],
                "mejor_deteccion_caidas_model": best_risk["model"],
                "mejor_deteccion_caidas": best_risk[risk_col],
                "mejor_loss_run_id": best_loss["run_id"],
                "mejor_loss_family": best_loss["family"],
                "mejor_loss_model": best_loss["model"],
                "mejor_loss_h": best_loss[loss_col],
            }
        )
    return pd.DataFrame(rows)


def discover_planned_run_ids(prompts_dir: Path) -> list[str]:
    planned: set[str] = set()
    for path in prompts_dir.rglob("*.md"):
        text = path.read_text()
        for match in RUN_ID_PATTERN.findall(text):
            planned.add(match)
    return sorted(planned)


def format_metric_value(value: Any) -> str:
    if value is None or pd.isna(value):
        return "NA"
    return f"{float(value):.6f}"


def build_horizon_metric_sections(compact_df: pd.DataFrame) -> str:
    sections: list[str] = []
    metric_order = ("mae", "rmse", "direction_accuracy", "deteccion_caidas", "loss_h")

    for horizonte in ("H1", "H2", "H3", "H4"):
        horizon_rows = compact_df[compact_df["horizonte"] == horizonte].copy()
        if horizon_rows.empty:
            continue
        sections.append(f"### {horizonte}")

        winners_in_horizon: list[str] = []
        for metric in metric_order:
            metric_row = horizon_rows[horizon_rows["metrica"] == metric]
            if metric_row.empty:
                sections.append(f"- mejor en {metric}: no disponible")
                continue
            row = metric_row.iloc[0]
            winners_in_horizon.append(str(row["mejor_run_id"]))
            tie_note = ""
            if bool(row.get("empate_exacto_top1_top2")) and row.get("segundo_run_id"):
                tie_note = f" | empate exacto con `{row['segundo_run_id']}`"
            sections.append(
                f"- mejor en {metric}: `{row['mejor_run_id']}` con `{format_metric_value(row['mejor_valor'])}`{tie_note}"
            )

        winner_counts = pd.Series(winners_in_horizon).value_counts()
        if not winner_counts.empty and int(winner_counts.iloc[0]) > 1:
            dominant_run = winner_counts.index[0]
            sections.append(
                f"- dominio_multiple: `{dominant_run}` gana {int(winner_counts.iloc[0])} de 5 metricas en {horizonte}."
            )
        else:
            sections.append(f"- lectura: no hay un dominador unico en {horizonte}; el horizonte reparte ventajas.")

    return "\n".join(sections)


def render_dataframe_markdown(df: pd.DataFrame, empty_message: str) -> str:
    if df.empty:
        return empty_message
    return "```\n" + df.to_string(index=False) + "\n```"


def build_stacking_readiness_documentation(
    *,
    runs_catalog_df: pd.DataFrame,
    stacking_readiness_df: pd.DataFrame,
    stacking_eligibility_df: pd.DataFrame,
    stacking_base_summary_df: pd.DataFrame,
) -> str:
    reconstruction_counts = (
        runs_catalog_df["reconstruccion_hiperparams_status"].fillna("no_recuperable").value_counts().to_dict()
        if not runs_catalog_df.empty
        else {}
    )
    eligible_by_horizon = (
        stacking_eligibility_df.groupby("horizonte_label")["stacking_eligible_horizonte"].sum().to_dict()
        if not stacking_eligibility_df.empty
        else {}
    )
    return f"""# Readiness para Stacking y Master Table Enriquecida

## Alcance de este documento

Este documento audita elegibilidad para stacking de runs base reutilizables. No fija por si solo el estado canónico completo de cada familia.

En particular:

- `E10_v1_clean` ya existe como corrida canonica de la familia `E10`, pero no aparece como elegible para stacking porque es un meta-selector final y sus predicciones no son mergeables como base simple de `E9`.
- `C1_v1_clean`, `C1_v2_clean` y `C1_v3_clean` si existen como corridas canonicas de clasificacion, pero quedan fuera de stacking por cambio de tarea (`task_type=clasificacion`).

## Que significa un run elegible para stacking en Radar

Un run se considera elegible para stacking solo si cumple simultaneamente estas reglas:

1. Tiene identidad canonica clara (`run_id`, directorio canonico y metadata recuperable).
2. No es un intento abortado, inconsistente o parcial.
3. Tiene predicciones fuera de muestra por horizonte en archivos reales `predicciones_h*.csv`.
4. Esas predicciones incluyen fecha, `y_true` y `y_pred`, sin duplicados de fecha y con merge estructuralmente valido.
5. Tiene metricas trazables por horizonte (`loss_h` y resto del bloque Radar).
6. Su configuracion puede reconstruirse con certeza suficiente desde artefactos reales (`completo` o `parcial`, nunca inventada).
7. No presenta una inconsistencia critica que invalide su reutilizacion.

La elegibilidad se distingue en dos niveles:

- Global: el run es elegible en `H1`, `H2`, `H3` y `H4`.
- Por horizonte: el run puede ser elegible solo en algunos horizontes.

## Campos nuevos incorporados a la tabla maestra

Se amplian cuatro capas:

- Identidad y linaje: `script_nombre`, `script_ruta`, `timestamp_run`, `version_canonica`, `status_canonico`, `motivo_exclusion_si_aplica`.
- Configuracion y reconstruccion: `validation_scheme`, `tuning_metric`, `hyperparams_json`, `hyperparams_hash`, `hyperparams_firma`, `hyperparams_resumen`, `reconstruccion_hiperparams_status`.
- Features y datos: `feature_count_h1..h4`, `features_artifact_available`, `source_dataset_period`, `exogenas_o_no`, `usa_target_delta`, `usa_target_nivel`.
- Stacking y explicabilidad: `mergeable_h1..h4`, `stacking_eligible_h1..h4`, `stacking_eligible_global`, `motivo_no_elegibilidad`, `tiene_coeficientes`, `tiene_importancias`, `tiene_shap_o_equivalente`.

## Como se reconstruyo retrospectivamente

Fuentes utilizadas:

- `metadata_run.json`
- `parametros_run.json`
- `metricas_horizonte.json`
- `resumen_modeling_horizontes.json`
- `predicciones_h1..h4.csv`
- `features_seleccionadas_h1..h4.csv`
- workbook maestro actual
- nombres de script y directorios

Supuestos permitidos:

- Se normalizan hiperparametros a partir de los parametros realmente guardados por cada familia.
- Cuando un run historico no trae exactamente la misma estructura que los recientes, se marca como reconstruccion `parcial`, no se rellena de manera ficticia.

Supuestos no permitidos:

- No se inventan hiperparametros ausentes.
- No se marcan runs elegibles solo por score.
- No se tratan directorios abortados como canonicos.

## Estado de reconstruccion retrospectiva

- Hiperparametros completos: {int(reconstruction_counts.get('completo', 0))}
- Hiperparametros parciales: {int(reconstruction_counts.get('parcial', 0))}
- Hiperparametros no recuperables: {int(reconstruction_counts.get('no_recuperable', 0))}
- Runs elegibles globalmente para stacking: {int(stacking_readiness_df['compatible_para_stacking'].sum()) if not stacking_readiness_df.empty else 0}
- Elegibles por horizonte: {json.dumps(eligible_by_horizon, ensure_ascii=False)}

## Bases stacking por horizonte

{render_dataframe_markdown(stacking_base_summary_df, 'No se construyeron bases stacking.')}

## Limitaciones vigentes

- La explicabilidad transversal entre familias sigue sin estar homogenea.
- Algunos runs historicos mantienen estructura antigua y su reconstruccion es solo parcial.
- La elegibilidad global y la elegibilidad por horizonte no coinciden siempre.
- Las bases stacking siguen siendo preparatorias: no contienen todavia meta-features ni entrenamiento del meta-modelo.
- La no elegibilidad de `E10_v1_clean` en este documento no equivale a que `E10` siga en premodelado; solo significa que `E10` no es un modelo base reutilizable dentro de stacking clasico.

## Uso hacia adelante

La automatizacion hacia adelante no requiere infraestructura paralela:

- `experiment_logger.py` ya refresca la auditoria maestra al cerrar runs completos.
- `backfill_runs.py` ya rehidrata el workbook y dispara el refresh unico al final.
- Esta ampliacion hace que cada corrida nueva quede retroproyectada automaticamente al catalogo enriquecido y a las bases stacking por horizonte.
"""


def build_constructs_documentation(
    *,
    family_constructs_df: pd.DataFrame,
    run_constructs_df: pd.DataFrame,
) -> str:
    return f"""# Diccionario de Constructos Canonicos Radar

## Proposito

Este documento fija los constructos canonicos de familias y runs para mantener una lectura metodologica consistente del proyecto.

## Reglas de lectura

- `constructo_familia` describe el papel vigente de la familia en el programa experimental.
- `constructo_run` describe el papel metodologico del run dentro de su familia o dentro de la historia experimental.
- Un constructo ordena la lectura; no reemplaza metricas, artefactos ni evidencia numerica.
- Los runs tecnicos o no canonicos se conservan por trazabilidad, no como evidencia competitiva principal.

## Constructos de familia

{render_dataframe_markdown(family_constructs_df, 'Sin constructos de familia.')}

## Constructos de run

{render_dataframe_markdown(run_constructs_df, 'Sin constructos de run.')}
"""


def build_markdown_summary(
    *,
    master_df: pd.DataFrame,
    runs_catalog_df: pd.DataFrame,
    inventory_df: pd.DataFrame,
    stacking_readiness_df: pd.DataFrame,
    stacking_eligibility_df: pd.DataFrame,
    coverage_df: pd.DataFrame,
    stacking_base_sheets: dict[str, pd.DataFrame],
    stacking_base_summary_df: pd.DataFrame,
    explainability_status_df: pd.DataFrame,
    planned_run_ids: list[str],
    family_constructs_df: pd.DataFrame,
) -> str:
    family_status_df = build_family_status_vigente_df()
    metric_rankings_df = build_metric_rankings_long(master_df)
    winners_compact_df = build_metric_winners_compact(metric_rankings_df)
    complete_master = master_df[
        (master_df["status_run"] == "completo")
        & (master_df["task_type"].fillna("regresion") == "regresion")
    ].copy()
    best_global = complete_master.nsmallest(1, "L_total_Radar").iloc[0]
    best_h1 = complete_master.nsmallest(1, "H1_loss").iloc[0]
    best_h2 = complete_master.nsmallest(1, "H2_loss").iloc[0]
    best_h3 = complete_master.nsmallest(1, "H3_loss").iloc[0]
    best_h4 = complete_master.nsmallest(1, "H4_loss").iloc[0]
    best_dir_global = complete_master.nlargest(1, "direction_accuracy_promedio").iloc[0]
    best_risk_global = complete_master.nlargest(1, "deteccion_caidas_promedio").iloc[0]
    best_dir_h1 = complete_master.nlargest(1, "H1_direction_accuracy").iloc[0]
    best_dir_h2 = complete_master.nlargest(1, "H2_direction_accuracy").iloc[0]
    best_dir_h3 = complete_master.nlargest(1, "H3_direction_accuracy").iloc[0]
    best_dir_h4 = complete_master.nlargest(1, "H4_direction_accuracy").iloc[0]
    best_risk_h1 = complete_master.nlargest(1, "H1_deteccion_caidas").iloc[0]
    best_risk_h2 = complete_master.nlargest(1, "H2_deteccion_caidas").iloc[0]
    best_risk_h3 = complete_master.nlargest(1, "H3_deteccion_caidas").iloc[0]
    best_risk_h4 = complete_master.nlargest(1, "H4_deteccion_caidas").iloc[0]

    family_counts = complete_master["family"].fillna("desconocida").value_counts().to_dict()
    canonical_run_ids = set(master_df["run_id"])
    planned_without_artifacts = [run_id for run_id in planned_run_ids if run_id not in canonical_run_ids]
    partial_dirs = inventory_df[inventory_df["status_artifactos"] != "completo"]
    if partial_dirs.empty:
        partial_dirs_text = "- No se detectaron directorios parciales."
    else:
        partial_lines = []
        for _, row in partial_dirs.iterrows():
            partial_lines.append(
                f"- `{row['run_id']}` | `{row['status_artifactos']}` | `{row['path_run']}` | {row['comentario']}"
            )
        partial_dirs_text = "\n".join(partial_lines)

    selected_features_count = int((inventory_df["features_seleccionadas"] > 0).sum())
    resumen_count = int((inventory_df["resumen_horizontes"] > 0).sum())
    stacking_compatible = stacking_readiness_df[stacking_readiness_df["compatible_para_stacking"] == True]
    stacking_partial = stacking_readiness_df[
        (stacking_readiness_df["compatible_para_stacking"] != True)
        & (stacking_readiness_df["prioridad_para_meta_modelo"].isin(["media", "baja"]))
    ]
    reconstruction_counts = (
        runs_catalog_df["reconstruccion_hiperparams_status"].fillna("no_recuperable").value_counts().to_dict()
        if not runs_catalog_df.empty
        else {}
    )
    eligible_by_horizon = (
        stacking_eligibility_df.groupby("horizonte_label")["stacking_eligible_horizonte"].sum().to_dict()
        if not stacking_eligibility_df.empty
        else {}
    )
    per_horizon_mergeable = (
        coverage_df.groupby("horizonte_label")["compatible_para_merge"].sum().to_dict()
        if not coverage_df.empty
        else {}
    )
    stacking_base_lines = []
    for sheet_name, sheet_df in stacking_base_sheets.items():
        pred_columns = [
            column
            for column in sheet_df.columns
            if column
            not in {
                "fecha",
                "y_true",
                "n_modelos_disponibles_fila",
                "fila_completa_todos_modelos",
                "cobertura_modelos_fila",
            }
        ]
        stacking_base_lines.append(
            f"- `{sheet_name}`: filas={len(sheet_df)}, modelos_integrados={len(pred_columns)}"
        )
    stacking_base_text = "\n".join(stacking_base_lines) if stacking_base_lines else "- No se construyeron bases parciales."
    incompatibility_examples = stacking_eligibility_df[
        stacking_eligibility_df["stacking_eligible_horizonte"] != True
    ].head(10)
    incompatibility_text = (
        "\n".join(
            f"- `{row['run_id']}` {row['horizonte_label']}: {row['motivo_no_elegibilidad_horizonte']}"
            for _, row in incompatibility_examples.iterrows()
        )
        if not incompatibility_examples.empty
        else "- No se detectaron incompatibilidades de elegibilidad."
    )

    return f"""# Resumen Auditoria Experimentos Radar

## Hallazgos principales

- Runs maestros consolidados: {len(master_df)}
- Runs en catalogo integral: {len(runs_catalog_df)}
- Directorios auditados en inventario: {len(inventory_df)}
- Directorios con artefactos parciales o inconsistentes: {len(partial_dirs)}
- Familias con runs maestros de regresion: {json.dumps(family_counts, ensure_ascii=False)}
- Corridas canonicas de clasificacion detectadas: `C1={int(sum(master_df['run_id'].str.startswith('C1_')))}, C2={int(sum(master_df['run_id'].str.startswith('C2_')))}, C3={int(sum(master_df['run_id'].str.startswith('C3_')))}, C4={int(sum(master_df['run_id'].str.startswith('C4_')))}`

## Mejor desempeño encontrado

- Mejor global crudo por `L_total_Radar`: `{best_global['run_id']}` con `{best_global['L_total_Radar']:.6f}`
- Mejor `H1`: `{best_h1['run_id']}` con `loss_h={best_h1['H1_loss']:.6f}`
- Mejor `H2`: `{best_h2['run_id']}` con `loss_h={best_h2['H2_loss']:.6f}`
- Mejor `H3`: `{best_h3['run_id']}` con `loss_h={best_h3['H3_loss']:.6f}`
- Mejor `H4`: `{best_h4['run_id']}` con `loss_h={best_h4['H4_loss']:.6f}`
- Mejor `direction_accuracy` promedio: `{best_dir_global['run_id']}` con `{best_dir_global['direction_accuracy_promedio']:.6f}`
- Mejor `deteccion_caidas` promedio: `{best_risk_global['run_id']}` con `{best_risk_global['deteccion_caidas_promedio']:.6f}`

## Nota metodologica vigente

- El ranking crudo por `L_total_Radar` no agota la adjudicacion metodologica del proyecto.
- `E1_v5_clean` sigue siendo el referente numerico puro principal del Radar.
- `E9_v2_clean` queda como el mejor referente actual orientado a riesgo, direccion y deteccion de caidas.
- Esa diferencia no se interpreta como contradiccion, sino como dualidad funcional entre pronostico numerico del porcentaje y utilidad operativa del movimiento.
- En consecuencia, `E9` queda util pero pausada, `E10` ya fue probado formalmente con `E10_v1_clean` y queda cerrada para promocion bajo su formulacion actual, y `E11` ya fue abierta con tres variantes controladas y queda evaluada sin promocion tras su primera apertura dual.

## Fortalezas operativas por horizonte

- Mejor `direction_accuracy` en `H1`: `{best_dir_h1['run_id']}` con `{best_dir_h1['H1_direction_accuracy']:.6f}`
- Mejor `direction_accuracy` en `H2`: `{best_dir_h2['run_id']}` con `{best_dir_h2['H2_direction_accuracy']:.6f}`
- Mejor `direction_accuracy` en `H3`: `{best_dir_h3['run_id']}` con `{best_dir_h3['H3_direction_accuracy']:.6f}`
- Mejor `direction_accuracy` en `H4`: `{best_dir_h4['run_id']}` con `{best_dir_h4['H4_direction_accuracy']:.6f}`
- Mejor `deteccion_caidas` en `H1`: `{best_risk_h1['run_id']}` con `{best_risk_h1['H1_deteccion_caidas']:.6f}`
- Mejor `deteccion_caidas` en `H2`: `{best_risk_h2['run_id']}` con `{best_risk_h2['H2_deteccion_caidas']:.6f}`
- Mejor `deteccion_caidas` en `H3`: `{best_risk_h3['run_id']}` con `{best_risk_h3['H3_deteccion_caidas']:.6f}`
- Mejor `deteccion_caidas` en `H4`: `{best_risk_h4['run_id']}` con `{best_risk_h4['H4_deteccion_caidas']:.6f}`

## Ganadores por metrica y horizonte

{build_horizon_metric_sections(winners_compact_df)}

## Inventario real por familia

- `E1`: {sum(run_id.startswith('E1_') for run_id in master_df['run_id'])} runs maestros
- `E2`: {sum(run_id.startswith('E2_') for run_id in master_df['run_id'])} runs maestros
- `E3`: {sum(run_id.startswith('E3_') for run_id in master_df['run_id'])} runs maestros
- `E4`: {sum(run_id.startswith('E4_') for run_id in master_df['run_id'])} runs maestros

## Estado vigente de familias

{render_dataframe_markdown(family_status_df, 'No se pudo construir el estado vigente de familias.')}

## Constructos canonicos

{render_dataframe_markdown(family_constructs_df[['family', 'constructo_familia', 'mejor_run', 'estado_vigente']], 'No se pudo sintetizar constructos de familia.')}

Registro ampliado:

- El detalle completo de constructos por run se exporta en `registro_constructos_runs_radar.csv` y `diccionario_constructos_canonicos_radar.md`.

## Estado de la capa explicativa transversal

{render_dataframe_markdown(explainability_status_df, 'No se pudo sintetizar el estado de la capa explicativa.')}

## Runs parciales / inconsistentes

{partial_dirs_text}

## Runs planeados detectados sin artefactos consolidados

{', '.join(planned_without_artifacts) if planned_without_artifacts else 'No se detectaron run_ids planeados sin artefactos mediante el barrido de prompts.'}

Nota: esta lista sale de menciones en prompts/documentos `.md`; no distingue automaticamente entre runs cancelados, diferidos o nunca ejecutados.

## Inconsistencias detectadas

- `E2_v1_clean` tiene tres directorios: dos intentos abortados y un directorio canónico completo. Se consolidó el directorio `E2_v1_clean_20260323_050029`.
- Los runs históricos `E1_v1` y `E1_v2` usan una estructura de artefactos más antigua, pero contienen métricas y predicciones suficientes para tratarlos como completos.
- No se encontraron artefactos experimentales fuera de `Experimentos/runs/` para `metadata_run.json`, `parametros_run.json`, `metricas_horizonte.json` o `resumen_modeling_horizontes.json`.

## Artefactos útiles para explicación por horizonte

- Directorios con `features_seleccionadas_h*.csv`: {selected_features_count}
- Directorios con `resumen_modeling_horizontes.json`: {resumen_count}
- No se encontraron artefactos homogéneos y reutilizables de importancias de variables o coeficientes comparables entre familias. Esa capa explicativa sigue pendiente.
- La clasificacion principal de esa capa sigue siendo `parcial intra-familia`: hay seleccion de variables por horizonte en algunos runs, pero no una taxonomia homogenea y comparable entre familias.

## Readiness para stacking / hipermodelo

- Runs compatibles para stacking 1..4: {len(stacking_compatible)}
- Runs con utilidad parcial para merge por horizonte: {len(stacking_partial)}
- Reconstrucciones completas de hiperparametros: {int(reconstruction_counts.get('completo', 0))}
- Reconstrucciones parciales de hiperparametros: {int(reconstruction_counts.get('parcial', 0))}
- Runs con hiperparametros no recuperables: {int(reconstruction_counts.get('no_recuperable', 0))}
- Cobertura mergeable por horizonte: {json.dumps(per_horizon_mergeable, ensure_ascii=False)}
- Elegibilidad real para stacking por horizonte: {json.dumps(eligible_by_horizon, ensure_ascii=False)}

### Bases parciales reconstruidas

{stacking_base_text}

### Resumen de bases stacking por horizonte

{render_dataframe_markdown(stacking_base_summary_df, 'No se generaron resumenes de bases stacking.')}

### Lectura retrospectiva

- Se pudo reconstruir retrospectivamente catalogo, metricas por horizonte, cobertura de predicciones y compatibilidad estructural para stacking a partir de artefactos reales en disco.
- Se pudo normalizar retrospectivamente la configuracion e hiperparametros de las familias existentes sin inventar campos ausentes.
- Se pudieron construir hojas `stacking_base_h1` a `stacking_base_h4` con los runs que tienen predicciones mergeables por horizonte.
- Lo que todavia no se pudo homogeneizar retrospectivamente es la capa explicativa transversal entre familias: coeficientes, importancias e interpretabilidad comparable.

### Incompatibilidades tipicas detectadas

{incompatibility_text}

## Lectura preliminar

- La evidencia sí sugiere especialización por horizonte: `{best_h1['run_id']}` domina `H1`, `{best_h2['run_id']}` domina `H2`, `{best_h3['run_id']}` domina `H3` y `{best_h4['run_id']}` domina `H4`.
- También hay especialización operativa: dirección y detección de caídas no siempre coinciden con el mejor `loss_h`.
- En `E3`, la señal más útil quedó en `E3_v2_clean`; `E3_v3_clean` no mejora esa línea y `E4` no desplazó a bagging.
- Ya existe base suficiente para pensar en un ensamblado por horizonte como hipótesis metodológica futura, con cobertura real de `E1-E4` y auditoría explícita de compatibilidad de predicciones.
- Conviene automatizar a partir de ahora la reconstrucción de esta tabla maestra en cada corrida nueva para evitar que el análisis dependa del workbook manual.
"""


def build_experiments_master_table(
    *,
    runs_dir: Path,
    workbook: Path,
    prompts_dir: Path,
    output_dir: Path,
) -> MasterAuditBuildResult:
    runs_dir = runs_dir.expanduser().resolve()
    workbook = workbook.expanduser().resolve()
    prompts_dir = prompts_dir.expanduser().resolve()
    output_dir = output_dir.expanduser().resolve()
    records = scan_run_directories(runs_dir)
    canonical, grouped = select_canonical_records(records)
    summary_lookup, results_lookup = build_workbook_lookup(workbook)

    master_rows = build_master_rows(canonical, summary_lookup, results_lookup)
    master_df = pd.DataFrame(master_rows).sort_values("run_id").reset_index(drop=True)
    runs_catalog_rows = build_runs_catalog_rows(canonical, summary_lookup)
    runs_catalog_df = pd.DataFrame(runs_catalog_rows).sort_values("run_id").reset_index(drop=True)
    coverage_df, standardized_lookup = build_prediction_coverage(canonical)
    stacking_eligibility_df = build_stacking_eligibility_by_horizon(runs_catalog_df, coverage_df)
    stacking_readiness_df = build_stacking_readiness(runs_catalog_df, coverage_df, stacking_eligibility_df)
    runs_catalog_df = enrich_runs_catalog_with_stacking(
        runs_catalog_df,
        coverage_df,
        stacking_eligibility_df,
        stacking_readiness_df,
    )
    master_df = enrich_master_with_catalog(master_df, runs_catalog_df)
    metricas_long_df = build_metricas_por_horizonte_long(canonical, master_df, results_lookup)

    inventory_rows = build_inventory_rows(records, canonical, grouped)
    inventory_df = pd.DataFrame(inventory_rows).sort_values(["run_id", "path_run"]).reset_index(drop=True)

    ranking_df = build_rankings(master_df)
    ranking_dim_df = build_dimension_rankings(master_df)
    metric_rankings_df = build_metric_rankings_long(master_df)
    winners_compact_df = build_metric_winners_compact(metric_rankings_df)
    planned_run_ids = discover_planned_run_ids(prompts_dir)

    stacking_base_sheets = {
        f"stacking_base_h{horizon}": build_stacking_base_sheet(
            horizon=horizon,
            stacking_eligibility_df=stacking_eligibility_df,
            standardized_lookup=standardized_lookup,
        )
        for horizon in (1, 2, 3, 4)
    }
    stacking_base_summary_df = build_stacking_base_summary(
        stacking_eligibility_df=stacking_eligibility_df,
        stacking_base_sheets=stacking_base_sheets,
    )
    e9_construct_df = build_e9_construct_dictionary(runs_catalog_df)
    e9_metricas_df = build_e9_metricas_candidatos(runs_catalog_df, e9_construct_df)
    e9_base_sheets = {
        f"E9_base_h{horizon}": build_e9_curated_base_sheet(
            horizon=horizon,
            stacking_base_sheet=stacking_base_sheets.get(f"stacking_base_h{horizon}", pd.DataFrame()),
        )
        for horizon in (1, 2, 3, 4)
    }
    e9_diagnostic_sheets = {
        f"E9_diagnostico_h{horizon}": build_e9_diagnostic_sheet(
            horizon=horizon,
            master_df=master_df,
            construct_df=e9_construct_df,
            stacking_base_sheet=stacking_base_sheets.get(f"stacking_base_h{horizon}", pd.DataFrame()),
            e9_base_sheet=e9_base_sheets.get(f"E9_base_h{horizon}", pd.DataFrame()),
        )
        for horizon in (1, 2, 3, 4)
    }
    e9_base_summary_df = build_e9_bases_summary(e9_base_sheets)
    family_status_df = build_family_status_vigente_df()
    family_constructs_df = build_family_constructs_df()
    run_constructs_df = build_run_constructs_df(master_df, family_status_df)
    explainability_status_df = build_explainability_status_df(runs_catalog_df)
    operational_benchmarks_df = build_operational_benchmarks_df(master_df)
    radar_layers_df = build_radar_layers_df()
    dual_functional_metrics_df = build_dual_functional_metrics_df()
    dual_policy_summary_df = build_dual_policy_summary_df()
    reconstruction_counts = runs_catalog_df["reconstruccion_hiperparams_status"].fillna("no_recuperable").value_counts().to_dict()
    eligible_by_horizon = (
        stacking_eligibility_df.groupby("horizonte_label")["stacking_eligible_horizonte"].sum().to_dict()
        if not stacking_eligibility_df.empty
        else {}
    )

    inventory_json = {
        "summary": {
            "master_runs_total": int(len(master_df)),
            "runs_catalog_total": int(len(runs_catalog_df)),
            "inventory_directories_total": int(len(inventory_df)),
            "e1_master_runs": int(sum(master_df["run_id"].str.startswith("E1_"))),
            "e2_master_runs": int(sum(master_df["run_id"].str.startswith("E2_"))),
            "e3_master_runs": int(sum(master_df["run_id"].str.startswith("E3_"))),
            "e4_master_runs": int(sum(master_df["run_id"].str.startswith("E4_"))),
            "c1_master_runs": int(sum(master_df["run_id"].str.startswith("C1_"))),
            "c2_master_runs": int(sum(master_df["run_id"].str.startswith("C2_"))),
            "c3_master_runs": int(sum(master_df["run_id"].str.startswith("C3_"))),
            "c4_master_runs": int(sum(master_df["run_id"].str.startswith("C4_"))),
            "stacking_compatible_runs": int(stacking_readiness_df["compatible_para_stacking"].sum()),
            "hyperparams_reconstruccion_completa": int(reconstruction_counts.get("completo", 0)),
            "hyperparams_reconstruccion_parcial": int(reconstruction_counts.get("parcial", 0)),
            "hyperparams_no_recuperables": int(reconstruction_counts.get("no_recuperable", 0)),
            "stacking_eligible_h1": int(eligible_by_horizon.get("H1", 0)),
            "stacking_eligible_h2": int(eligible_by_horizon.get("H2", 0)),
            "stacking_eligible_h3": int(eligible_by_horizon.get("H3", 0)),
            "stacking_eligible_h4": int(eligible_by_horizon.get("H4", 0)),
        },
        "canonical_runs": master_df.to_dict(orient="records"),
        "runs_catalogo": runs_catalog_df.to_dict(orient="records"),
        "metricas_por_horizonte_long": metricas_long_df.to_dict(orient="records"),
        "inventory_directories": inventory_rows,
        "stacking_readiness": stacking_readiness_df.to_dict(orient="records"),
        "stacking_elegibilidad_por_horizonte": stacking_eligibility_df.to_dict(orient="records"),
        "cobertura_predicciones": coverage_df.to_dict(orient="records"),
        "stacking_bases_resumen": stacking_base_summary_df.to_dict(orient="records"),
        "stacking_base_summary": {
            sheet_name: {
                "filas": int(len(sheet_df)),
                "columnas_modelo": int(
                    len(
                        [
                            c
                            for c in sheet_df.columns
                            if c
                            not in {
                                "fecha",
                                "y_true",
                                "n_modelos_disponibles_fila",
                                "fila_completa_todos_modelos",
                                "cobertura_modelos_fila",
                            }
                        ]
                    )
                ),
            }
            for sheet_name, sheet_df in stacking_base_sheets.items()
        },
        "e9_curacion": {
            "candidatos_finales_por_horizonte": {
                f"H{horizon}": E9_FINAL_CANDIDATES_BY_HORIZON[horizon]
                for horizon in (1, 2, 3, 4)
            },
            "reservas_por_horizonte": {
                f"H{horizon}": E9_RESERVE_CANDIDATES_BY_HORIZON[horizon]
                for horizon in (1, 2, 3, 4)
            },
            "runs_principales": e9_construct_df[e9_construct_df["decision_global_E9"] == "entra"].to_dict(orient="records"),
            "runs_reserva": e9_construct_df[e9_construct_df["decision_global_E9"] == "reserva"].to_dict(orient="records"),
            "bases_resumen": e9_base_summary_df.to_dict(orient="records"),
        },
        "estado_familias_vigente": family_status_df.to_dict(orient="records"),
        "constructos_familias_vigente": family_constructs_df.to_dict(orient="records"),
        "constructos_runs_vigente": run_constructs_df.to_dict(orient="records"),
        "estado_capa_explicativa_transversal": explainability_status_df.to_dict(orient="records"),
        "benchmarks_operativos_vigentes": operational_benchmarks_df.to_dict(orient="records"),
        "capas_radar_vigentes": radar_layers_df.to_dict(orient="records"),
        "tabla_funcional_dual_vigente": dual_functional_metrics_df.to_dict(orient="records"),
        "politica_funcional_dual_vigente": dual_policy_summary_df.to_dict(orient="records"),
        "planned_run_ids_detected": planned_run_ids,
        "planned_without_artifacts": [
            run_id for run_id in planned_run_ids if run_id not in set(master_df["run_id"])
        ],
    }

    csv_path = output_dir / "tabla_maestra_experimentos_radar.csv"
    xlsx_path = output_dir / "tabla_maestra_experimentos_radar.xlsx"
    e9_curated_xlsx_path = output_dir / "tabla_maestra_experimentos_radar_e9_curada.xlsx"
    json_path = output_dir / "inventario_experimentos_radar.json"
    md_path = output_dir / "resumen_auditoria_experimentos.md"
    readiness_doc_path = output_dir / "documentacion_stacking_readiness_radar.md"
    e9_markdown_path = output_dir / "preparacion_tabla_e9_stacking_controlado.md"
    constructs_doc_path = output_dir / "diccionario_constructos_canonicos_radar.md"
    constructs_csv_path = output_dir / "registro_constructos_runs_radar.csv"
    operational_json_path = output_dir / "registro_operacion_controlada_radar.json"

    master_df.to_csv(csv_path, index=False)
    run_constructs_df.to_csv(constructs_csv_path, index=False)
    operational_json_path.write_text(
        json.dumps(
            {
                "benchmarks_operativos_vigentes": operational_benchmarks_df.to_dict(orient="records"),
                "capas_radar_vigentes": radar_layers_df.to_dict(orient="records"),
                "tabla_funcional_dual_vigente": dual_functional_metrics_df.to_dict(orient="records"),
                "politica_funcional_dual_vigente": dual_policy_summary_df.to_dict(orient="records"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        runs_catalog_df.to_excel(writer, sheet_name="runs_catalogo", index=False)
        master_df.to_excel(writer, sheet_name="runs_consolidados", index=False)
        metricas_long_df.to_excel(writer, sheet_name="metricas_por_horizonte_long", index=False)
        ranking_df.to_excel(writer, sheet_name="ranking_por_horizonte", index=False)
        ranking_dim_df.to_excel(writer, sheet_name="ranking_dimensiones", index=False)
        metric_rankings_df.to_excel(writer, sheet_name="ranking_metricas_por_horizonte", index=False)
        winners_compact_df.to_excel(writer, sheet_name="ganadores_por_metrica_horizonte", index=False)
        stacking_readiness_df.to_excel(writer, sheet_name="stacking_readiness", index=False)
        stacking_eligibility_df.to_excel(writer, sheet_name="stacking_elegibilidad_h", index=False)
        coverage_df.to_excel(writer, sheet_name="cobertura_predicciones", index=False)
        stacking_base_summary_df.to_excel(writer, sheet_name="stacking_bases_resumen", index=False)
        family_status_df.to_excel(writer, sheet_name="estado_familias_vigente", index=False)
        family_constructs_df.to_excel(writer, sheet_name="constructos_familias_vigente", index=False)
        run_constructs_df.to_excel(writer, sheet_name="constructos_runs_vigente", index=False)
        explainability_status_df.to_excel(writer, sheet_name="estado_explicabilidad_transversal", index=False)
        operational_benchmarks_df.to_excel(writer, sheet_name="benchmarks_operativos_vigentes", index=False)
        radar_layers_df.to_excel(writer, sheet_name="capas_radar_vigentes", index=False)
        dual_functional_metrics_df.to_excel(writer, sheet_name="tabla_funcional_dual_vigente", index=False)
        dual_policy_summary_df.to_excel(writer, sheet_name="politica_funcional_dual_vigente", index=False)
        inventory_df.to_excel(writer, sheet_name="inventario_runs", index=False)
        for sheet_name, sheet_df in stacking_base_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    e9_curacion_resumen_df = build_e9_curacion_resumen_df(
        source_xlsx_path=xlsx_path,
        curated_xlsx_path=e9_curated_xlsx_path,
    )
    with pd.ExcelWriter(e9_curated_xlsx_path, engine="openpyxl") as writer:
        runs_catalog_df.to_excel(writer, sheet_name="runs_catalogo", index=False)
        master_df.to_excel(writer, sheet_name="runs_consolidados", index=False)
        metricas_long_df.to_excel(writer, sheet_name="metricas_por_horizonte_long", index=False)
        ranking_df.to_excel(writer, sheet_name="ranking_por_horizonte", index=False)
        ranking_dim_df.to_excel(writer, sheet_name="ranking_dimensiones", index=False)
        metric_rankings_df.to_excel(writer, sheet_name="ranking_metricas_por_horizonte", index=False)
        winners_compact_df.to_excel(writer, sheet_name="ganadores_por_metrica_horizonte", index=False)
        stacking_readiness_df.to_excel(writer, sheet_name="stacking_readiness", index=False)
        stacking_eligibility_df.to_excel(writer, sheet_name="stacking_elegibilidad_h", index=False)
        coverage_df.to_excel(writer, sheet_name="cobertura_predicciones", index=False)
        stacking_base_summary_df.to_excel(writer, sheet_name="stacking_bases_resumen", index=False)
        family_status_df.to_excel(writer, sheet_name="estado_familias_vigente", index=False)
        family_constructs_df.to_excel(writer, sheet_name="constructos_familias_vigente", index=False)
        run_constructs_df.to_excel(writer, sheet_name="constructos_runs_vigente", index=False)
        explainability_status_df.to_excel(writer, sheet_name="estado_explicabilidad_transversal", index=False)
        operational_benchmarks_df.to_excel(writer, sheet_name="benchmarks_operativos_vigentes", index=False)
        radar_layers_df.to_excel(writer, sheet_name="capas_radar_vigentes", index=False)
        dual_functional_metrics_df.to_excel(writer, sheet_name="tabla_funcional_dual_vigente", index=False)
        dual_policy_summary_df.to_excel(writer, sheet_name="politica_funcional_dual_vigente", index=False)
        inventory_df.to_excel(writer, sheet_name="inventario_runs", index=False)
        for sheet_name, sheet_df in stacking_base_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        e9_curacion_resumen_df.to_excel(writer, sheet_name="E9_curacion_resumen", index=False)
        e9_construct_df.to_excel(writer, sheet_name="E9_diccionario_constructos", index=False)
        e9_metricas_df.to_excel(writer, sheet_name="E9_metricas_candidatos", index=False)
        for sheet_name, sheet_df in e9_diagnostic_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        for sheet_name, sheet_df in e9_base_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        e9_base_summary_df.to_excel(writer, sheet_name="E9_bases_resumen", index=False)

    json_path.write_text(json.dumps(inventory_json, ensure_ascii=False, indent=2))
    md_path.write_text(
        build_markdown_summary(
            master_df=master_df,
            runs_catalog_df=runs_catalog_df,
            inventory_df=inventory_df,
            stacking_readiness_df=stacking_readiness_df,
            stacking_eligibility_df=stacking_eligibility_df,
            coverage_df=coverage_df,
            stacking_base_sheets=stacking_base_sheets,
            stacking_base_summary_df=stacking_base_summary_df,
            explainability_status_df=explainability_status_df,
            planned_run_ids=planned_run_ids,
            family_constructs_df=family_constructs_df,
        )
    )
    readiness_doc_path.write_text(
        build_stacking_readiness_documentation(
            runs_catalog_df=runs_catalog_df,
            stacking_readiness_df=stacking_readiness_df,
            stacking_eligibility_df=stacking_eligibility_df,
            stacking_base_summary_df=stacking_base_summary_df,
        )
    )
    e9_markdown_path.write_text(
        build_e9_markdown_report(
            source_xlsx_path=xlsx_path,
            curated_xlsx_path=e9_curated_xlsx_path,
            construct_df=e9_construct_df,
            e9_bases_summary_df=e9_base_summary_df,
        )
    )
    constructs_doc_path.write_text(
        build_constructs_documentation(
            family_constructs_df=family_constructs_df,
            run_constructs_df=run_constructs_df,
        )
    )

    return MasterAuditBuildResult(
        csv_path=csv_path,
        xlsx_path=xlsx_path,
        json_path=json_path,
        markdown_path=md_path,
        master_runs_total=len(master_df),
        inventory_directories_total=len(inventory_df),
        e9_curated_xlsx_path=e9_curated_xlsx_path,
        e9_markdown_path=e9_markdown_path,
    )


def main() -> None:
    args = parse_args()
    result = build_experiments_master_table(
        runs_dir=args.runs_dir,
        workbook=args.workbook,
        prompts_dir=args.prompts_dir,
        output_dir=args.output_dir,
    )

    print(f"CSV: {result.csv_path}")
    print(f"XLSX: {result.xlsx_path}")
    if result.e9_curated_xlsx_path is not None:
        print(f"E9_CURATED_XLSX: {result.e9_curated_xlsx_path}")
    print(f"JSON: {result.json_path}")
    print(f"MD: {result.markdown_path}")
    if result.e9_markdown_path is not None:
        print(f"E9_MD: {result.e9_markdown_path}")
    print(f"Runs maestros: {result.master_runs_total}")
    print(f"Directorios inventariados: {result.inventory_directories_total}")


if __name__ == "__main__":
    main()
