#!/usr/bin/env python3
"""
Calcula sentimiento semanal para Facebook, Twitter, YouTube y medios usando
diccionarios de polaridad y la ponderacion de redes planteada en el script
original.

Salida:
    - Un solo archivo de Excel
    - Una sola hoja
    - Una fila por semana ISO
    - Columnas con:
        1. sentimiento ponderado de redes
        2. sentimiento de medios
        3. promedio de ambos
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURACION
# =============================================================================
BASE_DIR = Path("/home/emilio/Documentos/RAdAR")
TEXTOS_DIR = BASE_DIR / "Datos_RadaR_Texto"
DICT_DIR = BASE_DIR / "Diccionarios_NLP" / "Diccionarios_Polaridad"
OUTPUT_PATH = (
    BASE_DIR
    / "Datos_Modelo_ML"
    / "aceptacion_digital_redes_medios_sentimiento_semanal.xlsx"
)

FUENTES = {
    "facebook": TEXTOS_DIR / "Facebook_Semana_Texto",
    "twitter": TEXTOS_DIR / "Twitter_Semana_Texto",
    "youtube": TEXTOS_DIR / "Youtube_Semana_Texto",
    "medios": TEXTOS_DIR / "Medios_Semana_Texto",
}

STOPLIST_PATH = DICT_DIR / "stop_list_espanol_limpia.txt"
DICT_POS_PATH = DICT_DIR / "diccionario_palabras_positivas.txt"
DICT_NEG_PATH = DICT_DIR / "diccionario_palabras_negativas.txt"

# Parametros del modelo de ponderacion del script original
U_FACEBOOK = 93.0
U_TWITTER = 16.9
U_YOUTUBE = 84.0

F_FACEBOOK = 0.40
F_TWITTER = 1.25
F_YOUTUBE = 1.00  # Escenario moderado del script original

ARCHIVO_RE = re.compile(r"^(?P<anio>\d{2})_(?P<semana>\d{2})_(?P<fuente>[a-z]+)\.txt$")
TOKEN_RE = re.compile(r"[a-z]+")

HOJA_EXCEL = "sentimiento_semanal"
COLUMNAS_SENTIMIENTO = [
    "sentimiento_facebook",
    "sentimiento_twitter",
    "sentimiento_youtube",
    "sentimiento_redes_ponderado",
    "sentimiento_medios",
    "promedio_redes_medios",
]


# =============================================================================
# FUNCIONES BASE
# =============================================================================
def normalizar(texto: str) -> str:
    texto = texto.lower().strip().replace("\ufeff", "")
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(char for char in nfkd if not unicodedata.combining(char))


def cargar_lista(path: Path) -> set[str]:
    palabras = set()
    with path.open(encoding="utf-8") as handle:
        for linea in handle:
            palabra = normalizar(linea.replace("\t", " ").strip())
            if palabra:
                palabras.add(palabra)
    return palabras


def tokenizar_y_limpiar(texto: str, stopwords: set[str]) -> list[str]:
    tokens = TOKEN_RE.findall(normalizar(texto))
    return [token for token in tokens if len(token) > 2 and token not in stopwords]


def calcular_pesos_redes() -> dict[str, float]:
    n_facebook = U_FACEBOOK * F_FACEBOOK
    n_twitter = U_TWITTER * F_TWITTER
    n_youtube = U_YOUTUBE * F_YOUTUBE
    total = n_facebook + n_twitter + n_youtube

    return {
        "facebook": n_facebook / total,
        "twitter": n_twitter / total,
        "youtube": n_youtube / total,
    }


def extraer_periodo_iso(path: Path, fuente_esperada: str) -> tuple[int, int]:
    match = ARCHIVO_RE.match(path.name)
    if not match:
        raise ValueError(f"Nombre de archivo no valido: {path.name}")

    fuente = match.group("fuente")
    if fuente != fuente_esperada:
        raise ValueError(
            f"Se esperaba fuente '{fuente_esperada}' pero el archivo es '{path.name}'"
        )

    anio_iso = 2000 + int(match.group("anio"))
    semana_iso = int(match.group("semana"))
    return anio_iso, semana_iso


def indexar_archivos(directorio: Path, fuente: str) -> dict[tuple[int, int], Path]:
    if not directorio.exists():
        raise FileNotFoundError(f"No existe el directorio: {directorio}")

    archivos = {}
    for path in sorted(directorio.glob("*.txt")):
        periodo = extraer_periodo_iso(path, fuente)
        if periodo in archivos:
            raise ValueError(
                f"Periodo duplicado para {fuente}: {periodo} en {path} y {archivos[periodo]}"
            )
        archivos[periodo] = path
    return archivos


def analizar_archivo(
    path: Path,
    stopwords: set[str],
    dict_pos: set[str],
    dict_neg: set[str],
) -> dict[str, float | int]:
    texto = path.read_text(encoding="utf-8", errors="ignore")
    tokens = tokenizar_y_limpiar(texto, stopwords)

    positivas = sum(token in dict_pos for token in tokens)
    negativas = sum(token in dict_neg for token in tokens)
    total_sentimiento = positivas + negativas

    if total_sentimiento:
        sentimiento = (positivas - negativas) / total_sentimiento
    else:
        sentimiento = 0.0

    return {
        "tokens_limpios": len(tokens),
        "positivas": int(positivas),
        "negativas": int(negativas),
        "total_sentimiento": int(total_sentimiento),
        "sentimiento": float(sentimiento),
    }


def promedio_ponderado(valores: dict[str, float], pesos: dict[str, float]) -> float:
    disponibles = {
        fuente: valor
        for fuente, valor in valores.items()
        if valor is not None and not pd.isna(valor)
    }
    if not disponibles:
        return float("nan")

    peso_total = sum(pesos[fuente] for fuente in disponibles)
    if peso_total == 0:
        return float("nan")

    return sum(disponibles[fuente] * pesos[fuente] for fuente in disponibles) / peso_total


def promedio_simple(*valores: float) -> float:
    disponibles = [valor for valor in valores if valor is not None and not pd.isna(valor)]
    if not disponibles:
        return float("nan")
    return sum(disponibles) / len(disponibles)


def construir_fila(
    periodo: tuple[int, int],
    analisis_por_fuente: dict[str, dict[str, float | int] | None],
    pesos_redes: dict[str, float],
) -> dict[str, float | int | str]:
    anio_iso, semana_iso = periodo

    fila: dict[str, float | int | str] = {
        "anio_iso": anio_iso,
        "semana_iso": semana_iso,
        "periodo_iso": f"{anio_iso}-W{semana_iso:02d}",
    }

    for fuente in FUENTES:
        analisis = analisis_por_fuente.get(fuente)
        if analisis is None:
            fila[f"sentimiento_{fuente}"] = float("nan")
            continue

        fila[f"sentimiento_{fuente}"] = analisis["sentimiento"]

    sentimiento_redes = promedio_ponderado(
        {
            "facebook": fila["sentimiento_facebook"],
            "twitter": fila["sentimiento_twitter"],
            "youtube": fila["sentimiento_youtube"],
        },
        pesos_redes,
    )

    fila["sentimiento_redes_ponderado"] = sentimiento_redes
    fila["promedio_redes_medios"] = promedio_simple(
        sentimiento_redes,
        fila["sentimiento_medios"],
    )

    return fila


def ajustar_hoja_excel(writer: pd.ExcelWriter, dataframe: pd.DataFrame) -> None:
    hoja = writer.sheets[HOJA_EXCEL]
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    encabezado_fill = PatternFill("solid", fgColor="1F4E78")
    encabezado_font = Font(color="FFFFFF", bold=True)

    for cell in hoja[1]:
        cell.fill = encabezado_fill
        cell.font = encabezado_font

    columnas_sentimiento = {col: idx + 1 for idx, col in enumerate(dataframe.columns)}

    for column_cells in hoja.columns:
        letra = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        hoja.column_dimensions[letra].width = min(max(max_len + 2, 12), 28)

    for nombre_columna in COLUMNAS_SENTIMIENTO:
        indice = columnas_sentimiento.get(nombre_columna)
        if not indice:
            continue
        letra = get_column_letter(indice)
        for row in range(2, hoja.max_row + 1):
            hoja[f"{letra}{row}"].number_format = "0.0000"


def guardar_excel(dataframe: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name=HOJA_EXCEL, index=False)
        ajustar_hoja_excel(writer, dataframe)


# =============================================================================
# EJECUCION
# =============================================================================
def main() -> None:
    print("=" * 72)
    print("SENTIMIENTO SEMANAL DE REDES Y MEDIOS")
    print("=" * 72)

    stopwords = cargar_lista(STOPLIST_PATH)
    dict_pos = cargar_lista(DICT_POS_PATH)
    dict_neg = cargar_lista(DICT_NEG_PATH)
    pesos_redes = calcular_pesos_redes()

    print("\nDiccionarios cargados:")
    print(f"  Stopwords: {len(stopwords):,}")
    print(f"  Positivas: {len(dict_pos):,}")
    print(f"  Negativas: {len(dict_neg):,}")

    print("\nPesos de redes:")
    print(f"  Facebook: {pesos_redes['facebook']:.4f}")
    print(f"  Twitter:  {pesos_redes['twitter']:.4f}")
    print(f"  YouTube:  {pesos_redes['youtube']:.4f}")
    print(f"  f_youtube usado: {F_YOUTUBE:.2f}")

    archivos_por_fuente = {
        fuente: indexar_archivos(directorio, fuente)
        for fuente, directorio in FUENTES.items()
    }

    for fuente, archivos in archivos_por_fuente.items():
        print(f"  {fuente.title():<8}: {len(archivos):>3} archivos")

    periodos = sorted(
        {
            periodo
            for archivos in archivos_por_fuente.values()
            for periodo in archivos
        }
    )

    filas = []
    for anio_iso, semana_iso in periodos:
        periodo = (anio_iso, semana_iso)
        analisis_por_fuente = {}
        faltantes = []

        for fuente, archivos in archivos_por_fuente.items():
            path = archivos.get(periodo)
            if path is None:
                analisis_por_fuente[fuente] = None
                faltantes.append(fuente)
                continue
            analisis_por_fuente[fuente] = analizar_archivo(path, stopwords, dict_pos, dict_neg)

        if faltantes:
            print(
                f"Aviso: faltan fuentes en {anio_iso}-W{semana_iso:02d}: "
                + ", ".join(faltantes)
            )

        filas.append(construir_fila(periodo, analisis_por_fuente, pesos_redes))

    dataframe = pd.DataFrame(filas).sort_values(["anio_iso", "semana_iso"]).reset_index(drop=True)

    columnas_principales = [
        "anio_iso",
        "semana_iso",
        "periodo_iso",
        "sentimiento_facebook",
        "sentimiento_twitter",
        "sentimiento_youtube",
        "sentimiento_redes_ponderado",
        "sentimiento_medios",
        "promedio_redes_medios",
    ]
    dataframe = dataframe[columnas_principales]

    guardar_excel(dataframe, OUTPUT_PATH)

    print("\nResumen:")
    print(f"  Semanas procesadas: {len(dataframe):,}")
    print(f"  Archivo generado: {OUTPUT_PATH}")
    print("=" * 72)


if __name__ == "__main__":
    main()
