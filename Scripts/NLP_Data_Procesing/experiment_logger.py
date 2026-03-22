#!/usr/bin/env python3
"""
Bitacora automatica de experimentos para el radar politico.

Idea base:
- El Excel `Experimentos/grid_experimentos_radar.xlsx` sigue siendo el indice maestro.
- Los archivos pesados o reproducibles viven en `Experimentos/runs/<run_id>_<timestamp>/`.
- El Excel guarda metadata del run, rutas al script, snapshot, JSON de parametros,
  JSON de resultados y una hoja adicional con todos los artefactos.

Uso rapido dentro de un script de modelado:

    from experiment_logger import RadarExperimentTracker

    tracker = RadarExperimentTracker()
    run = tracker.start_run(
        run_id="E4_v2",
        experiment_id="E4",
        family="boosting",
        model="LightGBM",
        script_path=__file__,
        parametros={"learning_rate": 0.03, "num_leaves": 31},
    )

    run.save_json({"nota": "configuracion inicial"}, "config_extra.json", artifact_type="config")
    run.save_json({"mae": 1.42}, "resumen_h1.json", artifact_type="metricas")

    run.finalize(
        horizon_results=[
            {
                "horizonte_sem": 1,
                "target": "nivel",
                "variables_temporales": "lags IAD, MA, momentum",
                "variables_tematicas": "sentimiento, temas, intensidad",
                "transformacion": "original",
                "seleccion_variables": "importancia_arboles",
                "validacion": "walk-forward",
                "dataset_periodo": "2024-10-22 a 2026-03-09",
                "notas_config": "LightGBM con lags 1..4 y sentimiento",
                "l_num": 0.11,
                "l_trend": 0.10,
                "l_risk": 0.08,
                "l_tol": 0.07,
                "mae": 1.42,
                "rmse": 1.86,
                "direccion_accuracy": 0.81,
                "deteccion_caidas": 0.75,
            },
        ],
        estado="corrido",
        comentarios="Primer baseline trazable en grid automatizado.",
    )

Preparacion inicial del workbook:
    python3 Scripts/NLP_Data_Procesing/experiment_logger.py --prepare-workbook
"""

from __future__ import annotations

import argparse
import fcntl
import json
import shutil
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = ROOT_DIR / "Experimentos" / "grid_experimentos_radar.xlsx"
DEFAULT_RUNS_DIR = ROOT_DIR / "Experimentos" / "runs"
RESULTADOS_SHEET = "RESULTADOS_GRID"
SUMMARY_SHEET = "RUN_SUMMARY"
ARTEFACTOS_SHEET = "RUN_ARTEFACTOS"

RESULTADOS_EXTRA_HEADERS = [
    "Script_nombre",
    "Script_ruta",
    "Run_dir",
    "Snapshot_script",
    "Parametros_ruta_json",
    "Resultados_ruta_json",
    "Timestamp_run",
]

SUMMARY_EXTRA_HEADERS = [
    "Script_nombre",
    "Script_ruta",
    "Run_dir",
    "Snapshot_script",
    "Parametros_ruta_json",
    "Resultados_ruta_json",
    "Timestamp_run",
]

ARTEFACTOS_HEADERS = [
    "Fecha",
    "Run_ID",
    "Experiment_ID",
    "Tipo",
    "Nombre",
    "Ruta",
    "Notas",
]

DEFAULT_COLUMN_WIDTHS = {
    "Script_nombre": 24,
    "Script_ruta": 42,
    "Run_dir": 40,
    "Snapshot_script": 42,
    "Parametros_ruta_json": 38,
    "Resultados_ruta_json": 38,
    "Timestamp_run": 22,
}


def now_dt() -> datetime:
    return datetime.now().replace(microsecond=0)


def now_text() -> str:
    return now_dt().isoformat(sep=" ")


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def prefer_repo_relative(path_like: str | Path | None) -> str:
    if not path_like:
        return ""
    path = Path(path_like).expanduser()
    if not path.is_absolute():
        path = path.resolve()
    try:
        return str(path.relative_to(ROOT_DIR))
    except ValueError:
        return str(path)


def json_dumps_pretty(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True)


def copy_cell_format(source_cell, target_cell) -> None:
    target_cell._style = copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)


def row_has_value(worksheet, row_idx: int) -> bool:
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row_idx, col_idx).value not in (None, ""):
            return True
    return False


def find_header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, col_idx).value
        if header not in (None, ""):
            headers[str(header)] = col_idx
    return headers


class WorkbookLock:
    def __init__(self, workbook_path: Path) -> None:
        self.lock_path = workbook_path.with_suffix(workbook_path.suffix + ".lock")
        self._handle = None

    def __enter__(self):
        ensure_parent(self.lock_path)
        self._handle = self.lock_path.open("w", encoding="utf-8")
        fcntl.flock(self._handle.fileno(), fcntl.LOCK_EX)
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        if self._handle is not None:
            fcntl.flock(self._handle.fileno(), fcntl.LOCK_UN)
            self._handle.close()
        if self.lock_path.exists():
            self.lock_path.unlink()


@dataclass
class RunContext:
    tracker: "RadarExperimentTracker"
    run_id: str
    experiment_id: str
    family: str
    model: str
    script_path: Path
    run_dir: Path
    snapshot_path: Path
    created_at: str
    parametros_path: Path | None = None
    artifacts: list[dict[str, str]] = field(default_factory=list)

    def register_artifact(
        self,
        path: str | Path,
        artifact_type: str,
        label: str | None = None,
        notes: str = "",
    ) -> Path:
        artifact_path = Path(path).resolve()
        self.artifacts.append(
            {
                "artifact_type": artifact_type,
                "label": label or artifact_path.name,
                "path": str(artifact_path),
                "notes": notes,
            }
        )
        return artifact_path

    def save_json(
        self,
        payload: Any,
        filename: str,
        artifact_type: str = "json",
        notes: str = "",
    ) -> Path:
        output_path = self.run_dir / filename
        ensure_parent(output_path)
        output_path.write_text(json_dumps_pretty(payload), encoding="utf-8")
        return self.register_artifact(output_path, artifact_type=artifact_type, notes=notes)

    def save_text(
        self,
        content: str,
        filename: str,
        artifact_type: str = "texto",
        notes: str = "",
    ) -> Path:
        output_path = self.run_dir / filename
        ensure_parent(output_path)
        output_path.write_text(content, encoding="utf-8")
        return self.register_artifact(output_path, artifact_type=artifact_type, notes=notes)

    def save_dataframe(
        self,
        dataframe: Any,
        filename: str,
        artifact_type: str = "tabla",
        notes: str = "",
        index: bool = False,
    ) -> Path:
        output_path = self.run_dir / filename
        ensure_parent(output_path)
        suffix = output_path.suffix.lower()
        if suffix == ".xlsx":
            dataframe.to_excel(output_path, index=index)
        else:
            dataframe.to_csv(output_path, index=index)
        return self.register_artifact(output_path, artifact_type=artifact_type, notes=notes)

    def copy_file(
        self,
        source_path: str | Path,
        artifact_type: str = "archivo",
        label: str | None = None,
        notes: str = "",
        subdir: str = "copias",
    ) -> Path:
        source = Path(source_path).resolve()
        destination = self.run_dir / subdir / source.name
        ensure_parent(destination)
        shutil.copy2(source, destination)
        return self.register_artifact(destination, artifact_type=artifact_type, label=label, notes=notes)

    def finalize(
        self,
        horizon_results: list[dict[str, Any]],
        *,
        target: str = "",
        variables_temporales: str = "",
        variables_tematicas: str = "",
        transformacion: str = "",
        seleccion_variables: str = "",
        validacion: str = "",
        dataset_periodo: str = "",
        notas_config: str = "",
        estado: str = "corrido",
        comentarios: str = "",
        l_coh: float | None = None,
    ) -> Path:
        results_payload = {
            "run_id": self.run_id,
            "experiment_id": self.experiment_id,
            "family": self.family,
            "model": self.model,
            "created_at": self.created_at,
            "script_path": str(self.script_path),
            "snapshot_path": str(self.snapshot_path),
            "horizon_results": horizon_results,
        }
        results_path = self.save_json(
            results_payload,
            "metricas_horizonte.json",
            artifact_type="metricas",
            notes="Resumen estructurado por horizonte.",
        )

        metadata_payload = {
            "run_id": self.run_id,
            "experiment_id": self.experiment_id,
            "family": self.family,
            "model": self.model,
            "script_path": str(self.script_path),
            "snapshot_path": str(self.snapshot_path),
            "run_dir": str(self.run_dir),
            "parametros_path": str(self.parametros_path) if self.parametros_path else "",
            "created_at": self.created_at,
        }
        self.save_json(
            metadata_payload,
            "metadata_run.json",
            artifact_type="metadata",
            notes="Metadata base del run.",
        )

        self.tracker.log_run(
            run_id=self.run_id,
            experiment_id=self.experiment_id,
            family=self.family,
            model=self.model,
            script_path=self.script_path,
            run_dir=self.run_dir,
            snapshot_path=self.snapshot_path,
            parametros_path=self.parametros_path,
            resultados_path=results_path,
            horizon_results=horizon_results,
            artifacts=self.artifacts,
            target=target,
            variables_temporales=variables_temporales,
            variables_tematicas=variables_tematicas,
            transformacion=transformacion,
            seleccion_variables=seleccion_variables,
            validacion=validacion,
            dataset_periodo=dataset_periodo,
            notas_config=notas_config,
            estado=estado,
            comentarios=comentarios,
            l_coh=l_coh,
            timestamp_run=self.created_at,
        )
        return results_path


class RadarExperimentTracker:
    def __init__(
        self,
        workbook_path: Path = DEFAULT_WORKBOOK,
        runs_dir: Path = DEFAULT_RUNS_DIR,
    ) -> None:
        self.workbook_path = Path(workbook_path).expanduser().resolve()
        self.runs_dir = Path(runs_dir).expanduser().resolve()

    def prepare_workbook(self) -> Path:
        with WorkbookLock(self.workbook_path):
            workbook = load_workbook(self.workbook_path)
            self._prepare_workbook_in_memory(workbook)
            workbook.save(self.workbook_path)
        return self.workbook_path

    def start_run(
        self,
        *,
        run_id: str,
        experiment_id: str,
        family: str,
        model: str,
        script_path: str | Path,
        parametros: dict[str, Any] | None = None,
    ) -> RunContext:
        script = Path(script_path).expanduser().resolve()
        if not script.exists():
            raise FileNotFoundError(f"No existe el script a registrar: {script}")

        created_at = now_text()
        run_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_dir = self.runs_dir / f"{run_id}_{run_stamp}"
        run_dir.mkdir(parents=True, exist_ok=False)

        snapshot_path = run_dir / "scripts" / script.name
        ensure_parent(snapshot_path)
        shutil.copy2(script, snapshot_path)

        context = RunContext(
            tracker=self,
            run_id=run_id,
            experiment_id=experiment_id,
            family=family,
            model=model,
            script_path=script,
            run_dir=run_dir,
            snapshot_path=snapshot_path,
            created_at=created_at,
        )
        context.register_artifact(
            snapshot_path,
            artifact_type="script_snapshot",
            notes="Copia del script ejecutado para reproducibilidad.",
        )

        if parametros is not None:
            parametros_path = run_dir / "parametros_run.json"
            parametros_path.write_text(json_dumps_pretty(parametros), encoding="utf-8")
            context.parametros_path = parametros_path
            context.register_artifact(
                parametros_path,
                artifact_type="parametros",
                notes="Parametros del experimento.",
            )

        return context

    def log_run(
        self,
        *,
        run_id: str,
        experiment_id: str,
        family: str,
        model: str,
        script_path: str | Path,
        run_dir: str | Path,
        snapshot_path: str | Path,
        parametros_path: str | Path | None,
        resultados_path: str | Path | None,
        horizon_results: list[dict[str, Any]],
        artifacts: list[dict[str, str]] | None = None,
        target: str = "",
        variables_temporales: str = "",
        variables_tematicas: str = "",
        transformacion: str = "",
        seleccion_variables: str = "",
        validacion: str = "",
        dataset_periodo: str = "",
        notas_config: str = "",
        estado: str = "corrido",
        comentarios: str = "",
        l_coh: float | None = None,
        timestamp_run: str | None = None,
    ) -> Path:
        if not horizon_results:
            raise ValueError("Se requiere al menos un resultado por horizonte.")

        run_timestamp = timestamp_run or now_text()

        with WorkbookLock(self.workbook_path):
            workbook = load_workbook(self.workbook_path)
            self._prepare_workbook_in_memory(workbook)

            result_ws = workbook[RESULTADOS_SHEET]
            summary_ws = workbook[SUMMARY_SHEET]
            artifacts_ws = workbook[ARTEFACTOS_SHEET]

            result_headers = find_header_map(result_ws)
            summary_headers = find_header_map(summary_ws)

            for horizon_result in horizon_results:
                row_idx = self._ensure_result_row(result_ws, run_id, int(horizon_result["horizonte_sem"]))
                self._write_result_row(
                    worksheet=result_ws,
                    headers=result_headers,
                    row_idx=row_idx,
                    run_id=run_id,
                    experiment_id=experiment_id,
                    family=family,
                    model=model,
                    horizon_result=horizon_result,
                    script_path=script_path,
                    run_dir=run_dir,
                    snapshot_path=snapshot_path,
                    parametros_path=parametros_path,
                    resultados_path=resultados_path,
                    target=target,
                    variables_temporales=variables_temporales,
                    variables_tematicas=variables_tematicas,
                    transformacion=transformacion,
                    seleccion_variables=seleccion_variables,
                    validacion=validacion,
                    dataset_periodo=dataset_periodo,
                    notas_config=notas_config,
                    estado=estado,
                    comentarios=comentarios,
                    timestamp_run=run_timestamp,
                )

            summary_row = self._ensure_summary_row(summary_ws, run_id)
            self._write_summary_row(
                worksheet=summary_ws,
                headers=summary_headers,
                result_headers=result_headers,
                row_idx=summary_row,
                run_id=run_id,
                l_coh=l_coh,
            )

            self._refresh_artifacts_sheet(
                worksheet=artifacts_ws,
                run_id=run_id,
                experiment_id=experiment_id,
                artifacts=artifacts or [],
            )

            workbook.save(self.workbook_path)

        return self.workbook_path

    def _prepare_workbook_in_memory(self, workbook) -> None:
        if RESULTADOS_SHEET not in workbook.sheetnames:
            raise KeyError(f"No existe la hoja '{RESULTADOS_SHEET}' en {self.workbook_path}")
        if SUMMARY_SHEET not in workbook.sheetnames:
            raise KeyError(f"No existe la hoja '{SUMMARY_SHEET}' en {self.workbook_path}")

        result_ws = workbook[RESULTADOS_SHEET]
        summary_ws = workbook[SUMMARY_SHEET]

        self._ensure_extra_columns(
            worksheet=result_ws,
            extra_headers=RESULTADOS_EXTRA_HEADERS,
            style_source_header="Comentarios",
            style_source_data="Comentarios",
        )
        self._ensure_extra_columns(
            worksheet=summary_ws,
            extra_headers=SUMMARY_EXTRA_HEADERS,
            style_source_header="Comentarios",
            style_source_data="Comentarios",
        )

        if ARTEFACTOS_SHEET not in workbook.sheetnames:
            artifacts_ws = workbook.create_sheet(ARTEFACTOS_SHEET)
            self._initialize_artifacts_sheet(artifacts_ws)

    def _ensure_extra_columns(
        self,
        *,
        worksheet,
        extra_headers: list[str],
        style_source_header: str,
        style_source_data: str,
    ) -> None:
        header_map = find_header_map(worksheet)
        header_source_col = header_map[style_source_header]
        data_source_col = header_map[style_source_data]

        for header in extra_headers:
            if header in header_map:
                continue

            new_col = worksheet.max_column + 1
            worksheet.cell(1, new_col).value = header
            copy_cell_format(worksheet.cell(1, header_source_col), worksheet.cell(1, new_col))
            worksheet.column_dimensions[get_column_letter(new_col)].width = DEFAULT_COLUMN_WIDTHS.get(header, 24)

            for row_idx in range(2, worksheet.max_row + 1):
                copy_cell_format(worksheet.cell(row_idx, data_source_col), worksheet.cell(row_idx, new_col))

            header_map[header] = new_col

    def _initialize_artifacts_sheet(self, worksheet) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        body_fill = PatternFill("solid", fgColor="D9EAF7")

        for col_idx, header in enumerate(ARTEFACTOS_HEADERS, start=1):
            cell = worksheet.cell(1, col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 26 if header != "Notas" else 36

        for col_idx in range(1, len(ARTEFACTOS_HEADERS) + 1):
            worksheet.cell(2, col_idx).fill = body_fill

    def _ensure_result_row(self, worksheet, run_id: str, horizon: int) -> int:
        headers = find_header_map(worksheet)
        run_col = headers["Run_ID"]
        horizon_col = headers["Horizonte_sem"]

        for row_idx in range(2, worksheet.max_row + 1):
            run_value = worksheet.cell(row_idx, run_col).value
            horizon_value = worksheet.cell(row_idx, horizon_col).value
            if str(run_value).strip() == run_id and horizon_value == horizon:
                return row_idx

        new_row = worksheet.max_row + 1
        self._copy_row_template(worksheet, source_row=2, target_row=new_row)
        return new_row

    def _ensure_summary_row(self, worksheet, run_id: str) -> int:
        run_col = find_header_map(worksheet)["Run_ID"]

        for row_idx in range(2, worksheet.max_row + 1):
            current = worksheet.cell(row_idx, run_col).value
            if str(current).strip() == run_id:
                return row_idx

        for row_idx in range(2, worksheet.max_row + 1):
            current = worksheet.cell(row_idx, run_col).value
            if current in (None, "") and not row_has_value(worksheet, row_idx):
                self._copy_row_template(worksheet, source_row=2, target_row=row_idx)
                return row_idx

        new_row = worksheet.max_row + 1
        self._copy_row_template(worksheet, source_row=2, target_row=new_row)
        return new_row

    def _copy_row_template(self, worksheet, *, source_row: int, target_row: int) -> None:
        for col_idx in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(source_row, col_idx)
            target_cell = worksheet.cell(target_row, col_idx)
            copy_cell_format(source_cell, target_cell)
            if target_cell.value not in (None, ""):
                target_cell.value = None
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height

    def _write_result_row(
        self,
        *,
        worksheet,
        headers: dict[str, int],
        row_idx: int,
        run_id: str,
        experiment_id: str,
        family: str,
        model: str,
        horizon_result: dict[str, Any],
        script_path: str | Path,
        run_dir: str | Path,
        snapshot_path: str | Path,
        parametros_path: str | Path | None,
        resultados_path: str | Path | None,
        target: str,
        variables_temporales: str,
        variables_tematicas: str,
        transformacion: str,
        seleccion_variables: str,
        validacion: str,
        dataset_periodo: str,
        notas_config: str,
        estado: str,
        comentarios: str,
        timestamp_run: str,
    ) -> None:
        horizon = int(horizon_result["horizonte_sem"])
        overrides = {
            "Target": horizon_result.get("target", target),
            "Variables_temporales": horizon_result.get("variables_temporales", variables_temporales),
            "Variables_tematicas": horizon_result.get("variables_tematicas", variables_tematicas),
            "Transformacion": horizon_result.get("transformacion", transformacion),
            "Seleccion_variables": horizon_result.get("seleccion_variables", seleccion_variables),
            "Validacion": horizon_result.get("validacion", validacion),
            "Dataset_Periodo": horizon_result.get("dataset_periodo", dataset_periodo),
            "Notas_config": horizon_result.get("notas_config", notas_config),
            "Estado": horizon_result.get("estado", estado),
            "Comentarios": horizon_result.get("comentarios", comentarios),
        }

        values = {
            "Fecha": timestamp_run,
            "Run_ID": run_id,
            "Experiment_ID": experiment_id,
            "Familia_modelo": family,
            "Modelo": model,
            "Horizonte_sem": horizon,
            "Target": overrides["Target"],
            "Variables_temporales": overrides["Variables_temporales"],
            "Variables_tematicas": overrides["Variables_tematicas"],
            "Transformacion": overrides["Transformacion"],
            "Seleccion_variables": overrides["Seleccion_variables"],
            "Validacion": overrides["Validacion"],
            "Dataset_Periodo": overrides["Dataset_Periodo"],
            "Notas_config": overrides["Notas_config"],
            "L_num": horizon_result.get("l_num"),
            "L_trend": horizon_result.get("l_trend"),
            "L_risk": horizon_result.get("l_risk"),
            "L_tol": horizon_result.get("l_tol"),
            "MAE": horizon_result.get("mae"),
            "RMSE": horizon_result.get("rmse"),
            "Direccion_accuracy": horizon_result.get("direccion_accuracy"),
            "Deteccion_caidas": horizon_result.get("deteccion_caidas"),
            "Estado": overrides["Estado"],
            "Comentarios": overrides["Comentarios"],
            "Script_nombre": Path(script_path).name,
            "Script_ruta": prefer_repo_relative(script_path),
            "Run_dir": prefer_repo_relative(run_dir),
            "Snapshot_script": prefer_repo_relative(snapshot_path),
            "Parametros_ruta_json": prefer_repo_relative(parametros_path),
            "Resultados_ruta_json": prefer_repo_relative(resultados_path),
            "Timestamp_run": timestamp_run,
        }

        for header, value in values.items():
            worksheet.cell(row_idx, headers[header]).value = value

        worksheet.cell(row_idx, headers["alpha"]).value = "=CONFIG_REFERENCIA!$E$3"
        worksheet.cell(row_idx, headers["beta"]).value = "=CONFIG_REFERENCIA!$E$4"
        worksheet.cell(row_idx, headers["gamma"]).value = "=CONFIG_REFERENCIA!$E$5"
        worksheet.cell(row_idx, headers["delta"]).value = "=CONFIG_REFERENCIA!$E$6"
        worksheet.cell(row_idx, headers["w_h"]).value = (
            f'=IF($F{row_idx}="","",INDEX(CONFIG_REFERENCIA!$B$3:$B$6,'
            f'MATCH($F{row_idx},CONFIG_REFERENCIA!$A$3:$A$6,0)))'
        )
        worksheet.cell(row_idx, headers["Loss_h"]).value = (
            f'=IF(COUNTA($T{row_idx}:$W{row_idx})<4,"",'
            f'$S{row_idx}*($O{row_idx}*$T{row_idx}+$P{row_idx}*$U{row_idx}+'
            f'$Q{row_idx}*$V{row_idx}+$R{row_idx}*$W{row_idx}))'
        )

    def _summary_lookup_formula(
        self,
        *,
        result_col: int,
        result_run_col: int,
        row_idx: int,
    ) -> str:
        result_col_letter = get_column_letter(result_col)
        run_col_letter = get_column_letter(result_run_col)
        return (
            f'=IF($A{row_idx}="","",IFERROR(INDEX({RESULTADOS_SHEET}!${result_col_letter}:${result_col_letter},'
            f'MATCH($A{row_idx},{RESULTADOS_SHEET}!${run_col_letter}:${run_col_letter},0)),""))'
        )

    def _write_summary_row(
        self,
        *,
        worksheet,
        headers: dict[str, int],
        result_headers: dict[str, int],
        row_idx: int,
        run_id: str,
        l_coh: float | None,
    ) -> None:
        result_run_col = result_headers["Run_ID"]

        worksheet.cell(row_idx, headers["Run_ID"]).value = run_id
        worksheet.cell(row_idx, headers["Experiment_ID"]).value = self._summary_lookup_formula(
            result_col=result_headers["Experiment_ID"],
            result_run_col=result_run_col,
            row_idx=row_idx,
        )
        worksheet.cell(row_idx, headers["Modelo"]).value = self._summary_lookup_formula(
            result_col=result_headers["Modelo"],
            result_run_col=result_run_col,
            row_idx=row_idx,
        )
        worksheet.cell(row_idx, headers["Familia"]).value = self._summary_lookup_formula(
            result_col=result_headers["Familia_modelo"],
            result_run_col=result_run_col,
            row_idx=row_idx,
        )
        worksheet.cell(row_idx, headers["Horizons_loaded"]).value = (
            f'=IF($A{row_idx}="","",COUNTIF({RESULTADOS_SHEET}!$B:$B,$A{row_idx}))'
        )
        worksheet.cell(row_idx, headers["Sum_weighted_Loss_h"]).value = (
            f'=IF($A{row_idx}="","",SUMIF({RESULTADOS_SHEET}!$B:$B,$A{row_idx},{RESULTADOS_SHEET}!$X:$X))'
        )
        if l_coh is not None:
            worksheet.cell(row_idx, headers["L_coh"]).value = l_coh
        worksheet.cell(row_idx, headers["eta"]).value = "=CONFIG_REFERENCIA!$E$7"
        worksheet.cell(row_idx, headers["L_total_Radar"]).value = (
            f'=IF($A{row_idx}="","",$F{row_idx}+N($G{row_idx})*$H{row_idx})'
        )
        worksheet.cell(row_idx, headers["Avg_MAE"]).value = (
            f'=IF($A{row_idx}="","",IFERROR(AVERAGEIF({RESULTADOS_SHEET}!$B:$B,$A{row_idx},'
            f'{RESULTADOS_SHEET}!$Y:$Y),""))'
        )
        worksheet.cell(row_idx, headers["Avg_RMSE"]).value = (
            f'=IF($A{row_idx}="","",IFERROR(AVERAGEIF({RESULTADOS_SHEET}!$B:$B,$A{row_idx},'
            f'{RESULTADOS_SHEET}!$Z:$Z),""))'
        )
        worksheet.cell(row_idx, headers["Dir_acc_prom"]).value = (
            f'=IF($A{row_idx}="","",IFERROR(AVERAGEIF({RESULTADOS_SHEET}!$B:$B,$A{row_idx},'
            f'{RESULTADOS_SHEET}!$AA:$AA),""))'
        )
        worksheet.cell(row_idx, headers["Det_caidas_prom"]).value = (
            f'=IF($A{row_idx}="","",IFERROR(AVERAGEIF({RESULTADOS_SHEET}!$B:$B,$A{row_idx},'
            f'{RESULTADOS_SHEET}!$AB:$AB),""))'
        )
        worksheet.cell(row_idx, headers["Ranking"]).value = (
            f'=IF($I{row_idx}="","",COUNTIFS($I$2:$I$1000,"<"&$I{row_idx})+1)'
        )

        for header in SUMMARY_EXTRA_HEADERS:
            worksheet.cell(row_idx, headers[header]).value = self._summary_lookup_formula(
                result_col=result_headers[header],
                result_run_col=result_run_col,
                row_idx=row_idx,
            )

    def _refresh_artifacts_sheet(
        self,
        *,
        worksheet,
        run_id: str,
        experiment_id: str,
        artifacts: list[dict[str, str]],
    ) -> None:
        run_col = 2
        rows_to_delete: list[int] = []
        for row_idx in range(2, worksheet.max_row + 1):
            current = worksheet.cell(row_idx, run_col).value
            if str(current).strip() == run_id:
                rows_to_delete.append(row_idx)
        for row_idx in reversed(rows_to_delete):
            worksheet.delete_rows(row_idx, 1)

        body_template_row = 2 if worksheet.max_row >= 2 else None
        for artifact in artifacts:
            new_row = worksheet.max_row + 1
            if body_template_row is not None and new_row != body_template_row:
                self._copy_row_template(worksheet, source_row=body_template_row, target_row=new_row)

            worksheet.cell(new_row, 1).value = now_text()
            worksheet.cell(new_row, 2).value = run_id
            worksheet.cell(new_row, 3).value = experiment_id
            worksheet.cell(new_row, 4).value = artifact.get("artifact_type", "")
            worksheet.cell(new_row, 5).value = artifact.get("label", "")
            worksheet.cell(new_row, 6).value = prefer_repo_relative(artifact.get("path", ""))
            worksheet.cell(new_row, 7).value = artifact.get("notes", "")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Prepara el workbook de experimentos del radar para tracking automatico.",
    )
    parser.add_argument(
        "--prepare-workbook",
        action="store_true",
        help="Agrega columnas de trazabilidad y la hoja RUN_ARTEFACTOS al workbook.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Ruta del workbook. Default: {DEFAULT_WORKBOOK}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    tracker = RadarExperimentTracker(workbook_path=args.workbook)
    if args.prepare_workbook:
        updated_path = tracker.prepare_workbook()
        print(f"Workbook preparado: {updated_path}")
    else:
        raise SystemExit("Usa --prepare-workbook para preparar el Excel.")


if __name__ == "__main__":
    main()
