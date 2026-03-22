#!/usr/bin/env python3
"""
Genera variables rezagadas (lags) para datasets semanales de ML.

Comportamiento:
- Procesa todas las hojas del Excel excepto `README`.
- Ordena cronologicamente cada hoja.
- Genera lags `lag1..lagN` para columnas numericas no meta y no flag.
- Genera `delta1` para los targets disponibles.
- Renombra hojas `ML_Ready_*` a `ML_Lagged_*`.

Ejemplos:
    python3 Scripts/NLP_Data_Procesing/Lageado_Datos_ML.py
    python3 Scripts/NLP_Data_Procesing/Lageado_Datos_ML.py \
        --input Datos_Modelo_ML/datos_ML_3.xlsx \
        --output Datos_Modelo_ML/datos_ML_4.xlsx \
        --lag-weeks 4
    python3 Scripts/NLP_Data_Procesing/Lageado_Datos_ML.py \
        --input Datos_Modelo_ML/datos_ML_3.xlsx \
        --output Datos_Modelo_ML/datos_ML_4.xlsx \
        --lags 1 2 3 4
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pandas.api.types import is_bool_dtype, is_numeric_dtype


BASE_DIR = Path("/home/emilio/Documentos/RAdAR")
DEFAULT_INPUT = BASE_DIR / "Datos_Modelo_ML" / "datos_ML_3.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "Datos_Modelo_ML" / "datos_ML_4.xlsx"
README_SHEET = "README"

META_COLUMNS = {
    "fecha_inicio_semana",
    "semana_iso",
    "iso_year",
    "iso_week",
    "iso_yearweek_num",
}
TARGET_COLUMNS = ("target_serie_3e", "target_serie_4e")
FLAG_PREFIXES = ("flag_",)
FLAG_COLUMNS = {"has_full_pmi_features"}
LAG_OR_DELTA_PATTERN = re.compile(r"_(lag\d+|delta\d+)$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera lags semanales para todas las hojas de datos de un archivo Excel.",
    )
    parser.add_argument(
        "-i",
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Archivo Excel de entrada. Default: {DEFAULT_INPUT}",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Archivo Excel de salida. Default: {DEFAULT_OUTPUT}",
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "-w",
        "--lag-weeks",
        type=int,
        choices=(1, 2, 3, 4),
        help="Numero maximo de semanas de lag a generar (1, 2, 3 o 4). Produce desde lag1 hasta lagN.",
    )
    group.add_argument(
        "--lags",
        type=int,
        nargs="+",
        choices=(1, 2, 3, 4),
        help="Lista explicita de lags a generar. Ejemplo: --lags 1 2 3 4",
    )
    return parser.parse_args()


def resolve_path(path: Path) -> Path:
    return path.expanduser()


def resolve_lag_values(args: argparse.Namespace) -> list[int]:
    if args.lags:
        return sorted(set(args.lags))
    if args.lag_weeks:
        return list(range(1, args.lag_weeks + 1))
    return [1, 2, 3]


def load_data_sheets(input_path: Path) -> dict[str, pd.DataFrame]:
    workbook = pd.ExcelFile(input_path)
    sheets: dict[str, pd.DataFrame] = {}

    for sheet_name in workbook.sheet_names:
        if sheet_name.strip().lower() == README_SHEET.lower():
            continue
        sheets[sheet_name] = pd.read_excel(input_path, sheet_name=sheet_name)

    if not sheets:
        raise SystemExit(f"No se encontraron hojas de datos en {input_path}")

    return sheets


def sort_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    if "iso_yearweek_num" in out.columns:
        out = out.sort_values("iso_yearweek_num", kind="stable")
    else:
        sort_columns = [col for col in ("iso_year", "iso_week", "fecha_inicio_semana") if col in out.columns]
        if sort_columns:
            out = out.sort_values(sort_columns, kind="stable")

    out.reset_index(drop=True, inplace=True)
    return out


def is_laggable_column(column_name: str, series: pd.Series) -> bool:
    if column_name in META_COLUMNS:
        return False
    if column_name in FLAG_COLUMNS:
        return False
    if column_name.startswith(FLAG_PREFIXES):
        return False
    if LAG_OR_DELTA_PATTERN.search(column_name):
        return False
    if is_bool_dtype(series):
        return False
    return is_numeric_dtype(series)


def detect_laggable_columns(df: pd.DataFrame) -> list[str]:
    return [column for column in df.columns if is_laggable_column(column, df[column])]


def output_sheet_name(input_sheet_name: str) -> str:
    if input_sheet_name.startswith("ML_Ready_"):
        return input_sheet_name.replace("ML_Ready_", "ML_Lagged_", 1)
    if input_sheet_name.startswith("ML_Lagged_"):
        return input_sheet_name
    return f"{input_sheet_name}_Lagged"


def generate_lags(
    df: pd.DataFrame,
    lag_values: list[int],
) -> tuple[pd.DataFrame, dict[str, int | str]]:
    ordered = sort_dataframe(df)
    laggable_columns = detect_laggable_columns(ordered)

    extra_parts = []
    lag_columns_created = 0
    delta_columns_created = 0

    for column in laggable_columns:
        for lag in lag_values:
            lag_name = f"{column}_lag{lag}"
            extra_parts.append(ordered[column].shift(lag).rename(lag_name))
            lag_columns_created += 1

    for target in TARGET_COLUMNS:
        if target not in ordered.columns:
            continue
        delta_name = f"{target}_delta1"
        extra_parts.append(ordered[target].diff(1).rename(delta_name))
        delta_columns_created += 1

    lagged = pd.concat([ordered] + extra_parts, axis=1) if extra_parts else ordered.copy()

    summary = {
        "rows": len(lagged),
        "original_columns": len(df.columns),
        "laggable_columns": len(laggable_columns),
        "lag_columns_created": lag_columns_created,
        "delta_columns_created": delta_columns_created,
        "usable_rows_after_max_lag": max(len(lagged) - max(lag_values, default=0), 0),
    }
    return lagged, summary


def build_readme(
    input_path: Path,
    output_path: Path,
    lag_values: list[int],
    summaries: list[dict[str, int | str]],
) -> pd.DataFrame:
    lag_text = ", ".join(f"lag{lag}" for lag in lag_values)
    rows: list[tuple[str, str]] = [
        (
            "Objetivo",
            "Generar variables lag para cada hoja de datos del archivo de entrada usando la estructura real del Excel.",
        ),
        ("Archivo fuente", str(input_path)),
        ("Archivo salida", str(output_path)),
        ("Semanas de lag", ", ".join(str(lag) for lag in lag_values)),
        ("Lags generados", lag_text),
        ("Hojas procesadas", ", ".join(str(item["input_sheet"]) for item in summaries)),
        (
            "Regla de lag",
            "Se aplican lags a columnas numericas no meta, no flag y que no sean lags/deltas previos.",
        ),
        (
            "Regla de delta",
            "Se genera delta1 para target_serie_3e y target_serie_4e solo si existen en la hoja.",
        ),
    ]

    for item in summaries:
        sheet_label = str(item["input_sheet"])
        rows.extend(
            [
                (f"{sheet_label} -> hoja salida", str(item["output_sheet"])),
                (f"{sheet_label} -> filas", str(item["rows"])),
                (f"{sheet_label} -> columnas originales", str(item["original_columns"])),
                (f"{sheet_label} -> columnas con lag", str(item["laggable_columns"])),
                (f"{sheet_label} -> nuevas columnas lag", str(item["lag_columns_created"])),
                (f"{sheet_label} -> nuevas columnas delta", str(item["delta_columns_created"])),
                (
                    f"{sheet_label} -> filas usables tras lag maximo",
                    str(item["usable_rows_after_max_lag"]),
                ),
            ]
        )

    return pd.DataFrame(rows, columns=["Campo", "Valor"])


def apply_format(writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame) -> None:
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font

    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 36)


def write_output(
    output_path: Path,
    sheets: dict[str, pd.DataFrame],
    readme: pd.DataFrame,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_format(writer, sheet_name, dataframe)

        readme.to_excel(writer, sheet_name=README_SHEET, index=False)
        apply_format(writer, README_SHEET, readme)


def main() -> None:
    args = parse_args()
    input_path = resolve_path(args.input)
    output_path = resolve_path(args.output)
    lag_values = resolve_lag_values(args)

    if not input_path.exists():
        raise SystemExit(f"No existe el archivo de entrada: {input_path}")

    print("=" * 72)
    print("GENERADOR DE LAGS")
    print("=" * 72)
    print(f"Input:      {input_path}")
    print(f"Output:     {output_path}")
    print(f"Lags:       {', '.join(str(lag) for lag in lag_values)}")

    input_sheets = load_data_sheets(input_path)
    output_sheets: dict[str, pd.DataFrame] = {}
    summaries: list[dict[str, int | str]] = []

    for input_sheet_name, dataframe in input_sheets.items():
        lagged_df, summary = generate_lags(dataframe, lag_values)
        lagged_sheet_name = output_sheet_name(input_sheet_name)

        if lagged_sheet_name in output_sheets:
            raise SystemExit(
                f"Conflicto de nombres de hoja de salida: {lagged_sheet_name}. "
                "Ajusta los nombres de las hojas de entrada."
            )

        output_sheets[lagged_sheet_name] = lagged_df
        summary["input_sheet"] = input_sheet_name
        summary["output_sheet"] = lagged_sheet_name
        summaries.append(summary)

        print(
            f"- {input_sheet_name} -> {lagged_sheet_name}: "
            f"{summary['lag_columns_created']} lags, "
            f"{summary['delta_columns_created']} deltas, "
            f"{summary['rows']} filas"
        )

    readme = build_readme(
        input_path=input_path,
        output_path=output_path,
        lag_values=lag_values,
        summaries=summaries,
    )
    write_output(output_path, output_sheets, readme)

    print(f"Archivo generado: {output_path}")
    print("=" * 72)


if __name__ == "__main__":
    main()
