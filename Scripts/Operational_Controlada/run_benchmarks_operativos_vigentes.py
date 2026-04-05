#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
SCRIPTS_DIR = ROOT_DIR / "Scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from common_runtime_logging import log_event


REGISTRY_PATH = ROOT_DIR / "Experimentos" / "registro_operacion_controlada_radar.json"
GRID_WORKBOOK_PATH = ROOT_DIR / "Experimentos" / "grid_experimentos_radar.xlsx"
MASTER_WORKBOOK_PATH = ROOT_DIR / "Experimentos" / "tabla_maestra_experimentos_radar.xlsx"
INVENTORY_PATH = ROOT_DIR / "Experimentos" / "inventario_experimentos_radar.json"
DEFAULT_REPORT_JSON_PATH = (
    ROOT_DIR / "Experimentos" / "auditoria_benchmarks_operativos_controlados.json"
)
DEFAULT_REPORT_MD_PATH = (
    ROOT_DIR / "Experimentos" / "auditoria_benchmarks_operativos_controlados.md"
)
DEFAULT_PYTHON_BIN = Path("/home/emilio/anaconda3/envs/radar-exp-py311/bin/python")
REQUIRED_RUN_ARTIFACTS = [
    "metadata_run.json",
    "parametros_run.json",
    "metricas_horizonte.json",
    "resumen_modeling_horizontes.json",
]
REQUIRED_PREDICTION_COLUMNS = {"y_true", "y_pred"}
ACCEPTED_DATE_COLUMNS = {"fecha", "fecha_inicio_semana"}

FROZEN_EXECUTION_SPECS = {
    "benchmark_numerico_puro_vigente": {
        "runner": ROOT_DIR / "Scripts" / "Modeling" / "run_e1_ridge_clean.py",
        "args": [
            "--reference-run-id",
            "E1_v5_clean",
            "--target-mode",
            "nivel",
            "--feature-mode",
            "corr",
            "--lags",
            "1,2,3,4,5,6",
            "--transform-mode",
            "standard",
            "--initial-train-size",
            "40",
            "--horizons",
            "1,2,3,4",
        ],
    },
    "benchmark_operativo_riesgo_vigente": {
        "runner": ROOT_DIR / "Scripts" / "Modeling" / "run_e9_stacking.py",
        "args": [
            "--reference-run-id",
            "E1_v5_clean",
            "--extra-reference-run-ids",
            "E5_v4_clean,E3_v2_clean,E2_v3_clean,E7_v3_clean,E9_v2_clean",
            "--hypothesis-note",
            "benchmark_operativo_riesgo_controlado",
            "--table-path",
            str(ROOT_DIR / "Experimentos" / "tabla_maestra_experimentos_radar_e9_curada.xlsx"),
            "--horizons",
            "1,2,3,4",
            "--initial-train-size",
            "12",
            "--meta-model",
            "huber",
            "--alpha-grid-size",
            "40",
            "--inner-splits",
            "3",
            "--alpha-selection-metric",
            "mae",
            "--use-only-complete-rows",
        ],
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Inspector/dispatcher de benchmarks operativos vigentes del Radar.",
    )
    parser.add_argument(
        "--registry-path",
        type=Path,
        default=REGISTRY_PATH,
        help="JSON canonico de benchmarks operativos vigentes.",
    )
    parser.add_argument(
        "--benchmark-id",
        choices=tuple(FROZEN_EXECUTION_SPECS),
        help="Benchmark a inspeccionar o ejecutar.",
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Ejecuta el runner subyacente con los parametros congelados.",
    )
    parser.add_argument(
        "--run-id",
        help="Run_ID a usar cuando se ejecuta un benchmark congelado.",
    )
    parser.add_argument(
        "--python-bin",
        type=Path,
        default=DEFAULT_PYTHON_BIN,
        help="Python a usar para ejecutar el runner congelado.",
    )
    parser.add_argument(
        "--grid-workbook-path",
        type=Path,
        default=GRID_WORKBOOK_PATH,
        help="Workbook del tracker operativo con RUN_SUMMARY/RESULTADOS_GRID/RUN_ARTEFACTOS.",
    )
    parser.add_argument(
        "--master-workbook-path",
        type=Path,
        default=MASTER_WORKBOOK_PATH,
        help="Workbook maestro con hojas canonicas de benchmarks y capas vigentes.",
    )
    parser.add_argument(
        "--inventory-path",
        type=Path,
        default=INVENTORY_PATH,
        help="Inventario JSON canonico del proyecto.",
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        help="Audita integridad operativa de benchmarks y genera reporte JSON/MD.",
    )
    parser.add_argument(
        "--report-json-path",
        type=Path,
        default=DEFAULT_REPORT_JSON_PATH,
        help="Salida JSON de la auditoria operativa.",
    )
    parser.add_argument(
        "--report-md-path",
        type=Path,
        default=DEFAULT_REPORT_MD_PATH,
        help="Salida Markdown de la auditoria operativa.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Si hay gaps en la auditoria, termina con error.",
    )
    return parser.parse_args()


def load_registry(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def load_inventory(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def build_command(benchmark_id: str, run_id: str, python_bin: Path) -> list[str]:
    spec = FROZEN_EXECUTION_SPECS[benchmark_id]
    return [str(python_bin), str(spec["runner"]), "--run-id", run_id, *spec["args"]]


def load_sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header = [
        str(value).strip().lower() if value is not None else ""
        for value in rows[0]
    ]
    return [
        {header[idx]: row[idx] for idx in range(len(header))}
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def inspect_prediction_file(prediction_path: Path) -> dict[str, Any]:
    result = {
        "path": str(prediction_path),
        "exists": prediction_path.exists(),
        "rows": 0,
        "columns": [],
        "has_required_columns": False,
        "has_date_column": False,
    }
    if not prediction_path.exists():
        return result
    with prediction_path.open("r", encoding="utf-8") as handle:
        lines = handle.read().splitlines()
    if not lines:
        return result
    header = lines[0].split(",")
    result["rows"] = max(len(lines) - 1, 0)
    result["columns"] = header
    result["has_required_columns"] = REQUIRED_PREDICTION_COLUMNS.issubset(set(header))
    result["has_date_column"] = bool(set(header) & ACCEPTED_DATE_COLUMNS)
    return result


def validate_benchmark(
    benchmark: dict[str, Any],
    inventory: dict[str, Any],
    grid_rows: dict[str, list[dict[str, Any]]],
    master_rows: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    run_id = benchmark["run_id"]
    run_dir = Path(benchmark["run_dir"])
    gaps: list[str] = []

    artifact_status = {}
    for filename in REQUIRED_RUN_ARTIFACTS:
        exists = (run_dir / filename).exists()
        artifact_status[filename] = exists
        if not exists:
            gaps.append(f"faltante artefacto base: {filename}")

    predictions = {}
    for horizon in (1, 2, 3, 4):
        prediction_info = inspect_prediction_file(run_dir / f"predicciones_h{horizon}.csv")
        predictions[f"H{horizon}"] = prediction_info
        if not prediction_info["exists"]:
            gaps.append(f"faltante predicciones_h{horizon}.csv")
        elif not prediction_info["has_required_columns"]:
            gaps.append(f"predicciones_h{horizon}.csv sin columnas y_true/y_pred")
        elif not prediction_info["has_date_column"]:
            gaps.append(f"predicciones_h{horizon}.csv sin fecha canonica")

    snapshot_dir = run_dir / "scripts"
    snapshot_files = sorted(path.name for path in snapshot_dir.glob("*.py")) if snapshot_dir.exists() else []
    if not snapshot_files:
        gaps.append("snapshot de scripts ausente")

    run_summary_rows = [
        row for row in grid_rows["RUN_SUMMARY"] if str(row.get("run_id", "")) == run_id
    ]
    resultados_rows = [
        row for row in grid_rows["RESULTADOS_GRID"] if str(row.get("run_id", "")) == run_id
    ]
    artefact_rows = [
        row for row in grid_rows["RUN_ARTEFACTOS"] if str(row.get("run_id", "")) == run_id
    ]
    if len(run_summary_rows) != 1:
        gaps.append(f"RUN_SUMMARY esperado=1 observado={len(run_summary_rows)}")
    observed_horizons = sorted(
        {
            int(row["horizonte_sem"])
            for row in resultados_rows
            if row.get("horizonte_sem") is not None
        }
    )
    if observed_horizons != [1, 2, 3, 4]:
        gaps.append(f"RESULTADOS_GRID horizontes incompletos: {observed_horizons}")
    if not artefact_rows:
        gaps.append("RUN_ARTEFACTOS sin filas para el benchmark")

    inventory_benchmark_rows = [
        row
        for row in inventory.get("benchmarks_operativos_vigentes", [])
        if row.get("benchmark_id") == benchmark["benchmark_id"]
        and row.get("run_id") == run_id
    ]
    inventory_run_rows = [
        row for row in inventory.get("canonical_runs", []) if row.get("run_id") == run_id
    ]
    if len(inventory_benchmark_rows) != 1:
        gaps.append(
            "inventario JSON no refleja exactamente una fila de benchmark operativo vigente"
        )
    if not inventory_run_rows:
        gaps.append("inventario JSON no contiene el run como canonico")

    master_benchmark_rows = [
        row
        for row in master_rows["benchmarks_operativos_vigentes"]
        if str(row.get("benchmark_id", "")) == benchmark["benchmark_id"]
        and str(row.get("run_id", "")) == run_id
    ]
    if len(master_benchmark_rows) != 1:
        gaps.append(
            "tabla_maestra_experimentos_radar.xlsx no refleja exactamente una fila de benchmark vigente"
        )

    return {
        "benchmark_id": benchmark["benchmark_id"],
        "run_id": run_id,
        "plano_funcional": benchmark.get("plano_funcional"),
        "run_dir": str(run_dir),
        "runner_script_canonico": benchmark.get("runner_script_canonico"),
        "artefactos_base": artifact_status,
        "predicciones_por_horizonte": predictions,
        "snapshot_scripts": snapshot_files,
        "grid_alignment": {
            "run_summary_rows": len(run_summary_rows),
            "resultados_grid_rows": len(resultados_rows),
            "resultados_grid_horizontes": observed_horizons,
            "run_artefactos_rows": len(artefact_rows),
        },
        "inventory_alignment": {
            "benchmark_rows": len(inventory_benchmark_rows),
            "canonical_run_rows": len(inventory_run_rows),
        },
        "master_alignment": {
            "benchmark_rows": len(master_benchmark_rows),
        },
        "gaps": gaps,
        "estado_integridad": "ok" if not gaps else "gaps_detectados",
    }


def build_validation_report(
    registry: dict[str, Any],
    registry_path: Path,
    inventory: dict[str, Any],
    inventory_path: Path,
    grid_workbook_path: Path,
    master_workbook_path: Path,
) -> dict[str, Any]:
    grid_rows = {
        "RUN_SUMMARY": load_sheet_rows(grid_workbook_path, "RUN_SUMMARY"),
        "RESULTADOS_GRID": load_sheet_rows(grid_workbook_path, "RESULTADOS_GRID"),
        "RUN_ARTEFACTOS": load_sheet_rows(grid_workbook_path, "RUN_ARTEFACTOS"),
    }
    master_rows = {
        "benchmarks_operativos_vigentes": load_sheet_rows(
            master_workbook_path,
            "benchmarks_operativos_vigentes",
        ),
        "capas_radar_vigentes": load_sheet_rows(master_workbook_path, "capas_radar_vigentes"),
        "tabla_funcional_dual_vigente": load_sheet_rows(
            master_workbook_path,
            "tabla_funcional_dual_vigente",
        ),
        "politica_funcional_dual_vigente": load_sheet_rows(
            master_workbook_path,
            "politica_funcional_dual_vigente",
        ),
    }
    benchmark_reports = [
        validate_benchmark(benchmark, inventory, grid_rows, master_rows)
        for benchmark in registry.get("benchmarks_operativos_vigentes", [])
    ]
    dual_policy_gaps: list[str] = []
    registry_dual_table_rows = registry.get("tabla_funcional_dual_vigente", [])
    registry_dual_policy_rows = registry.get("politica_funcional_dual_vigente", [])
    inventory_dual_table_rows = inventory.get("tabla_funcional_dual_vigente", [])
    inventory_dual_policy_rows = inventory.get("politica_funcional_dual_vigente", [])
    master_dual_table_rows = master_rows["tabla_funcional_dual_vigente"]
    master_dual_policy_rows = master_rows["politica_funcional_dual_vigente"]
    if len(registry_dual_table_rows) != 20:
        dual_policy_gaps.append(
            f"registro_operacion_controlada_radar.json tabla_funcional_dual_vigente esperado=20 observado={len(registry_dual_table_rows)}"
        )
    if len(registry_dual_policy_rows) != 6:
        dual_policy_gaps.append(
            f"registro_operacion_controlada_radar.json politica_funcional_dual_vigente esperado=6 observado={len(registry_dual_policy_rows)}"
        )
    if len(inventory_dual_table_rows) != 20:
        dual_policy_gaps.append(
            f"inventario_experimentos_radar.json tabla_funcional_dual_vigente esperado=20 observado={len(inventory_dual_table_rows)}"
        )
    if len(inventory_dual_policy_rows) != 6:
        dual_policy_gaps.append(
            f"inventario_experimentos_radar.json politica_funcional_dual_vigente esperado=6 observado={len(inventory_dual_policy_rows)}"
        )
    if len(master_dual_table_rows) != 20:
        dual_policy_gaps.append(
            f"tabla_maestra_experimentos_radar.xlsx tabla_funcional_dual_vigente esperado=20 observado={len(master_dual_table_rows)}"
        )
    if len(master_dual_policy_rows) != 6:
        dual_policy_gaps.append(
            f"tabla_maestra_experimentos_radar.xlsx politica_funcional_dual_vigente esperado=6 observado={len(master_dual_policy_rows)}"
        )

    gaps_total = sum(len(item["gaps"]) for item in benchmark_reports) + len(dual_policy_gaps)
    return {
        "registry_path": str(registry_path),
        "inventory_path": str(inventory_path),
        "grid_workbook_path": str(grid_workbook_path),
        "master_workbook_path": str(master_workbook_path),
        "benchmark_count": len(benchmark_reports),
        "summary": {
            "benchmarks_ok": sum(item["estado_integridad"] == "ok" for item in benchmark_reports),
            "benchmarks_con_gaps": sum(
                item["estado_integridad"] != "ok" for item in benchmark_reports
            ),
            "dual_policy_gaps": len(dual_policy_gaps),
            "gaps_total": gaps_total,
            "estado_global": "ok" if gaps_total == 0 else "gaps_detectados",
        },
        "benchmarks": benchmark_reports,
        "dual_policy_alignment": {
            "registry_dual_table_rows": len(registry_dual_table_rows),
            "registry_dual_policy_rows": len(registry_dual_policy_rows),
            "inventory_dual_table_rows": len(inventory_dual_table_rows),
            "inventory_dual_policy_rows": len(inventory_dual_policy_rows),
            "master_dual_table_rows": len(master_dual_table_rows),
            "master_dual_policy_rows": len(master_dual_policy_rows),
            "gaps": dual_policy_gaps,
            "estado_integridad": "ok" if not dual_policy_gaps else "gaps_detectados",
        },
    }


def save_validation_report(report: dict[str, Any], json_path: Path, md_path: Path) -> None:
    json_path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    lines = [
        "# Auditoria de Benchmarks Operativos Controlados",
        "",
        f"- Estado global: `{report['summary']['estado_global']}`",
        f"- Benchmarks auditados: `{report['benchmark_count']}`",
        f"- Benchmarks sin gaps: `{report['summary']['benchmarks_ok']}`",
        f"- Benchmarks con gaps: `{report['summary']['benchmarks_con_gaps']}`",
        f"- Gaps politica dual: `{report['summary']['dual_policy_gaps']}`",
        f"- Gaps totales: `{report['summary']['gaps_total']}`",
        "",
        "## Politica Dual Canonica",
        "",
        f"- `estado_integridad`: `{report['dual_policy_alignment']['estado_integridad']}`",
        f"- `registry tabla_funcional_dual_vigente`: `{report['dual_policy_alignment']['registry_dual_table_rows']}` fila(s)",
        f"- `registry politica_funcional_dual_vigente`: `{report['dual_policy_alignment']['registry_dual_policy_rows']}` fila(s)",
        f"- `inventory tabla_funcional_dual_vigente`: `{report['dual_policy_alignment']['inventory_dual_table_rows']}` fila(s)",
        f"- `inventory politica_funcional_dual_vigente`: `{report['dual_policy_alignment']['inventory_dual_policy_rows']}` fila(s)",
        f"- `master tabla_funcional_dual_vigente`: `{report['dual_policy_alignment']['master_dual_table_rows']}` fila(s)",
        f"- `master politica_funcional_dual_vigente`: `{report['dual_policy_alignment']['master_dual_policy_rows']}` fila(s)",
    ]
    if report["dual_policy_alignment"]["gaps"]:
        lines.append("- Gaps detectados:")
        lines.extend([f"  - {gap}" for gap in report["dual_policy_alignment"]["gaps"]])
    else:
        lines.append("- Gaps detectados: ninguno")
    lines.append("")
    for benchmark in report["benchmarks"]:
        lines.extend(
            [
                f"## {benchmark['benchmark_id']}",
                "",
                f"- `run_id`: `{benchmark['run_id']}`",
                f"- `plano_funcional`: `{benchmark['plano_funcional']}`",
                f"- `estado_integridad`: `{benchmark['estado_integridad']}`",
                f"- `run_dir`: `{benchmark['run_dir']}`",
                f"- `runner_script_canonico`: `{benchmark['runner_script_canonico']}`",
                f"- `RUN_SUMMARY`: `{benchmark['grid_alignment']['run_summary_rows']}` fila(s)",
                f"- `RESULTADOS_GRID`: `{benchmark['grid_alignment']['resultados_grid_rows']}` fila(s), horizontes={benchmark['grid_alignment']['resultados_grid_horizontes']}",
                f"- `RUN_ARTEFACTOS`: `{benchmark['grid_alignment']['run_artefactos_rows']}` fila(s)",
                f"- `inventory benchmark rows`: `{benchmark['inventory_alignment']['benchmark_rows']}`",
                f"- `inventory canonical run rows`: `{benchmark['inventory_alignment']['canonical_run_rows']}`",
                f"- `master benchmark rows`: `{benchmark['master_alignment']['benchmark_rows']}`",
                f"- `snapshot_scripts`: `{len(benchmark['snapshot_scripts'])}`",
            ]
        )
        for horizon, prediction_info in benchmark["predicciones_por_horizonte"].items():
            lines.append(
                f"- `{horizon}`: exists={prediction_info['exists']} rows={prediction_info['rows']} required_cols={prediction_info['has_required_columns']} date_col={prediction_info['has_date_column']}"
            )
        if benchmark["gaps"]:
            lines.append("- Gaps detectados:")
            lines.extend([f"  - {gap}" for gap in benchmark["gaps"]])
        else:
            lines.append("- Gaps detectados: ninguno")
        lines.append("")
    md_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    args = parse_args()
    registry = load_registry(args.registry_path)
    inventory = load_inventory(args.inventory_path)
    benchmarks = registry.get("benchmarks_operativos_vigentes", [])

    if args.validate:
        report = build_validation_report(
            registry=registry,
            registry_path=args.registry_path,
            inventory=inventory,
            inventory_path=args.inventory_path,
            grid_workbook_path=args.grid_workbook_path,
            master_workbook_path=args.master_workbook_path,
        )
        save_validation_report(
            report=report,
            json_path=args.report_json_path,
            md_path=args.report_md_path,
        )
        log_event(
            "operacional_controlada",
            "INFO",
            "Auditoria operativa de benchmarks completada",
            estado_global=report["summary"]["estado_global"],
            benchmarks_ok=report["summary"]["benchmarks_ok"],
            benchmarks_con_gaps=report["summary"]["benchmarks_con_gaps"],
            report_json_path=args.report_json_path,
            report_md_path=args.report_md_path,
        )
        if args.strict and report["summary"]["gaps_total"] > 0:
            raise SystemExit("Auditoria operativa con gaps detectados.")

    if not args.benchmark_id:
        log_event(
            "operacional_controlada",
            "INFO",
            "Benchmarks operativos vigentes disponibles",
            benchmark_ids=",".join(item["benchmark_id"] for item in benchmarks),
            registry_path=args.registry_path,
        )
        return

    benchmark = next(
        (item for item in benchmarks if item.get("benchmark_id") == args.benchmark_id),
        None,
    )
    if benchmark is None:
        raise SystemExit(f"No existe benchmark_id={args.benchmark_id} en {args.registry_path}")

    command_preview = build_command(
        benchmark_id=args.benchmark_id,
        run_id=args.run_id or "<run_id_requerido>",
        python_bin=args.python_bin,
    )
    log_event(
        "operacional_controlada",
        "INFO",
        "Benchmark operativo seleccionado",
        benchmark_id=args.benchmark_id,
        plano_funcional=benchmark.get("plano_funcional"),
        run_id_referencia=benchmark.get("run_id"),
        runner_script=benchmark.get("runner_script_canonico"),
        parametros_fuente=benchmark.get("parametros_fuente"),
    )
    print("Comando congelado:")
    print(" ".join(command_preview))

    if not args.execute:
        return

    if not args.run_id:
        raise SystemExit("Usa --run-id cuando ejecutes un benchmark operativo congelado.")

    report = build_validation_report(
        registry=registry,
        registry_path=args.registry_path,
        inventory=inventory,
        inventory_path=args.inventory_path,
        grid_workbook_path=args.grid_workbook_path,
        master_workbook_path=args.master_workbook_path,
    )
    selected_report = next(
        item for item in report["benchmarks"] if item["benchmark_id"] == args.benchmark_id
    )
    if selected_report["gaps"]:
        raise SystemExit(
            f"No se ejecuta {args.benchmark_id}: integridad operativa incompleta: {selected_report['gaps']}"
        )

    command = build_command(args.benchmark_id, args.run_id, args.python_bin)
    log_event(
        "operacional_controlada",
        "INFO",
        "Ejecutando benchmark operativo congelado",
        benchmark_id=args.benchmark_id,
        run_id=args.run_id,
    )
    subprocess.run(command, check=True)


if __name__ == "__main__":
    main()
