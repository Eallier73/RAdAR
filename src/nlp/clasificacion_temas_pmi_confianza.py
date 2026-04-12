#!/usr/bin/env python3
"""
Clasificación de textos por categoría temática usando diccionarios PMI.

Scoring por palabra: Delta_PMI * hits * Confianza

Genera:
  - Un archivo .xlsx por corpus (Twitter, Medios, Youtube)
    con scores brutos (Positivo, Negativo, Neto) por semana
  - Un archivo consolidado (_Consolidado.xlsx) que suma los scores
    de todos los corpus, normalizados por mil tokens para que
    ningún corpus domine por volumen de texto

Uso:
    python clasificar_temas_pmi.py --dict_v5 <ruta> --dict_v10 <ruta> --output <carpeta> --nombre <prefijo>
"""

import re
import os
import sys
import argparse
import unicodedata
from collections import defaultdict, Counter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# ARGUMENTOS
# ============================================================
parser = argparse.ArgumentParser(description='Clasificación temática PMI')
parser.add_argument('--dict_v5', type=str,
    default='/home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/Produccion_Diccionarios/diccionario_pmi_v5.xlsx',
    help='Ruta al diccionario PMI ventana 5')
parser.add_argument('--dict_v10', type=str,
    default='/home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/Produccion_Diccionarios/diccionario_pmi_v10.xlsx',
    help='Ruta al diccionario PMI ventana 10')
parser.add_argument('--output', type=str,
    default='/home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/Resultados_Clasificacio_Temas/resultados_pmi_1',
    help='Carpeta de salida')
parser.add_argument('--nombre', type=str, default='',
    help='Prefijo para los archivos de salida (ej: pmi_run1)')
parser.add_argument('--datos', type=str,
    default='/home/emilio/Documentos/RAdAR/data/text/radar_weekly_flat',
    help='Carpeta base de los corpus')

args = parser.parse_args()

DICT_PATHS = {
    'V5': args.dict_v5,
    'V10': args.dict_v10,
}

BASE_DATOS = args.datos
OUTPUT_DIR = args.output
PREFIJO = args.nombre + '_' if args.nombre else ''

# Sin Facebook — excluido por sesgo institucional
CORPUS_DIRS = {
    'Twitter': os.path.join(BASE_DATOS, 'Twitter_Semana_Texto'),
    'Medios': os.path.join(BASE_DATOS, 'Medios_Semana_Texto'),
    'Youtube': os.path.join(BASE_DATOS, 'Youtube_Semana_Texto'),
}

# Factor de normalización: scores por cada N tokens
NORM_FACTOR = 1000

os.makedirs(OUTPUT_DIR, exist_ok=True)

print(f"Diccionario V5:  {DICT_PATHS['V5']}")
print(f"Diccionario V10: {DICT_PATHS['V10']}")
print(f"Datos:           {BASE_DATOS}")
print(f"Salida:          {OUTPUT_DIR}")
print(f"Prefijo:         {PREFIJO if PREFIJO else '(ninguno)'}")
print(f"Normalizacion:   por {NORM_FACTOR} tokens")
print(f"Corpus:          {', '.join(CORPUS_DIRS.keys())} (Facebook excluido)")

# ============================================================
# FUNCIONES
# ============================================================

def normalizar(texto):
    texto = texto.lower().strip()
    nfkd = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def cargar_diccionario(path):
    """Carga diccionario con 4 columnas: Categoria, Palabra, Delta_PMI, Confianza."""
    wb = load_workbook(path)
    ws = wb.active
    diccionario = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        cat = ws.cell(row, 1).value
        palabra = ws.cell(row, 2).value
        delta = ws.cell(row, 3).value
        confianza = ws.cell(row, 4).value
        if cat and palabra and isinstance(delta, (int, float)):
            w = normalizar(str(palabra))
            conf = float(confianza) if isinstance(confianza, (int, float)) else 1.0
            diccionario[w].append((cat, delta, conf))
    return diccionario


def construir_regex(diccionario):
    patrones = {}
    for palabra in diccionario:
        patrones[palabra] = re.compile(r'\b' + re.escape(palabra) + r'\b')
    return patrones


def contar_tokens(texto):
    """Cuenta tokens (palabras de 3+ caracteres) en texto normalizado."""
    return len(re.findall(r'\b[a-záéíóúñü]{3,}\b', normalizar(texto)))


def clasificar_texto(texto, diccionario, patrones):
    """Scoring: delta * hits * confianza."""
    texto_norm = normalizar(texto)
    scores = defaultdict(lambda: {'pos': 0.0, 'neg': 0.0})
    hits = defaultdict(lambda: Counter())

    for palabra, patron in patrones.items():
        matches = patron.findall(texto_norm)
        n = len(matches)
        if n == 0:
            continue
        for cat, delta, conf in diccionario[palabra]:
            score = delta * n * conf
            if score > 0:
                scores[cat]['pos'] += score
            elif score < 0:
                scores[cat]['neg'] += score
            hits[cat][palabra] += n

    return scores, hits


def extraer_semana(nombre_archivo):
    """Extrae identificador de semana: '26_10_twitter.txt' -> '26_10'."""
    base = os.path.splitext(nombre_archivo)[0]
    for sufijo in ['_facebook', '_twitter', '_medios', '_youtube']:
        if base.lower().endswith(sufijo):
            base = base[:-len(sufijo)]
            break
    return base


def listar_archivos_txt(carpeta):
    archivos = [f for f in os.listdir(carpeta) if f.endswith('.txt')]
    archivos.sort()
    return archivos


# ============================================================
# ESTILOS
# ============================================================
hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
hdr_fill = PatternFill('solid', fgColor='2F5496')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cf = Font(name='Arial', size=10)
cf_pos = Font(name='Arial', size=10, color='006100')
cf_neg = Font(name='Arial', size=10, color='9C0006')
cf_blue = Font(name='Arial', size=10, color='2F5496')
num_fmt = '0.0000'

# ============================================================
# CARGAR DICCIONARIOS
# ============================================================
print("\nCargando diccionarios...")
diccionarios = {}
patrones_dict = {}
categorias_por_dict = {}

for vname, path in DICT_PATHS.items():
    if not os.path.exists(path):
        print(f"  ERROR: No se encontro {path}")
        sys.exit(1)
    d = cargar_diccionario(path)
    diccionarios[vname] = d
    patrones_dict[vname] = construir_regex(d)
    cats = set()
    for entries in d.values():
        for cat, _, _ in entries:
            cats.add(cat)
    categorias_por_dict[vname] = sorted(cats)
    print(f"  {vname}: {len(d)} palabras, {len(cats)} categorias")

# ============================================================
# ALMACÉN PARA CONSOLIDADO
# scores_norm[vname][semana][cat] = {'pos': float, 'neg': float, 'n_corpus': int}
# ============================================================
scores_norm = {vname: defaultdict(lambda: defaultdict(lambda: {'pos': 0.0, 'neg': 0.0, 'n_corpus': 0}))
               for vname in DICT_PATHS}

# ============================================================
# PROCESAR CADA CORPUS
# ============================================================

for corpus_name, corpus_dir in CORPUS_DIRS.items():
    print(f"\n{'='*60}")
    print(f"CORPUS: {corpus_name}")
    print(f"{'='*60}")

    if not os.path.exists(corpus_dir):
        print(f"  AVISO: No existe {corpus_dir}, saltando.")
        continue

    archivos = listar_archivos_txt(corpus_dir)
    print(f"  Archivos: {len(archivos)}")

    if not archivos:
        print(f"  AVISO: No se encontraron archivos .txt")
        continue

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for vname in ['V5', 'V10']:
        diccionario = diccionarios[vname]
        patrones = patrones_dict[vname]
        categorias = categorias_por_dict[vname]

        print(f"\n  --- Diccionario {vname} ---")

        hits_global = defaultdict(lambda: Counter())

        # ========== HOJA DE RESULTADOS ==========
        ws = wb_out.create_sheet(f"Resultados_{vname}")

        ws.cell(1, 1, "Semana").font = hdr_font
        ws.cell(1, 1).fill = hdr_fill
        ws.cell(1, 1).alignment = hdr_align

        col = 2
        cat_cols = {}
        for cat in categorias:
            c_pos = ws.cell(1, col, f"{cat}\nPositivo")
            c_pos.font = hdr_font
            c_pos.fill = PatternFill('solid', fgColor='548235')
            c_pos.alignment = hdr_align

            c_neg = ws.cell(1, col + 1, f"{cat}\nNegativo")
            c_neg.font = hdr_font
            c_neg.fill = PatternFill('solid', fgColor='C00000')
            c_neg.alignment = hdr_align

            c_net = ws.cell(1, col + 2, f"{cat}\nNeto")
            c_net.font = hdr_font
            c_net.fill = PatternFill('solid', fgColor='2F5496')
            c_net.alignment = hdr_align

            cat_cols[cat] = (col, col + 1, col + 2)
            col += 3

        for file_idx, archivo in enumerate(archivos):
            filepath = os.path.join(corpus_dir, archivo)
            with open(filepath, encoding='utf-8', errors='ignore') as f:
                texto = f.read()

            scores, hits = clasificar_texto(texto, diccionario, patrones)
            n_tokens = contar_tokens(texto)

            for cat in hits:
                hits_global[cat].update(hits[cat])

            # Acumular scores normalizados para consolidado
            semana = extraer_semana(archivo)
            if n_tokens > 0:
                for cat in categorias:
                    norm_pos = (scores[cat]['pos'] / n_tokens) * NORM_FACTOR
                    norm_neg = (scores[cat]['neg'] / n_tokens) * NORM_FACTOR
                    scores_norm[vname][semana][cat]['pos'] += norm_pos
                    scores_norm[vname][semana][cat]['neg'] += norm_neg
                    scores_norm[vname][semana][cat]['n_corpus'] += 1

            # Escribir fila bruta
            row = file_idx + 2
            ws.cell(row, 1, os.path.splitext(archivo)[0]).font = cf

            for cat in categorias:
                col_pos, col_neg, col_net = cat_cols[cat]
                s_pos = round(scores[cat]['pos'], 4)
                s_neg = round(scores[cat]['neg'], 4)
                s_net = round(s_pos + s_neg, 4)

                cp = ws.cell(row, col_pos, s_pos)
                cp.number_format = num_fmt
                cp.font = cf_pos if s_pos > 0 else cf

                cn = ws.cell(row, col_neg, s_neg)
                cn.number_format = num_fmt
                cn.font = cf_neg if s_neg < 0 else cf

                cne = ws.cell(row, col_net, s_net)
                cne.number_format = num_fmt
                if s_net > 0:
                    cne.font = cf_pos
                elif s_net < 0:
                    cne.font = cf_neg
                else:
                    cne.font = cf

        # Fila TOTAL
        total_row = len(archivos) + 2
        ws.cell(total_row, 1, "TOTAL").font = Font(bold=True, name='Arial', size=10)
        for cat in categorias:
            col_pos, col_neg, col_net = cat_cols[cat]
            first = 2
            last = total_row - 1
            col_letter_p = get_column_letter(col_pos)
            col_letter_n = get_column_letter(col_neg)
            col_letter_ne = get_column_letter(col_net)

            ws.cell(total_row, col_pos).value = f"=SUM({col_letter_p}{first}:{col_letter_p}{last})"
            ws.cell(total_row, col_pos).font = Font(bold=True, name='Arial', size=10, color='006100')
            ws.cell(total_row, col_pos).number_format = num_fmt

            ws.cell(total_row, col_neg).value = f"=SUM({col_letter_n}{first}:{col_letter_n}{last})"
            ws.cell(total_row, col_neg).font = Font(bold=True, name='Arial', size=10, color='9C0006')
            ws.cell(total_row, col_neg).number_format = num_fmt

            ws.cell(total_row, col_net).value = f"=SUM({col_letter_ne}{first}:{col_letter_ne}{last})"
            ws.cell(total_row, col_net).font = Font(bold=True, name='Arial', size=10, color='2F5496')
            ws.cell(total_row, col_net).number_format = num_fmt

        ws.column_dimensions['A'].width = 28
        for c in range(2, col):
            ws.column_dimensions[get_column_letter(c)].width = 14
        ws.freeze_panes = 'B2'
        ws.row_dimensions[1].height = 35

        print(f"    Resultados: {len(archivos)} semanas x {len(categorias)} categorias")

        # ========== HOJA TOP 30 ==========
        ws_top = wb_out.create_sheet(f"Top30_{vname}")

        col_offset = 0
        for cat in categorias:
            base_col = col_offset * 4 + 1

            c_cat = ws_top.cell(1, base_col, cat)
            c_cat.font = Font(bold=True, name='Arial', size=11, color='2F5496')
            ws_top.merge_cells(start_row=1, start_column=base_col, end_row=1, end_column=base_col + 2)

            for i, h in enumerate(['Palabra', 'Hits', 'Delta_PMI']):
                c = ws_top.cell(2, base_col + i, h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align

            top30 = hits_global[cat].most_common(30)
            for rank, (palabra, count) in enumerate(top30, 3):
                ws_top.cell(rank, base_col, palabra).font = cf
                ws_top.cell(rank, base_col + 1, count).font = cf

                delta = None
                for entry_cat, entry_delta, _ in diccionario.get(palabra, []):
                    if entry_cat == cat:
                        delta = entry_delta
                        break

                if delta is not None:
                    cd = ws_top.cell(rank, base_col + 2, round(delta, 4))
                    cd.number_format = num_fmt
                    if delta > 0:
                        cd.font = cf_pos
                    elif delta < 0:
                        cd.font = cf_neg
                    else:
                        cd.font = cf

            ws_top.column_dimensions[get_column_letter(base_col)].width = 20
            ws_top.column_dimensions[get_column_letter(base_col + 1)].width = 8
            ws_top.column_dimensions[get_column_letter(base_col + 2)].width = 12

            col_offset += 1

        ws_top.freeze_panes = 'A3'
        print(f"    Top 30: {len(categorias)} categorias")

    # Guardar archivo por corpus
    out_path = os.path.join(OUTPUT_DIR, f"{PREFIJO}{corpus_name}.xlsx")
    wb_out.save(out_path)
    print(f"\n  Guardado: {out_path}")

# ============================================================
# ARCHIVO CONSOLIDADO
# ============================================================
print(f"\n{'='*60}")
print(f"GENERANDO CONSOLIDADO")
print(f"{'='*60}")

wb_con = Workbook()
wb_con.remove(wb_con.active)

for vname in ['V5', 'V10']:
    categorias = categorias_por_dict[vname]
    semanas_dict = scores_norm[vname]
    semanas_sorted = sorted(semanas_dict.keys())

    if not semanas_sorted:
        print(f"  {vname}: Sin datos para consolidar")
        continue

    ws = wb_con.create_sheet(f"Consolidado_{vname}")

    # Header
    ws.cell(1, 1, "Semana").font = hdr_font
    ws.cell(1, 1).fill = hdr_fill
    ws.cell(1, 1).alignment = hdr_align

    col = 2
    cat_cols = {}
    for cat in categorias:
        c_pos = ws.cell(1, col, f"{cat}\nPositivo")
        c_pos.font = hdr_font
        c_pos.fill = PatternFill('solid', fgColor='548235')
        c_pos.alignment = hdr_align

        c_neg = ws.cell(1, col + 1, f"{cat}\nNegativo")
        c_neg.font = hdr_font
        c_neg.fill = PatternFill('solid', fgColor='C00000')
        c_neg.alignment = hdr_align

        c_net = ws.cell(1, col + 2, f"{cat}\nNeto")
        c_net.font = hdr_font
        c_net.fill = PatternFill('solid', fgColor='2F5496')
        c_net.alignment = hdr_align

        cat_cols[cat] = (col, col + 1, col + 2)
        col += 3

    # Datos
    for row_idx, semana in enumerate(semanas_sorted, 2):
        ws.cell(row_idx, 1, semana).font = cf

        for cat in categorias:
            col_pos, col_neg, col_net = cat_cols[cat]
            s = semanas_dict[semana][cat]
            s_pos = round(s['pos'], 4)
            s_neg = round(s['neg'], 4)
            s_net = round(s_pos + s_neg, 4)

            cp = ws.cell(row_idx, col_pos, s_pos)
            cp.number_format = num_fmt
            cp.font = cf_pos if s_pos > 0 else cf

            cn = ws.cell(row_idx, col_neg, s_neg)
            cn.number_format = num_fmt
            cn.font = cf_neg if s_neg < 0 else cf

            cne = ws.cell(row_idx, col_net, s_net)
            cne.number_format = num_fmt
            if s_net > 0:
                cne.font = cf_pos
            elif s_net < 0:
                cne.font = cf_neg
            else:
                cne.font = cf

    # Fila TOTAL
    total_row = len(semanas_sorted) + 2
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True, name='Arial', size=10)
    for cat in categorias:
        col_pos, col_neg, col_net = cat_cols[cat]
        first = 2
        last = total_row - 1
        col_letter_p = get_column_letter(col_pos)
        col_letter_n = get_column_letter(col_neg)
        col_letter_ne = get_column_letter(col_net)

        ws.cell(total_row, col_pos).value = f"=SUM({col_letter_p}{first}:{col_letter_p}{last})"
        ws.cell(total_row, col_pos).font = Font(bold=True, name='Arial', size=10, color='006100')
        ws.cell(total_row, col_pos).number_format = num_fmt

        ws.cell(total_row, col_neg).value = f"=SUM({col_letter_n}{first}:{col_letter_n}{last})"
        ws.cell(total_row, col_neg).font = Font(bold=True, name='Arial', size=10, color='9C0006')
        ws.cell(total_row, col_neg).number_format = num_fmt

        ws.cell(total_row, col_net).value = f"=SUM({col_letter_ne}{first}:{col_letter_ne}{last})"
        ws.cell(total_row, col_net).font = Font(bold=True, name='Arial', size=10, color='2F5496')
        ws.cell(total_row, col_net).number_format = num_fmt

    ws.column_dimensions['A'].width = 20
    for c in range(2, col):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = 'B2'
    ws.row_dimensions[1].height = 35

    print(f"  {vname}: {len(semanas_sorted)} semanas consolidadas")

con_path = os.path.join(OUTPUT_DIR, f"{PREFIJO}Consolidado.xlsx")
wb_con.save(con_path)
print(f"\n  Guardado: {con_path}")

print(f"\n{'='*60}")
print(f"LISTO. Resultados en: {OUTPUT_DIR}")
print(f"{'='*60}")
