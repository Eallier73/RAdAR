#!/usr/bin/env python3
"""
Fusiona el dataset semanal listo para ML con las variables semanales de
sentimiento digital calculadas por semana ISO.

Salida:
    /home/emilio/Documentos/RAdAR/data/processed/modeling/datos_ML_0.xlsx
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path("/home/emilio/Documentos/RAdAR")
ML_INPUT_PATH = BASE_DIR / "data" / "processed" / "modeling" / "ML_Ready_Monica_Villarreal_Encuestas_PMI (1).xlsx"
SENT_INPUT_PATH = (
    BASE_DIR / "data" / "processed" / "modeling" / "aceptacion_digital_redes_medios_sentimiento_semanal.xlsx"
)
OUTPUT_PATH = BASE_DIR / "data" / "processed" / "modeling" / "datos_ML_0.xlsx"

TRAIN_SHEET = "ML_Ready_Train"
ALL_WEEKS_SHEET = "ML_Ready_AllWeeks"
README_SHEET = "README"

SENT_FEATURES = [
    "sentimiento_facebook",
    "sentimiento_twitter",
    "sentimiento_youtube",
    "sentimiento_redes_ponderado",
    "sentimiento_medios",
    "promedio_redes_medios",
]
SENT_JOIN_KEYS = ["iso_year", "iso_week"]
INSERT_AFTER = "has_full_pmi_features"


def cargar_sentimiento() -> pd.DataFrame:
    sentimiento = pd.read_excel(SENT_INPUT_PATH, sheet_name="sentimiento_semanal")
    sentimiento = sentimiento.rename(
        columns={
            "anio_iso": "iso_year",
            "semana_iso": "iso_week",
        }
    )
    columnas = ["iso_year", "iso_week"] + SENT_FEATURES
    sentimiento = sentimiento[columnas].copy()
    sentimiento["iso_year"] = sentimiento["iso_year"].astype(int)
    sentimiento["iso_week"] = sentimiento["iso_week"].astype(int)
    return sentimiento


def reordenar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    if INSERT_AFTER not in df.columns:
        return df

    nuevas = ["flag_missing_sentimiento_digital"] + SENT_FEATURES
    existentes = [col for col in df.columns if col not in nuevas]

    indice = existentes.index(INSERT_AFTER) + 1
    columnas = existentes[:indice] + nuevas + existentes[indice:]
    return df[columnas]


def fusionar_hoja(df_ml: pd.DataFrame, sentimiento: pd.DataFrame) -> pd.DataFrame:
    salida = df_ml.copy()
    salida["iso_year"] = salida["iso_year"].astype(int)
    salida["iso_week"] = salida["iso_week"].astype(int)

    salida = salida.merge(
        sentimiento,
        on=SENT_JOIN_KEYS,
        how="left",
        validate="one_to_one",
    )
    salida["flag_missing_sentimiento_digital"] = salida["sentimiento_redes_ponderado"].isna().astype(int)
    salida = reordenar_columnas(salida)
    return salida


def construir_readme(
    train_rows: int,
    all_rows: int,
    faltantes_allweeks: int,
) -> pd.DataFrame:
    filas = [
        (
            "Objetivo",
            "Dataset semanal listo para ML con encuestas, PMI y sentimiento digital fusionados por semana ISO.",
        ),
        (
            "Hoja principal",
            "ML_Ready_Train: semanas con features originales del dataset ML mas sentimiento digital.",
        ),
        (
            "Hoja secundaria",
            "ML_Ready_AllWeeks: todas las semanas del dataset ML; las semanas sin texto quedan con NaN en sentimiento.",
        ),
        ("Archivo base ML", ML_INPUT_PATH.name),
        ("Archivo de sentimiento", SENT_INPUT_PATH.name),
        ("Clave de union", "iso_year + iso_week"),
        ("Nuevas variables", ", ".join(["flag_missing_sentimiento_digital"] + SENT_FEATURES)),
        ("Filas ML_Ready_Train", train_rows),
        ("Filas ML_Ready_AllWeeks", all_rows),
        ("Semanas sin sentimiento en AllWeeks", faltantes_allweeks),
    ]
    return pd.DataFrame(filas, columns=["Campo", "Valor"])


def aplicar_formato(writer: pd.ExcelWriter, nombre_hoja: str, dataframe: pd.DataFrame) -> None:
    hoja = writer.sheets[nombre_hoja]
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in hoja[1]:
        cell.fill = fill
        cell.font = font

    for column_cells in hoja.columns:
        letra = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        hoja.column_dimensions[letra].width = min(max(max_len + 2, 12), 32)

    formatos = {
        "mitofsky": "0.0",
        "mitofsky_locf": "0.0",
        "scr": "0.0",
        "scr_locf": "0.0",
        "demoscopia": "0.0",
        "demoscopia_locf": "0.0",
        "rubrum_eval_general": "0.00",
        "rubrum_eval_general_locf": "0.00",
        "rubrum_x10": "0.0",
        "rubrum_x10_locf": "0.0",
        "consenso_mensual_3e": "0.000",
        "consenso_mensual_3e_locf": "0.000",
        "consenso_mensual_4e": "0.000",
        "consenso_mensual_4e_locf": "0.000",
        "target_serie_3e": "0.000",
        "target_serie_4e": "0.000",
        "sentimiento_facebook": "0.0000",
        "sentimiento_twitter": "0.0000",
        "sentimiento_youtube": "0.0000",
        "sentimiento_redes_ponderado": "0.0000",
        "sentimiento_medios": "0.0000",
        "promedio_redes_medios": "0.0000",
    }

    columnas = {col: idx + 1 for idx, col in enumerate(dataframe.columns)}
    for nombre_columna, formato in formatos.items():
        indice = columnas.get(nombre_columna)
        if not indice:
            continue
        letra = get_column_letter(indice)
        for row in range(2, hoja.max_row + 1):
            hoja[f"{letra}{row}"].number_format = formato


def guardar_excel(train_df: pd.DataFrame, allweeks_df: pd.DataFrame, readme_df: pd.DataFrame) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        train_df.to_excel(writer, sheet_name=TRAIN_SHEET, index=False)
        allweeks_df.to_excel(writer, sheet_name=ALL_WEEKS_SHEET, index=False)
        readme_df.to_excel(writer, sheet_name=README_SHEET, index=False)

        aplicar_formato(writer, TRAIN_SHEET, train_df)
        aplicar_formato(writer, ALL_WEEKS_SHEET, allweeks_df)
        aplicar_formato(writer, README_SHEET, readme_df)


def main() -> None:
    print("=" * 72)
    print("FUSION ML + SENTIMIENTO DIGITAL")
    print("=" * 72)

    sentimiento = cargar_sentimiento()
    ml_train = pd.read_excel(ML_INPUT_PATH, sheet_name=TRAIN_SHEET)
    ml_allweeks = pd.read_excel(ML_INPUT_PATH, sheet_name=ALL_WEEKS_SHEET)

    train_merge = fusionar_hoja(ml_train, sentimiento)
    allweeks_merge = fusionar_hoja(ml_allweeks, sentimiento)

    faltantes_allweeks = int(allweeks_merge["flag_missing_sentimiento_digital"].sum())
    readme = construir_readme(
        train_rows=len(train_merge),
        all_rows=len(allweeks_merge),
        faltantes_allweeks=faltantes_allweeks,
    )

    guardar_excel(train_merge, allweeks_merge, readme)

    print(f"Filas {TRAIN_SHEET}: {len(train_merge):,}")
    print(f"Filas {ALL_WEEKS_SHEET}: {len(allweeks_merge):,}")
    print(f"Semanas sin sentimiento en {ALL_WEEKS_SHEET}: {faltantes_allweeks}")
    print(f"Archivo generado: {OUTPUT_PATH}")
    print("=" * 72)


if __name__ == "__main__":
    main()
