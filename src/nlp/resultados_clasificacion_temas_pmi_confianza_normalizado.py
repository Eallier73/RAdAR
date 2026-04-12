#!/usr/bin/env python3
"""
Normalización de resultados de clasificación temática PMI.

Lee un archivo consolidado (.xlsx) con columnas Positivo/Negativo/Neto
por categoría y aplica log-transform para comprimir outliers:

    valor_norm = signo(x) * log_base(1 + |x|)

Esto convierte, por ejemplo (con base=2):
    -453  ->  -8.83
    -30   ->  -4.95
    -3    ->  -2.00
    10    ->   3.46
    300   ->   8.24

Preserva la dirección (signo), el orden relativo, y comprime
los picos extremos sin perder información.

Uso:
    python normalizar_consolidado.py --input <entrada.xlsx> --output <salida.xlsx>

Ejemplo:
    python normalizar_consolidado.py \
        --input /home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/Resultados_Clasificacio_Temas/consolidado.xlsx \
        --output /home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/Resultados_Clasificacio_Temas/consolidado_norm.xlsx

Parámetros opcionales:
    --base    Base del logaritmo (default: 2). Mayor base = menos compresión.
    --nombre  Sufijo para el archivo de salida (se ignora si se usa --output)
"""

import sys
import os
import math
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# PARÁMETROS MODIFICABLES
# ============================================================
# Base del logaritmo: 2 = compresión fuerte, 10 = compresión suave
LOG_BASE_DEFAULT = 2

# Rutas por defecto
INPUT_DEFAULT  = "/home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/Resultados_Clasificacio_Temas/consolidado.xlsx"
OUTPUT_DEFAULT = "/home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/Resultados_Clasificacio_Temas/consolidado_norm.xlsx"

# ============================================================
# ARGUMENTOS
# ============================================================
parser = argparse.ArgumentParser(description='Normalización log-transform de consolidado PMI')
parser.add_argument('--input', type=str, default=INPUT_DEFAULT,
    help='Ruta al archivo consolidado de entrada (.xlsx)')
parser.add_argument('--output', type=str, default=OUTPUT_DEFAULT,
    help='Ruta al archivo normalizado de salida (.xlsx)')
parser.add_argument('--base', type=float, default=LOG_BASE_DEFAULT,
    help=f'Base del logaritmo (default: {LOG_BASE_DEFAULT}). Mayor = menos compresión.')
parser.add_argument('--nombre', type=str, default='',
    help='Sufijo para nombre de salida (se genera automáticamente si no se pasa --output)')

args = parser.parse_args()

INPUT_PATH = args.input
LOG_BASE = args.base

# Si se pasó --nombre pero no --output explícito, construir nombre de salida
if args.nombre and args.output == OUTPUT_DEFAULT:
    directorio = os.path.dirname(INPUT_PATH)
    OUTPUT_PATH = os.path.join(directorio, f"{args.nombre}.xlsx")
else:
    OUTPUT_PATH = args.output

if not os.path.exists(INPUT_PATH):
    print(f"ERROR: No se encontró {INPUT_PATH}")
    sys.exit(1)

# ============================================================
# TRANSFORMACIÓN
# ============================================================
def log_transform(x, base):
    """signo(x) * log_base(1 + |x|)"""
    if x == 0:
        return 0.0
    return math.copysign(math.log(1 + abs(x)) / math.log(base), x)

# ============================================================
# ESTILOS
# ============================================================
hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
hdr_fill = PatternFill('solid', fgColor='2F5496')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cf = Font(name='Arial', size=10)
cf_pos = Font(name='Arial', size=10, color='006100')
cf_neg = Font(name='Arial', size=10, color='9C0006')
num_fmt = '0.0000'

# ============================================================
# PROCESAR
# ============================================================
print(f"Entrada:    {INPUT_PATH}")
print(f"Salida:     {OUTPUT_PATH}")
print(f"Log base:   {LOG_BASE}")

wb_in = load_workbook(INPUT_PATH)
wb_out = Workbook()
wb_out.remove(wb_out.active)

for sheet_name in wb_in.sheetnames:
    ws_in = wb_in[sheet_name]
    ws_out = wb_out.create_sheet(sheet_name)

    print(f"\n  Hoja: {sheet_name}")
    print(f"    Filas: {ws_in.max_row}, Columnas: {ws_in.max_column}")

    # Copiar headers
    for col in range(1, ws_in.max_column + 1):
        val = ws_in.cell(1, col).value
        cell = ws_out.cell(1, col, val)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    # Detectar fila TOTAL
    last_data_row = ws_in.max_row
    last_val = ws_in.cell(ws_in.max_row, 1).value
    if isinstance(last_val, str) and last_val.upper() == 'TOTAL':
        last_data_row = ws_in.max_row - 1

    # Transformar datos
    transformed = 0
    for row in range(2, last_data_row + 1):
        semana = ws_in.cell(row, 1).value
        ws_out.cell(row, 1, semana).font = cf

        for col in range(2, ws_in.max_column + 1):
            val = ws_in.cell(row, col).value
            if isinstance(val, (int, float)):
                norm = round(log_transform(val, LOG_BASE), 4)
                c = ws_out.cell(row, col, norm)
                c.number_format = num_fmt
                if norm > 0:
                    c.font = cf_pos
                elif norm < 0:
                    c.font = cf_neg
                else:
                    c.font = cf
                transformed += 1
            else:
                ws_out.cell(row, col, val).font = cf

    # Fila TOTAL con fórmulas
    total_row = last_data_row + 1
    ws_out.cell(total_row, 1, "TOTAL").font = Font(bold=True, name='Arial', size=10)
    for col in range(2, ws_in.max_column + 1):
        col_letter = get_column_letter(col)
        ws_out.cell(total_row, col).value = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
        ws_out.cell(total_row, col).number_format = num_fmt
        ws_out.cell(total_row, col).font = Font(bold=True, name='Arial', size=10)

    # Formato columnas
    ws_out.column_dimensions['A'].width = 20
    for col in range(2, ws_in.max_column + 1):
        ws_out.column_dimensions[get_column_letter(col)].width = 14
    ws_out.freeze_panes = 'B2'
    ws_out.row_dimensions[1].height = 35

    print(f"    Celdas transformadas: {transformed:,}")

# ============================================================
# GUARDAR
# ============================================================
os.makedirs(os.path.dirname(OUTPUT_PATH) or '.', exist_ok=True)
wb_out.save(OUTPUT_PATH)
print(f"\nGuardado: {OUTPUT_PATH}")
print("Listo.")
