#!/usr/bin/env python3
"""
Cálculo de PMI y WPMI por palabra de categoría contra sentimiento positivo/negativo.
Ventanas deslizantes de 5 y 10 palabras sobre texto limpio (sin stopwords).

Métricas:
  - PMI clásico (log2 de asociación)
  - WPMI = P(w, sent) * PMI  →  pondera por probabilidad conjunta
  - Delta_PMI  = PMI_pos - PMI_neg       (dirección, sin frecuencia)
  - Delta_WPMI = WPMI_pos - WPMI_neg     (dirección + evidencia)

Exclusión por palabra: cuando una palabra de categoría también está en el
diccionario de sentimiento, NO se cuenta a sí misma como indicador de
sentimiento en su propio chunk. Sí cuenta como indicador para otras palabras.

Uso:
    python diccionario_temas_pmi.py <nombre_archivo> <carpeta_salida>

Ejemplo:
    python diccionario_temas_pmi.py pmi_corpus_limpio /home/emilio/Documentos/RAdAR/Resultados

    Genera: /home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/pmi_corpus_limpio.xlsx

Si no se pasan argumentos:
    Genera: /home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp/pmi_resultados.xlsx
"""

import re
import sys
import os
import math
import unicodedata
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# ARGUMENTOS
# ============================================================
BASE = "/home/emilio/Documentos/RAdAR/data/reference/dictionaries_nlp"

if len(sys.argv) >= 3:
    nombre_archivo = sys.argv[1]
    carpeta_salida = sys.argv[2]
elif len(sys.argv) == 2:
    nombre_archivo = sys.argv[1]
    carpeta_salida = BASE
else:
    nombre_archivo = "pmi_resultados"
    carpeta_salida = BASE

if not nombre_archivo.endswith('.xlsx'):
    nombre_archivo += '.xlsx'

os.makedirs(carpeta_salida, exist_ok=True)
OUTPUT_PATH = os.path.join(carpeta_salida, nombre_archivo)

# ============================================================
# RUTAS DE ENTRADA
# ============================================================
CORPUS_PATH     = f"{BASE}/Corpus_Diccionario/corpus_radar_texto_completo.txt"
STOPLIST_PATH   = f"{BASE}/stop_list_espanol_limpia.txt"
DICT_POS_PATH   = f"{BASE}/diccionario_palabras_positivas.txt"
DICT_NEG_PATH   = f"{BASE}/diccionario_palabras_negativas.txt"
CATEGORIAS_PATH = f"{BASE}/categorías_subcategorias_palabras.xlsx"

VENTANAS = [5, 10]

# ============================================================
# FUNCIONES
# ============================================================

def normalizar(texto):
    texto = texto.lower().strip()
    nfkd = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def cargar_lista(path):
    palabras = set()
    with open(path, encoding='utf-8') as f:
        for linea in f:
            w = normalizar(linea.strip().replace('\t', '').replace('\r', ''))
            if w:
                palabras.add(w)
    return palabras


def cargar_categorias(path):
    wb = load_workbook(path)
    ws = wb.active
    cats = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if not header:
            continue
        words = set()
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row, col).value
            if v:
                words.add(normalizar(str(v)))
        cats[header] = words
    return cats


def tokenizar_y_limpiar(texto, stopwords):
    tokens = re.findall(r'\b[a-záéíóúñü]+\b', normalizar(texto))
    return [t for t in tokens if t not in stopwords and len(t) > 2]


# ============================================================
# CARGA DE DATOS
# ============================================================
print("Cargando datos...")

stopwords = cargar_lista(STOPLIST_PATH)
dict_pos = cargar_lista(DICT_POS_PATH)
dict_neg = cargar_lista(DICT_NEG_PATH)
categorias = cargar_categorias(CATEGORIAS_PATH)

todas_cat_words = set()
for words in categorias.values():
    todas_cat_words.update(words)

word_to_cats = defaultdict(set)
for cat_name, words in categorias.items():
    for w in words:
        word_to_cats[w].add(cat_name)

print(f"  Stopwords: {len(stopwords)}")
print(f"  Dict positivo: {len(dict_pos)} palabras")
print(f"  Dict negativo: {len(dict_neg)} palabras")
print(f"  Categorias: {len(categorias)}")
print(f"  Palabras de categoria (unicas): {len(todas_cat_words)}")

traslape_pos = todas_cat_words & dict_pos
traslape_neg = todas_cat_words & dict_neg
if traslape_pos:
    print(f"  TRASLAPE con dict positivo: {len(traslape_pos)} palabras")
    print(f"    Ejemplos: {sorted(traslape_pos)[:15]}")
if traslape_neg:
    print(f"  TRASLAPE con dict negativo: {len(traslape_neg)} palabras")
    print(f"    Ejemplos: {sorted(traslape_neg)[:15]}")
print(f"  -> Se aplica exclusion por palabra (no cuenta consigo misma)")

# ============================================================
# LEER Y LIMPIAR CORPUS
# ============================================================
print("Leyendo y limpiando corpus...")

with open(CORPUS_PATH, encoding='utf-8') as f:
    texto_crudo = f.read()

tokens = tokenizar_y_limpiar(texto_crudo, stopwords)
N_tokens = len(tokens)
print(f"  Tokens limpios: {N_tokens:,}")

# ============================================================
# PROCESAR POR VENTANA
# ============================================================

wb_out = Workbook()
wb_out.remove(wb_out.active)

for VENTANA in VENTANAS:
    print(f"\n{'='*60}")
    print(f"VENTANA DE {VENTANA} PALABRAS")
    print(f"{'='*60}")

    total_chunks = N_tokens - VENTANA + 1
    if total_chunks <= 0:
        print(f"  ERROR: corpus demasiado corto para ventana de {VENTANA}")
        continue

    # Contadores globales (para P(sentimiento), sin exclusiones)
    count_pos_global = 0
    count_neg_global = 0

    # Contadores por palabra de categoría (con exclusión por palabra)
    count_word = defaultdict(int)
    count_word_pos = defaultdict(int)
    count_word_neg = defaultdict(int)

    print(f"  Total chunks: {total_chunks:,}")
    print(f"  Escaneando...")

    for i in range(total_chunks):
        chunk = tokens[i:i + VENTANA]
        chunk_set = set(chunk)

        # Conteo GLOBAL de sentimiento (sin exclusiones, para P(sent))
        has_pos_global = bool(chunk_set & dict_pos)
        has_neg_global = bool(chunk_set & dict_neg)
        if has_pos_global:
            count_pos_global += 1
        if has_neg_global:
            count_neg_global += 1

        # Palabras de categoría en este chunk
        cat_words_in_chunk = chunk_set & todas_cat_words
        if not cat_words_in_chunk:
            continue

        # Para cada palabra de categoría: exclusión por palabra
        for w in cat_words_in_chunk:
            count_word[w] += 1

            # Quitar SOLO la propia palabra w para evaluar sentimiento
            resto = chunk_set - {w}
            if resto & dict_pos:
                count_word_pos[w] += 1
            if resto & dict_neg:
                count_word_neg[w] += 1

        if i % 500000 == 0 and i > 0:
            print(f"    ...{i:,}/{total_chunks:,}")

    print(f"  Chunks con positivo (global): {count_pos_global:,}")
    print(f"  Chunks con negativo (global): {count_neg_global:,}")
    print(f"  Palabras de categoria encontradas: {len(count_word)}")

    # ============================================================
    # CALCULAR PMI y WPMI
    # ============================================================
    print(f"  Calculando PMI y WPMI...")

    P_pos = count_pos_global / total_chunks
    P_neg = count_neg_global / total_chunks

    resultados = []

    for cat_name, cat_words in categorias.items():
        for w in sorted(cat_words):
            cw = count_word.get(w, 0)
            if cw == 0:
                continue

            P_w = cw / total_chunks
            cwp = count_word_pos.get(w, 0)
            cwn = count_word_neg.get(w, 0)

            P_w_pos = cwp / total_chunks   # P(w, pos)
            P_w_neg = cwn / total_chunks   # P(w, neg)

            # --- PMI clásico ---
            if P_w_pos > 0 and P_pos > 0:
                pmi_pos = math.log2(P_w_pos / (P_w * P_pos))
            else:
                pmi_pos = None

            if P_w_neg > 0 and P_neg > 0:
                pmi_neg = math.log2(P_w_neg / (P_w * P_neg))
            else:
                pmi_neg = None

            if pmi_pos is not None and pmi_neg is not None:
                delta = pmi_pos - pmi_neg
            elif pmi_pos is not None:
                delta = pmi_pos
            elif pmi_neg is not None:
                delta = -pmi_neg
            else:
                delta = None

            # --- WPMI = P(w, sent) * PMI ---
            wpmi_pos = (P_w_pos * pmi_pos) if pmi_pos is not None else None
            wpmi_neg = (P_w_neg * pmi_neg) if pmi_neg is not None else None

            if wpmi_pos is not None and wpmi_neg is not None:
                delta_wpmi = wpmi_pos - wpmi_neg
            elif wpmi_pos is not None:
                delta_wpmi = wpmi_pos
            elif wpmi_neg is not None:
                delta_wpmi = -wpmi_neg
            else:
                delta_wpmi = None

            resultados.append((
                cat_name, w, cw, cwp, cwn,
                pmi_pos, pmi_neg, delta,
                wpmi_pos, wpmi_neg, delta_wpmi
            ))

    # ============================================================
    # HOJA DE DATOS
    # ============================================================
    ws = wb_out.create_sheet(f"Ventana_{VENTANA}")

    headers = [
        "Categoria", "Palabra", "Chunks_palabra",
        "Chunks_palabra_pos", "Chunks_palabra_neg",
        "PMI_pos", "PMI_neg", "Delta_PMI",
        "WPMI_pos", "WPMI_neg", "Delta_WPMI"
    ]
    hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    hdr_fill = PatternFill('solid', fgColor='2F5496')
    hdr_align = Alignment(horizontal='center', vertical='center')
    num_fmt_4 = '0.0000'
    num_fmt_8 = '0.00000000'
    cf = Font(name='Arial', size=10)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    for r, row_data in enumerate(resultados, 2):
        (cat, palabra, cw, cwp, cwn,
         pmi_p, pmi_n, delta,
         wpmi_p, wpmi_n, delta_w) = row_data

        ws.cell(r, 1, cat).font = cf
        ws.cell(r, 2, palabra).font = cf
        ws.cell(r, 3, cw).font = cf
        ws.cell(r, 4, cwp).font = cf
        ws.cell(r, 5, cwn).font = cf

        # PMI clásico
        if pmi_p is not None:
            c6 = ws.cell(r, 6, round(pmi_p, 4))
            c6.number_format = num_fmt_4
        else:
            ws.cell(r, 6, "N/A")

        if pmi_n is not None:
            c7 = ws.cell(r, 7, round(pmi_n, 4))
            c7.number_format = num_fmt_4
        else:
            ws.cell(r, 7, "N/A")

        if delta is not None:
            c8 = ws.cell(r, 8, round(delta, 4))
            c8.number_format = num_fmt_4
            if delta > 0:
                c8.font = Font(name='Arial', size=10, color='006100')
            elif delta < 0:
                c8.font = Font(name='Arial', size=10, color='9C0006')
        else:
            ws.cell(r, 8, "N/A")

        # WPMI
        if wpmi_p is not None:
            c9 = ws.cell(r, 9, round(wpmi_p, 8))
            c9.number_format = num_fmt_8
        else:
            ws.cell(r, 9, "N/A")

        if wpmi_n is not None:
            c10 = ws.cell(r, 10, round(wpmi_n, 8))
            c10.number_format = num_fmt_8
        else:
            ws.cell(r, 10, "N/A")

        if delta_w is not None:
            c11 = ws.cell(r, 11, round(delta_w, 8))
            c11.number_format = num_fmt_8
            if delta_w > 0:
                c11.font = Font(name='Arial', size=10, color='006100')
            elif delta_w < 0:
                c11.font = Font(name='Arial', size=10, color='9C0006')
        else:
            ws.cell(r, 11, "N/A")

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    for cl in ['C', 'D', 'E']:
        ws.column_dimensions[cl].width = 18
    for cl in ['F', 'G', 'H']:
        ws.column_dimensions[cl].width = 14
    for cl in ['I', 'J', 'K']:
        ws.column_dimensions[cl].width = 16
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:K{len(resultados) + 1}"

    # ============================================================
    # HOJA RESUMEN
    # ============================================================
    ws_r = wb_out.create_sheet(f"Resumen_V{VENTANA}")

    ws_r.cell(1, 1, "Estadisticas globales").font = Font(bold=True, name='Arial', size=11)
    info = [
        ("Total chunks", f"{total_chunks:,}"),
        ("Chunks con positivo", f"{count_pos_global:,}"),
        ("Chunks con negativo", f"{count_neg_global:,}"),
        ("P(positivo)", round(P_pos, 6)),
        ("P(negativo)", round(P_neg, 6)),
        ("Tokens limpios en corpus", f"{N_tokens:,}"),
        ("Tamano ventana", VENTANA),
        ("Traslape cat-positivo", f"{len(traslape_pos)} palabras"),
        ("Traslape cat-negativo", f"{len(traslape_neg)} palabras"),
        ("Metodo", "Exclusion por palabra (no cuenta consigo misma)"),
        ("Metrica principal", "WPMI = P(w, sent) * PMI"),
        ("Ponderacion resumen", "log2(chunks_palabra + 1)"),
    ]
    for i, (label, val) in enumerate(info, 2):
        ws_r.cell(i, 1, label).font = cf
        ws_r.cell(i, 2, val).font = cf

    # --- Promedios ponderados por log2(chunks_palabra + 1) ---
    # Se usa log2 para comprimir el rango de pesos: una palabra con
    # 5000 chunks pesa ~3.5x más que una con 10 (no 500x).
    cat_stats = defaultdict(lambda: {
        'n': 0,
        'sum_cw': 0,             # suma de chunks_palabra (informativo)
        'sum_log_w': 0.0,        # suma de log2(cw+1) (denominador)
        'sum_delta_w': 0.0,      # suma ponderada de delta_wpmi
        'sum_delta': 0.0,        # suma ponderada de delta_pmi (referencia)
        'pos': 0, 'neg': 0, 'na': 0
    })
    for row_data in resultados:
        (cat, palabra, cw, cwp, cwn,
         pmi_p, pmi_n, delta,
         wpmi_p, wpmi_n, delta_w) = row_data
        cs = cat_stats[cat]
        cs['n'] += 1
        cs['sum_cw'] += cw
        log_w = math.log2(cw + 1)
        cs['sum_log_w'] += log_w
        if delta_w is not None:
            cs['sum_delta_w'] += delta_w * log_w
            if delta is not None:
                cs['sum_delta'] += delta * log_w
            if delta_w > 0:
                cs['pos'] += 1
            else:
                cs['neg'] += 1
        else:
            cs['na'] += 1

    row = len(info) + 3
    cat_hdrs = [
        "Categoria", "Palabras", "Total_chunks_cat",
        "Prom_Delta_WPMI_pond", "Prom_Delta_PMI_pond",
        "Tend. Positiva", "Tend. Negativa", "Sin datos"
    ]
    for c, h in enumerate(cat_hdrs, 1):
        cell = ws_r.cell(row, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    for cat_name in categorias.keys():
        if cat_name not in cat_stats:
            continue
        row += 1
        cs = cat_stats[cat_name]
        ws_r.cell(row, 1, cat_name).font = cf
        ws_r.cell(row, 2, cs['n']).font = cf
        ws_r.cell(row, 3, cs['sum_cw']).font = cf

        # Promedio ponderado de Delta_WPMI (peso = log2(cw+1))
        if cs['sum_log_w'] > 0:
            avg_wpmi = cs['sum_delta_w'] / cs['sum_log_w']
        else:
            avg_wpmi = 0
        c4 = ws_r.cell(row, 4, round(avg_wpmi, 8))
        c4.number_format = num_fmt_8
        if avg_wpmi > 0:
            c4.font = Font(name='Arial', size=10, color='006100')
        elif avg_wpmi < 0:
            c4.font = Font(name='Arial', size=10, color='9C0006')

        # Promedio ponderado de Delta_PMI (peso = log2(cw+1), referencia)
        if cs['sum_log_w'] > 0:
            avg_pmi = cs['sum_delta'] / cs['sum_log_w']
        else:
            avg_pmi = 0
        c5 = ws_r.cell(row, 5, round(avg_pmi, 4))
        c5.number_format = num_fmt_4
        if avg_pmi > 0:
            c5.font = Font(name='Arial', size=10, color='006100')
        elif avg_pmi < 0:
            c5.font = Font(name='Arial', size=10, color='9C0006')

        ws_r.cell(row, 6, cs['pos']).font = cf
        ws_r.cell(row, 7, cs['neg']).font = cf
        ws_r.cell(row, 8, cs['na']).font = cf

    ws_r.column_dimensions['A'].width = 28
    ws_r.column_dimensions['B'].width = 12
    ws_r.column_dimensions['C'].width = 18
    ws_r.column_dimensions['D'].width = 22
    ws_r.column_dimensions['E'].width = 22
    for cl in ['F', 'G', 'H']:
        ws_r.column_dimensions[cl].width = 16

    # ============================================================
    # HOJA DICCIONARIO (formato limpio para scoring)
    # ============================================================
    ws_d = wb_out.create_sheet(f"Diccionario_V{VENTANA}")

    dict_hdrs = ["Categoria", "Palabra", "Delta_WPMI"]
    for c, h in enumerate(dict_hdrs, 1):
        cell = ws_d.cell(1, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    dr = 1
    for row_data in resultados:
        (cat, palabra, cw, cwp, cwn,
         pmi_p, pmi_n, delta,
         wpmi_p, wpmi_n, delta_w) = row_data
        if delta_w is None:
            continue
        dr += 1
        ws_d.cell(dr, 1, cat).font = cf
        ws_d.cell(dr, 2, palabra).font = cf
        cd = ws_d.cell(dr, 3, round(delta_w, 8))
        cd.number_format = num_fmt_8
        if delta_w > 0:
            cd.font = Font(name='Arial', size=10, color='006100')
        elif delta_w < 0:
            cd.font = Font(name='Arial', size=10, color='9C0006')

    ws_d.column_dimensions['A'].width = 28
    ws_d.column_dimensions['B'].width = 22
    ws_d.column_dimensions['C'].width = 16
    ws_d.freeze_panes = 'A2'
    ws_d.auto_filter.ref = f"A1:C{dr}"

    print(f"  Resultados: {len(resultados)} filas")

# ============================================================
# REORDENAR HOJAS: Diccionario primero
# ============================================================
sheet_names = wb_out.sheetnames
dict_sheets = [s for s in sheet_names if s.startswith('Diccionario')]
other_sheets = [s for s in sheet_names if not s.startswith('Diccionario')]
new_order = dict_sheets + other_sheets
wb_out._sheets = [wb_out[s] for s in new_order]

# ============================================================
# GUARDAR
# ============================================================
wb_out.save(OUTPUT_PATH)
print(f"\nArchivo guardado en: {OUTPUT_PATH}")
print("Listo.")
