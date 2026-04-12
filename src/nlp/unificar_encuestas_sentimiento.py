#!/usr/bin/env python3
"""
Une en un solo Excel:
1. Las encuestas mensuales por casa encuestadora
2. El sentimiento digital semanal agregado a nivel mensual

La integracion se hace por mes calendario. Para las semanas ISO se usa el
jueves de la semana como fecha representativa del mes.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
import unicodedata

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path("/home/emilio/Documentos/RAdAR")
ENCUESTAS_PATH = (
    BASE_DIR / "data" / "external" / "surveys" / "Aprobacion_Encuestas_Monica_Villarreal_Oct24_Marzo26.xlsx"
)
SENTIMIENTO_PATH = (
    BASE_DIR / "data" / "processed" / "modeling" / "aceptacion_digital_redes_medios_sentimiento_semanal.xlsx"
)
OUTPUT_PATH = BASE_DIR / "data" / "processed" / "modeling" / "encuestas_y_sentimiento_mensual_unificado.xlsx"

HOJA = "mensual_unificado"

MESES = {
    "enero": 1,
    "ene": 1,
    "febrero": 2,
    "feb": 2,
    "marzo": 3,
    "mar": 3,
    "abril": 4,
    "abr": 4,
    "mayo": 5,
    "may": 5,
    "junio": 6,
    "jun": 6,
    "julio": 7,
    "jul": 7,
    "agosto": 8,
    "ago": 8,
    "septiembre": 9,
    "setiembre": 9,
    "sep": 9,
    "octubre": 10,
    "oct": 10,
    "noviembre": 11,
    "nov": 11,
    "diciembre": 12,
    "dic": 12,
}

MES_NOMBRE = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

def normalizar(texto: str) -> str:
    texto = str(texto).strip().lower()
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(char for char in nfkd if not unicodedata.combining(char))


def parsear_mes(valor: str) -> int:
    clave = normalizar(valor)
    if clave not in MESES:
        raise ValueError(f"Mes no reconocido: {valor}")
    return MESES[clave]


def construir_periodo(df: pd.DataFrame, anio_col: str, mes_col: str) -> pd.DataFrame:
    salida = df.copy()
    salida["anio"] = salida[anio_col].astype(int)
    salida["mes_num"] = salida[mes_col].apply(parsear_mes).astype(int)
    salida["mes_nombre"] = salida["mes_num"].map(MES_NOMBRE)
    salida["anio_mes"] = salida["anio"].astype(str) + "-" + salida["mes_num"].map(lambda x: f"{x:02d}")
    salida["fecha_inicio_mes"] = pd.to_datetime(
        salida["anio_mes"] + "-01",
        format="%Y-%m-%d",
    )
    salida["fecha_fin_mes"] = salida["fecha_inicio_mes"] + pd.offsets.MonthEnd(0)
    return salida


def cargar_encuestas() -> pd.DataFrame:
    mitofsky = pd.read_excel(ENCUESTAS_PATH, sheet_name="Mitofsky")
    mitofsky = construir_periodo(mitofsky, "AÑO", "MES")
    mitofsky = mitofsky.rename(columns={"APROBACIÓN": "mitofsky_aprobacion_pct"})
    mitofsky = mitofsky[
        ["anio", "mes_num", "mes_nombre", "anio_mes", "fecha_inicio_mes", "fecha_fin_mes", "mitofsky_aprobacion_pct"]
    ]

    scr = pd.read_excel(ENCUESTAS_PATH, sheet_name="SCR")
    scr = scr.copy()
    scr["mes_txt"] = scr["Mes"].astype(str).str.split().str[0]
    scr["anio"] = scr["Mes"].astype(str).str.split().str[1].astype(int)
    scr["mes_num"] = scr["mes_txt"].apply(parsear_mes).astype(int)
    scr["mes_nombre"] = scr["mes_num"].map(MES_NOMBRE)
    scr["anio_mes"] = scr["anio"].astype(str) + "-" + scr["mes_num"].map(lambda x: f"{x:02d}")
    scr["fecha_inicio_mes"] = pd.to_datetime(scr["anio_mes"] + "-01", format="%Y-%m-%d")
    scr["fecha_fin_mes"] = scr["fecha_inicio_mes"] + pd.offsets.MonthEnd(0)
    scr["scr_aprobacion_pct"] = scr["Aprobación SRC\n(%)"].astype(float)
    if scr["scr_aprobacion_pct"].max() <= 1.5:
        scr["scr_aprobacion_pct"] = scr["scr_aprobacion_pct"] * 100
    scr = scr[
        ["anio", "mes_num", "mes_nombre", "anio_mes", "fecha_inicio_mes", "fecha_fin_mes", "scr_aprobacion_pct"]
    ]

    rubrum = pd.read_excel(ENCUESTAS_PATH, sheet_name="Rubrum")
    rubrum = construir_periodo(rubrum, "AÑO", "MES")
    rubrum = rubrum.rename(columns={"EVALUACIÓN GENERAL": "rubrum_evaluacion_general_0a10"})
    rubrum = rubrum[
        [
            "anio",
            "mes_num",
            "mes_nombre",
            "anio_mes",
            "fecha_inicio_mes",
            "fecha_fin_mes",
            "rubrum_evaluacion_general_0a10",
        ]
    ]

    demoscopia = pd.read_excel(ENCUESTAS_PATH, sheet_name="Demoscopia")
    demoscopia = construir_periodo(demoscopia, "AÑO", "MES")
    demoscopia = demoscopia.rename(
        columns={
            "APRUEBA": "demoscopia_aprueba_pct",
            "DESAPRUEBA": "demoscopia_desaprueba_pct",
        }
    )
    demoscopia = demoscopia[
        [
            "anio",
            "mes_num",
            "mes_nombre",
            "anio_mes",
            "fecha_inicio_mes",
            "fecha_fin_mes",
            "demoscopia_aprueba_pct",
            "demoscopia_desaprueba_pct",
        ]
    ]

    claves = ["anio", "mes_num", "mes_nombre", "anio_mes", "fecha_inicio_mes", "fecha_fin_mes"]
    encuestas = mitofsky.merge(scr, on=claves, how="outer")
    encuestas = encuestas.merge(rubrum, on=claves, how="outer")
    encuestas = encuestas.merge(demoscopia, on=claves, how="outer")

    return encuestas.sort_values(["anio", "mes_num"]).reset_index(drop=True)


def cargar_sentimiento_mensual() -> pd.DataFrame:
    sentimiento = pd.read_excel(SENTIMIENTO_PATH)
    sentimiento = sentimiento.copy()

    sentimiento["fecha_representativa"] = sentimiento.apply(
        lambda row: date.fromisocalendar(int(row["anio_iso"]), int(row["semana_iso"]), 4),
        axis=1,
    )
    sentimiento["anio"] = pd.to_datetime(sentimiento["fecha_representativa"]).dt.year
    sentimiento["mes_num"] = pd.to_datetime(sentimiento["fecha_representativa"]).dt.month
    sentimiento["mes_nombre"] = sentimiento["mes_num"].map(MES_NOMBRE)
    sentimiento["anio_mes"] = (
        sentimiento["anio"].astype(str) + "-" + sentimiento["mes_num"].map(lambda x: f"{x:02d}")
    )
    sentimiento["fecha_inicio_mes"] = pd.to_datetime(sentimiento["anio_mes"] + "-01", format="%Y-%m-%d")
    sentimiento["fecha_fin_mes"] = sentimiento["fecha_inicio_mes"] + pd.offsets.MonthEnd(0)

    agregaciones = {
        "periodo_iso": ["count", "min", "max"],
        "sentimiento_facebook": "mean",
        "sentimiento_twitter": "mean",
        "sentimiento_youtube": "mean",
        "sentimiento_redes_ponderado": "mean",
        "sentimiento_medios": "mean",
        "promedio_redes_medios": "mean",
    }

    mensual = (
        sentimiento.groupby(
            ["anio", "mes_num", "mes_nombre", "anio_mes", "fecha_inicio_mes", "fecha_fin_mes"],
            as_index=False,
        )
        .agg(agregaciones)
    )

    mensual.columns = [
        "anio",
        "mes_num",
        "mes_nombre",
        "anio_mes",
        "fecha_inicio_mes",
        "fecha_fin_mes",
        "n_semanas_digitales",
        "primera_semana_iso",
        "ultima_semana_iso",
        "sentimiento_facebook_prom_mes",
        "sentimiento_twitter_prom_mes",
        "sentimiento_youtube_prom_mes",
        "sentimiento_redes_ponderado_prom_mes",
        "sentimiento_medios_prom_mes",
        "promedio_redes_medios_prom_mes",
    ]

    return mensual.sort_values(["anio", "mes_num"]).reset_index(drop=True)


def ajustar_excel(writer: pd.ExcelWriter, dataframe: pd.DataFrame) -> None:
    hoja = writer.sheets[HOJA]
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in hoja[1]:
        cell.fill = fill
        cell.font = font

    for column_cells in hoja.columns:
        letra = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        hoja.column_dimensions[letra].width = min(max(max_len + 2, 12), 32)

    formatos = {
        "mitofsky_aprobacion_pct": "0.0",
        "scr_aprobacion_pct": "0.0",
        "rubrum_evaluacion_general_0a10": "0.00",
        "demoscopia_aprueba_pct": "0.0",
        "demoscopia_desaprueba_pct": "0.0",
        "sentimiento_facebook_prom_mes": "0.0000",
        "sentimiento_twitter_prom_mes": "0.0000",
        "sentimiento_youtube_prom_mes": "0.0000",
        "sentimiento_redes_ponderado_prom_mes": "0.0000",
        "sentimiento_medios_prom_mes": "0.0000",
        "promedio_redes_medios_prom_mes": "0.0000",
    }

    columnas = {col: idx + 1 for idx, col in enumerate(dataframe.columns)}
    for nombre_columna, formato in formatos.items():
        indice = columnas.get(nombre_columna)
        if not indice:
            continue
        letra = get_column_letter(indice)
        for row in range(2, hoja.max_row + 1):
            hoja[f"{letra}{row}"].number_format = formato


def guardar_excel(dataframe: pd.DataFrame) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name=HOJA, index=False)
        ajustar_excel(writer, dataframe)


def main() -> None:
    print("=" * 72)
    print("UNIFICACION DE ENCUESTAS Y SENTIMIENTO DIGITAL")
    print("=" * 72)

    encuestas = cargar_encuestas()
    sentimiento_mensual = cargar_sentimiento_mensual()

    claves = ["anio", "mes_num", "mes_nombre", "anio_mes", "fecha_inicio_mes", "fecha_fin_mes"]
    resultado = encuestas.merge(sentimiento_mensual, on=claves, how="outer")
    resultado = resultado.sort_values(["anio", "mes_num"]).reset_index(drop=True)

    columnas = [
        "anio",
        "mes_num",
        "mes_nombre",
        "anio_mes",
        "fecha_inicio_mes",
        "fecha_fin_mes",
        "mitofsky_aprobacion_pct",
        "scr_aprobacion_pct",
        "rubrum_evaluacion_general_0a10",
        "demoscopia_aprueba_pct",
        "demoscopia_desaprueba_pct",
        "n_semanas_digitales",
        "primera_semana_iso",
        "ultima_semana_iso",
        "sentimiento_facebook_prom_mes",
        "sentimiento_twitter_prom_mes",
        "sentimiento_youtube_prom_mes",
        "sentimiento_redes_ponderado_prom_mes",
        "sentimiento_medios_prom_mes",
        "promedio_redes_medios_prom_mes",
    ]
    resultado = resultado[columnas]

    guardar_excel(resultado)

    print(f"Filas unificadas: {len(resultado):,}")
    print(f"Archivo generado: {OUTPUT_PATH}")
    print("=" * 72)


if __name__ == "__main__":
    main()
