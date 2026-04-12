from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from config import (
    DATASET_PATH,
    DATE_COLUMN,
    DEFAULT_HORIZONS,
    DEFAULT_INITIAL_TRAIN_SIZE,
    DEFAULT_LAGS,
    DEFAULT_SHEET_NAME,
    FEATURE_MODE_ALL,
    FEATURE_MODE_CHOICES,
    TARGET_COLUMNS,
    TARGET_MODE_CHOICES,
    TARGET_MODE_LEVEL,
)
from data_master import get_base_feature_columns, load_master_dataset
from evaluation import compute_loss_h, compute_radar_metrics, compute_total_radar_loss, walk_forward_predict
from experiment_logger import DEFAULT_RUNS_DIR, DEFAULT_WORKBOOK, RadarExperimentTracker
from feature_engineering import build_model_frame


def parse_int_sequence(raw_value: str, label: str) -> tuple[int, ...]:
    values = []
    for token in raw_value.split(","):
        token = token.strip()
        if not token:
            continue
        values.append(int(token))
    if not values:
        raise ValueError(f"{label} no puede quedar vacio.")
    return tuple(values)


def parse_float_sequence(raw_value: str, label: str) -> tuple[float, ...]:
    values = []
    for token in raw_value.split(","):
        token = token.strip()
        if not token:
            continue
        values.append(float(token))
    if not values:
        raise ValueError(f"{label} no puede quedar vacio.")
    return tuple(values)


def parse_string_sequence(raw_value: str) -> tuple[str, ...]:
    values = []
    for token in raw_value.split(","):
        token = token.strip()
        if not token:
            continue
        values.append(token)
    return tuple(values)


def parse_lags(raw_value: str) -> tuple[int, ...]:
    return parse_int_sequence(raw_value, label="lags")


def parse_horizons(raw_value: str) -> tuple[int, ...]:
    horizons = parse_int_sequence(raw_value, label="horizons")
    invalid = [horizon for horizon in horizons if horizon not in TARGET_COLUMNS]
    if invalid:
        raise ValueError(f"Horizontes no soportados: {invalid}. Usa {sorted(TARGET_COLUMNS)}.")
    return horizons


def parse_order(raw_value: str) -> tuple[int, int, int]:
    values = parse_int_sequence(raw_value, label="order")
    if len(values) != 3:
        raise ValueError("order debe tener exactamente tres enteros, por ejemplo 1,0,0.")
    return values[0], values[1], values[2]


def add_common_experiment_args(parser: argparse.ArgumentParser, *, default_run_id: str) -> None:
    parser.add_argument("--run-id", default=default_run_id, help="Run_ID a registrar en el grid.")
    parser.add_argument(
        "--target-mode",
        choices=TARGET_MODE_CHOICES,
        default=TARGET_MODE_LEVEL,
        help="Modo de target del modelo: nivel o delta.",
    )
    parser.add_argument(
        "--lags",
        default=",".join(str(lag) for lag in DEFAULT_LAGS),
        help="Lags separados por coma. Ejemplo: 1,2,3,4",
    )
    parser.add_argument(
        "--feature-mode",
        choices=FEATURE_MODE_CHOICES,
        default=FEATURE_MODE_ALL,
        help="Modo de seleccion de features.",
    )
    parser.add_argument(
        "--initial-train-size",
        type=int,
        default=DEFAULT_INITIAL_TRAIN_SIZE,
        help=f"Tamano inicial de entrenamiento. Default: {DEFAULT_INITIAL_TRAIN_SIZE}",
    )
    parser.add_argument(
        "--horizons",
        default=",".join(str(horizon) for horizon in DEFAULT_HORIZONS),
        help="Horizontes a correr, separados por coma.",
    )
    parser.add_argument(
        "--dataset-path",
        type=Path,
        default=DATASET_PATH,
        help=f"Ruta del dataset maestro. Default: {DATASET_PATH}",
    )
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help=f"Hoja del dataset maestro. Default: {DEFAULT_SHEET_NAME}",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Workbook del grid de experimentos. Default: {DEFAULT_WORKBOOK}",
    )
    parser.add_argument(
        "--runs-dir",
        type=Path,
        default=DEFAULT_RUNS_DIR,
        help=f"Directorio donde se guardan runs y artefactos. Default: {DEFAULT_RUNS_DIR}",
    )


def add_reference_comparison_args(
    parser: argparse.ArgumentParser,
    *,
    default_reference_run_id: str,
    include_hypothesis_note: bool = True,
) -> None:
    parser.add_argument(
        "--reference-run-id",
        default=default_reference_run_id,
        help="Run_ID de referencia principal para artefactos comparativos.",
    )
    parser.add_argument(
        "--extra-reference-run-ids",
        default="",
        help="Run_IDs extra para comparaciones JSON, separados por coma.",
    )
    if include_hypothesis_note:
        parser.add_argument(
            "--hypothesis-note",
            default="",
            help="Nota corta de la hipotesis del run para comentarios y grid.",
        )


def finalize_common_args(args: argparse.Namespace) -> argparse.Namespace:
    args.lags = parse_lags(args.lags)
    args.horizons = parse_horizons(args.horizons)
    args.dataset_path = args.dataset_path.expanduser().resolve()
    args.workbook = args.workbook.expanduser().resolve()
    args.runs_dir = args.runs_dir.expanduser().resolve()
    return args


def serialize_model_params(model_params: dict[str, Any] | None) -> dict[str, Any]:
    if not model_params:
        return {}
    serialized: dict[str, Any] = {}
    for key, value in model_params.items():
        if isinstance(value, Path):
            serialized[key] = str(value)
        elif isinstance(value, tuple):
            serialized[key] = list(value)
        else:
            serialized[key] = value
    return serialized


def build_run_parameters(
    args: argparse.Namespace,
    model_name: str,
    feature_columns: list[str],
    model_params: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "dataset_path": str(args.dataset_path),
        "sheet_name": args.sheet_name,
        "model": model_name,
        "target_mode": args.target_mode,
        "feature_mode": args.feature_mode,
        "transform_mode": getattr(args, "transform_mode", ""),
        "feature_columns": feature_columns,
        "lags": list(args.lags),
        "horizons": list(args.horizons),
        "initial_train_size": args.initial_train_size,
        "model_params": serialize_model_params(model_params),
    }


def build_notas_config(
    args: argparse.Namespace,
    model_name: str,
    model_params: dict[str, Any] | None,
    horizon: int,
    summary_row: dict[str, Any],
) -> str:
    payload = {
        "model": model_name,
        "target_mode": args.target_mode,
        "feature_mode": args.feature_mode,
        "transform_mode": getattr(args, "transform_mode", ""),
        "lags": list(args.lags),
        "initial_train_size": args.initial_train_size,
        "horizon": horizon,
        "rows_modeling": summary_row["rows_modeling"],
        "rows_eval": summary_row["rows_eval"],
        "feature_candidates": summary_row["feature_candidates"],
        "selected_feature_count_avg": summary_row["selected_feature_count_avg"],
        "model_params": serialize_model_params(model_params),
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def save_run_outputs(
    run,
    predictions_by_horizon: dict[int, Any],
    rows_summary: list[dict[str, Any]],
) -> None:
    for horizon, predictions in predictions_by_horizon.items():
        run.save_dataframe(
            predictions,
            f"predicciones_h{horizon}.csv",
            artifact_type="predicciones",
            notes=f"Predicciones walk-forward para horizonte {horizon}.",
        )

    run.save_json(
        rows_summary,
        "resumen_modeling_horizontes.json",
        artifact_type="resumen",
        notes="Resumen de tamanos y seleccion de features por horizonte.",
    )


def build_selected_features_summary(predictions: pd.DataFrame) -> pd.DataFrame:
    feature_sets = (
        predictions["selected_features"]
        .fillna("")
        .astype(str)
        .map(lambda value: value.strip())
    )
    summary = (
        feature_sets.value_counts(dropna=False)
        .rename_axis("selected_features")
        .reset_index(name="fold_count")
    )
    summary["feature_count"] = summary["selected_features"].map(
        lambda value: 0 if not value else len([token for token in value.split(",") if token])
    )
    summary["share_folds"] = summary["fold_count"] / max(len(predictions), 1)
    return summary


def normalize_reference_run_ids(
    reference_run_id: str | None,
    extra_reference_run_ids: tuple[str, ...] | list[str] | None = None,
) -> tuple[str, ...]:
    ordered = []
    if reference_run_id:
        ordered.append(reference_run_id)
    for run_id in extra_reference_run_ids or ():
        if run_id and run_id not in ordered:
            ordered.append(run_id)
    return tuple(ordered)


def load_run_horizon_metrics(workbook_path: Path, run_id: str) -> dict[int, dict[str, float]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook["RESULTADOS_GRID"]
    headers = {worksheet.cell(1, c).value: c for c in range(1, worksheet.max_column + 1)}
    metrics_by_horizon: dict[int, dict[str, float]] = {}
    for row_idx in range(2, worksheet.max_row + 1):
        if worksheet.cell(row_idx, headers["Run_ID"]).value != run_id:
            continue
        horizon = int(worksheet.cell(row_idx, headers["Horizonte_sem"]).value)
        metrics_by_horizon[horizon] = {
            "loss_h": float(worksheet.cell(row_idx, headers["Loss_h"]).value),
            "mae": float(worksheet.cell(row_idx, headers["MAE"]).value),
            "rmse": float(worksheet.cell(row_idx, headers["RMSE"]).value),
            "direccion_accuracy": float(worksheet.cell(row_idx, headers["Direccion_accuracy"]).value),
            "deteccion_caidas": float(worksheet.cell(row_idx, headers["Deteccion_caidas"]).value),
        }
    if not metrics_by_horizon:
        raise ValueError(f"No se encontraron metricas para {run_id} en {workbook_path}.")
    return metrics_by_horizon


def build_comparison_payload(
    *,
    workbook_path: Path,
    reference_run_id: str,
    clean_run_id: str,
    horizon_results: list[dict[str, Any]],
    l_total_radar: float,
) -> dict[str, Any]:
    reference = load_run_horizon_metrics(workbook_path=workbook_path, run_id=reference_run_id)
    clean = {int(item["horizonte_sem"]): item for item in horizon_results}

    comparison_rows = []
    for horizon in sorted(clean):
        reference_row = reference[horizon]
        clean_row = clean[horizon]
        comparison_rows.append(
            {
                "horizonte_sem": horizon,
                "loss_h_original": reference_row["loss_h"],
                "loss_h_clean": float(clean_row["loss_h"]),
                "delta_loss_h": float(clean_row["loss_h"]) - reference_row["loss_h"],
                "mae_original": reference_row["mae"],
                "mae_clean": float(clean_row["mae"]),
                "delta_mae": float(clean_row["mae"]) - reference_row["mae"],
                "rmse_original": reference_row["rmse"],
                "rmse_clean": float(clean_row["rmse"]),
                "delta_rmse": float(clean_row["rmse"]) - reference_row["rmse"],
                "direction_accuracy_original": reference_row["direccion_accuracy"],
                "direction_accuracy_clean": float(clean_row["direccion_accuracy"]),
                "delta_direction_accuracy": float(clean_row["direccion_accuracy"]) - reference_row["direccion_accuracy"],
                "risk_detection_original": reference_row["deteccion_caidas"],
                "risk_detection_clean": float(clean_row["deteccion_caidas"]),
                "delta_risk_detection": float(clean_row["deteccion_caidas"]) - reference_row["deteccion_caidas"],
            }
        )

    return {
        "reference_run_id": reference_run_id,
        "clean_run_id": clean_run_id,
        "l_total_radar_clean": float(l_total_radar),
        "comparacion_por_horizonte": comparison_rows,
    }


def save_reference_comparisons(
    *,
    run,
    workbook_path: Path,
    clean_run_id: str,
    reference_run_ids: tuple[str, ...],
    horizon_results: list[dict[str, Any]],
    l_total_radar: float,
) -> None:
    for reference_run_id in reference_run_ids:
        comparison_payload = build_comparison_payload(
            workbook_path=workbook_path,
            reference_run_id=reference_run_id,
            clean_run_id=clean_run_id,
            horizon_results=horizon_results,
            l_total_radar=l_total_radar,
        )
        run.save_json(
            comparison_payload,
            f"comparacion_vs_{reference_run_id}.json",
            artifact_type="comparacion",
            notes=f"Comparacion de la corrida clean contra {reference_run_id}.",
        )


def run_tabular_experiment(
    *,
    args: argparse.Namespace,
    script_path: str | Path,
    experiment_id: str,
    family: str,
    model_name: str,
    estimator,
    model_params: dict[str, Any] | None = None,
    transformacion: str = "",
    comentarios: str = "",
    always_include_columns: list[str] | None = None,
    reference_run_ids: tuple[str, ...] = (),
    l_coh: float | None = None,
) -> dict[str, Any]:
    tracker = RadarExperimentTracker(workbook_path=args.workbook, runs_dir=args.runs_dir)
    reference_values = tracker.get_reference_values()
    df = load_master_dataset(dataset_path=args.dataset_path, sheet_name=args.sheet_name)
    base_feature_columns = get_base_feature_columns(df)

    run = tracker.start_run(
        run_id=args.run_id,
        experiment_id=experiment_id,
        family=family,
        model=model_name,
        script_path=script_path,
        parametros=build_run_parameters(
            args=args,
            model_name=model_name,
            feature_columns=base_feature_columns,
            model_params=model_params,
        ),
    )

    predictions_by_horizon: dict[int, Any] = {}
    horizon_results: list[dict[str, Any]] = []
    rows_summary: list[dict[str, Any]] = []
    dataset_periodo = f"{df[DATE_COLUMN].min().date()} a {df[DATE_COLUMN].max().date()}"

    for horizon in args.horizons:
        modeling_df, modeling_features, target_column = build_model_frame(
            df=df,
            horizon=horizon,
            feature_columns=base_feature_columns,
            lags=args.lags,
            target_mode=args.target_mode,
        )
        predictions = walk_forward_predict(
            estimator=estimator,
            data=modeling_df,
            feature_columns=modeling_features,
            target_column=target_column,
            actual_target_column=TARGET_COLUMNS[horizon],
            initial_train_size=args.initial_train_size,
            feature_mode=args.feature_mode,
            target_mode=args.target_mode,
            always_include_columns=always_include_columns,
        )
        predictions["horizonte_sem"] = horizon
        predictions["target_mode"] = args.target_mode
        predictions["feature_mode"] = args.feature_mode
        predictions["transform_mode"] = getattr(args, "transform_mode", "")
        predictions["run_id"] = args.run_id
        predictions["model_name"] = model_name
        predictions_by_horizon[horizon] = predictions
        if args.feature_mode != FEATURE_MODE_ALL:
            selected_features_summary = build_selected_features_summary(predictions)
            run.save_dataframe(
                selected_features_summary,
                f"features_seleccionadas_h{horizon}.csv",
                artifact_type="seleccion_features",
                notes=(
                    "Resumen de combinaciones de features seleccionadas por fold externo. "
                    "El filtro se calcula solo con el train de cada fold externo."
                ),
            )

        summary_row = {
            "horizonte_sem": horizon,
            "rows_modeling": len(modeling_df),
            "rows_eval": len(predictions),
            "feature_candidates": len(modeling_features),
            "selected_feature_count_avg": float(predictions["selected_feature_count"].mean()),
            "selected_feature_count_min": int(predictions["selected_feature_count"].min()),
            "selected_feature_count_max": int(predictions["selected_feature_count"].max()),
            "feature_mode": args.feature_mode,
            "target_mode": args.target_mode,
            "transform_mode": getattr(args, "transform_mode", ""),
        }
        rows_summary.append(summary_row)

        metrics = compute_radar_metrics(predictions)
        loss_h = compute_loss_h(metrics=metrics, horizon=horizon, reference_values=reference_values)
        horizon_results.append(
            {
                "horizonte_sem": horizon,
                "target": args.target_mode,
                "variables_temporales": f"y_t + lags {list(args.lags)}",
                "variables_tematicas": ", ".join(base_feature_columns),
                "transformacion": transformacion,
                "seleccion_variables": args.feature_mode,
                "validacion": "walk-forward_expanding",
                "dataset_periodo": dataset_periodo,
                "notas_config": build_notas_config(
                    args=args,
                    model_name=model_name,
                    model_params=model_params,
                    horizon=horizon,
                    summary_row=summary_row,
                ),
                "estado": "corrido",
                "comentarios": comentarios,
                "loss_h": loss_h,
                **metrics,
            }
        )

    save_run_outputs(run=run, predictions_by_horizon=predictions_by_horizon, rows_summary=rows_summary)
    total_radar = compute_total_radar_loss(
        horizon_results=horizon_results,
        reference_values=reference_values,
        l_coh=l_coh,
    )
    if reference_run_ids:
        save_reference_comparisons(
            run=run,
            workbook_path=tracker.workbook_path,
            clean_run_id=args.run_id,
            reference_run_ids=reference_run_ids,
            horizon_results=horizon_results,
            l_total_radar=float(total_radar["l_total_radar"]),
        )

    run.finalize(
        horizon_results=horizon_results,
        target=args.target_mode,
        variables_temporales=f"y_t + lags {list(args.lags)}",
        variables_tematicas=", ".join(base_feature_columns),
        transformacion=transformacion,
        seleccion_variables=args.feature_mode,
        validacion="walk-forward_expanding",
        dataset_periodo=dataset_periodo,
        notas_config=json.dumps(rows_summary, ensure_ascii=False),
        estado="corrido",
        comentarios=(
            f"{comentarios} | L_total_Radar={total_radar['l_total_radar']:.6f}"
            if comentarios
            else f"L_total_Radar={total_radar['l_total_radar']:.6f}"
        ),
        l_coh=l_coh,
    )

    return {
        "workbook_path": tracker.workbook_path,
        "run_dir": run.run_dir,
        "horizon_results": horizon_results,
        "rows_summary": rows_summary,
        "l_total_radar": total_radar["l_total_radar"],
    }


def run_time_series_experiment(**kwargs) -> dict[str, Any]:
    return run_tabular_experiment(**kwargs)
