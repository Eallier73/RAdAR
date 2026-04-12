#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.linear_model import Ridge
from sklearn.pipeline import Pipeline

from config import (
    DATE_COLUMN,
    FEATURE_MODE_ALL,
    TARGET_COLUMNS,
    TRANSFORM_MODE_CHOICES,
    TRANSFORM_MODE_STANDARD,
)
from data_master import get_base_feature_columns, load_master_dataset
from evaluation import (
    compute_loss_h,
    compute_radar_metrics,
    compute_total_radar_loss,
    walk_forward_predict_with_tscv_alpha_tuning,
)
from experiment_logger import RadarExperimentTracker
from feature_engineering import build_model_frame
from pipeline_common import (
    add_common_experiment_args,
    build_notas_config,
    build_run_parameters,
    finalize_common_args,
    save_run_outputs,
)
from preprocessing import build_feature_transformer, describe_transform_mode


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="E1_v2_clean | Ridge con tuning temporal interno por TimeSeriesSplit.",
    )
    add_common_experiment_args(parser, default_run_id="E1_v2_clean")
    parser.add_argument(
        "--reference-run-id",
        default="E1_v2_clean",
        help="Run_ID de referencia para artefactos comparativos.",
    )
    parser.add_argument(
        "--transform-mode",
        choices=TRANSFORM_MODE_CHOICES,
        default=TRANSFORM_MODE_STANDARD,
        help="Transformacion aplicada dentro del pipeline de Ridge.",
    )
    parser.add_argument("--winsor-lower-quantile", type=float, default=0.05)
    parser.add_argument("--winsor-upper-quantile", type=float, default=0.95)
    parser.add_argument("--alpha-grid-min-exp", type=float, default=-4.0)
    parser.add_argument("--alpha-grid-max-exp", type=float, default=4.0)
    parser.add_argument("--alpha-grid-points", type=int, default=40)
    parser.add_argument("--inner-splits", type=int, default=3)
    parser.add_argument("--alpha-selection-metric", choices=("mae", "rmse"), default="mae")
    args = finalize_common_args(parser.parse_args())
    if not 0.0 <= args.winsor_lower_quantile < args.winsor_upper_quantile <= 1.0:
        raise ValueError("Los cuantiles de winsor deben cumplir 0 <= lower < upper <= 1.")
    return args


def build_estimator(alpha: float, args: argparse.Namespace) -> Pipeline:
    return Pipeline(
        steps=[
            (
                "transform",
                build_feature_transformer(
                    transform_mode=args.transform_mode,
                    winsor_lower_quantile=args.winsor_lower_quantile,
                    winsor_upper_quantile=args.winsor_upper_quantile,
                ),
            ),
            ("model", Ridge(alpha=float(alpha))),
        ]
    )


def build_model_params(args: argparse.Namespace, alpha_grid: list[float]) -> dict[str, object]:
    return {
        "alpha_grid": alpha_grid,
        "inner_splits": args.inner_splits,
        "alpha_selection_metric": args.alpha_selection_metric,
        "transform_mode": args.transform_mode,
        "winsor_lower_quantile": args.winsor_lower_quantile,
        "winsor_upper_quantile": args.winsor_upper_quantile,
    }


def summarize_alpha_trace(alpha_trace: list[dict[str, object]]) -> dict[str, object]:
    best_alphas = [float(item["best_alpha"]) for item in alpha_trace]
    return {
        "best_alpha_mean": float(np.mean(best_alphas)),
        "best_alpha_median": float(np.median(best_alphas)),
        "best_alpha_min": float(np.min(best_alphas)),
        "best_alpha_max": float(np.max(best_alphas)),
        "best_alphas_by_fold": alpha_trace,
    }


def build_selected_features_summary(predictions: pd.DataFrame) -> pd.DataFrame:
    feature_sets = (
        predictions["selected_features"]
        .fillna("")
        .astype(str)
        .map(lambda value: value.strip())
    )
    summary = (
        feature_sets.value_counts(dropna=False)
        .rename_axis("selected_features")
        .reset_index(name="fold_count")
    )
    summary["feature_count"] = summary["selected_features"].map(
        lambda value: 0 if not value else len([token for token in value.split(",") if token])
    )
    summary["share_folds"] = summary["fold_count"] / max(len(predictions), 1)
    return summary


def load_run_horizon_metrics(workbook_path: Path, run_id: str) -> dict[int, dict[str, float]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook["RESULTADOS_GRID"]
    headers = {worksheet.cell(1, c).value: c for c in range(1, worksheet.max_column + 1)}
    metrics_by_horizon: dict[int, dict[str, float]] = {}
    for row_idx in range(2, worksheet.max_row + 1):
        if worksheet.cell(row_idx, headers["Run_ID"]).value != run_id:
            continue
        horizon = int(worksheet.cell(row_idx, headers["Horizonte_sem"]).value)
        metrics_by_horizon[horizon] = {
            "loss_h": float(worksheet.cell(row_idx, headers["Loss_h"]).value),
            "mae": float(worksheet.cell(row_idx, headers["MAE"]).value),
            "rmse": float(worksheet.cell(row_idx, headers["RMSE"]).value),
            "direccion_accuracy": float(worksheet.cell(row_idx, headers["Direccion_accuracy"]).value),
            "deteccion_caidas": float(worksheet.cell(row_idx, headers["Deteccion_caidas"]).value),
        }
    if not metrics_by_horizon:
        raise ValueError(f"No se encontraron metricas para {run_id} en {workbook_path}.")
    return metrics_by_horizon


def build_comparison_payload(
    *,
    workbook_path: Path,
    reference_run_id: str,
    clean_run_id: str,
    horizon_results: list[dict[str, object]],
    l_total_radar: float,
) -> dict[str, object]:
    reference = load_run_horizon_metrics(workbook_path=workbook_path, run_id=reference_run_id)
    clean = {int(item["horizonte_sem"]): item for item in horizon_results}

    comparison_rows = []
    for horizon in sorted(clean):
        reference_row = reference[horizon]
        clean_row = clean[horizon]
        comparison_rows.append(
            {
                "horizonte_sem": horizon,
                "loss_h_original": reference_row["loss_h"],
                "loss_h_clean": float(clean_row["loss_h"]),
                "delta_loss_h": float(clean_row["loss_h"]) - reference_row["loss_h"],
                "mae_original": reference_row["mae"],
                "mae_clean": float(clean_row["mae"]),
                "delta_mae": float(clean_row["mae"]) - reference_row["mae"],
                "rmse_original": reference_row["rmse"],
                "rmse_clean": float(clean_row["rmse"]),
                "delta_rmse": float(clean_row["rmse"]) - reference_row["rmse"],
                "direction_accuracy_original": reference_row["direccion_accuracy"],
                "direction_accuracy_clean": float(clean_row["direccion_accuracy"]),
                "delta_direction_accuracy": float(clean_row["direccion_accuracy"]) - reference_row["direccion_accuracy"],
                "risk_detection_original": reference_row["deteccion_caidas"],
                "risk_detection_clean": float(clean_row["deteccion_caidas"]),
                "delta_risk_detection": float(clean_row["deteccion_caidas"]) - reference_row["deteccion_caidas"],
            }
        )

    return {
        "reference_run_id": reference_run_id,
        "clean_run_id": clean_run_id,
        "l_total_radar_clean": float(l_total_radar),
        "comparacion_por_horizonte": comparison_rows,
    }


def main() -> None:
    args = parse_args()
    alpha_grid = list(
        np.logspace(
            args.alpha_grid_min_exp,
            args.alpha_grid_max_exp,
            args.alpha_grid_points,
        )
    )

    tracker = RadarExperimentTracker(workbook_path=args.workbook, runs_dir=args.runs_dir)
    reference_values = tracker.get_reference_values()
    df = load_master_dataset(dataset_path=args.dataset_path, sheet_name=args.sheet_name)
    base_feature_columns = get_base_feature_columns(df)

    run = tracker.start_run(
        run_id=args.run_id,
        experiment_id="E1",
        family="lineal_regularizado",
        model="ridge_tscv",
        script_path=__file__,
        parametros=build_run_parameters(
            args=args,
            model_name="ridge_tscv",
            feature_columns=base_feature_columns,
            model_params=build_model_params(args, alpha_grid),
        ),
    )

    predictions_by_horizon = {}
    horizon_results = []
    rows_summary = []
    alpha_summary_payload = []
    dataset_periodo = f"{df[DATE_COLUMN].min().date()} a {df[DATE_COLUMN].max().date()}"

    for horizon in args.horizons:
        modeling_df, modeling_features, target_column = build_model_frame(
            df=df,
            horizon=horizon,
            feature_columns=base_feature_columns,
            lags=args.lags,
            target_mode=args.target_mode,
        )
        predictions, alpha_trace = walk_forward_predict_with_tscv_alpha_tuning(
            estimator_builder=lambda alpha: build_estimator(alpha, args),
            alpha_grid=alpha_grid,
            data=modeling_df,
            feature_columns=modeling_features,
            target_column=target_column,
            actual_target_column=TARGET_COLUMNS[horizon],
            initial_train_size=args.initial_train_size,
            feature_mode=args.feature_mode,
            target_mode=args.target_mode,
            tuning_metric=args.alpha_selection_metric,
            inner_splits=args.inner_splits,
        )

        predictions["horizonte_sem"] = horizon
        predictions["target_mode"] = args.target_mode
        predictions["feature_mode"] = args.feature_mode
        predictions["transform_mode"] = args.transform_mode
        predictions["run_id"] = args.run_id
        predictions["model_name"] = "ridge_tscv"
        predictions_by_horizon[horizon] = predictions
        if args.feature_mode != FEATURE_MODE_ALL:
            selected_features_summary = build_selected_features_summary(predictions)
            run.save_dataframe(
                selected_features_summary,
                f"features_seleccionadas_h{horizon}.csv",
                artifact_type="seleccion_features",
                notes=(
                    "Resumen de combinaciones de features seleccionadas por fold externo. "
                    "El filtro se calcula solo con el train de cada fold externo."
                ),
            )

        alpha_summary = summarize_alpha_trace(alpha_trace)
        alpha_summary_payload.append(
            {
                "horizonte_sem": horizon,
                "tuning_metric": args.alpha_selection_metric,
                "inner_splits": args.inner_splits,
                **alpha_summary,
            }
        )

        summary_row = {
            "horizonte_sem": horizon,
            "rows_modeling": len(modeling_df),
            "rows_eval": len(predictions),
            "feature_candidates": len(modeling_features),
            "selected_feature_count_avg": float(predictions["selected_feature_count"].mean()),
            "selected_feature_count_min": int(predictions["selected_feature_count"].min()),
            "selected_feature_count_max": int(predictions["selected_feature_count"].max()),
            "feature_mode": args.feature_mode,
            "target_mode": args.target_mode,
            "transform_mode": args.transform_mode,
            "best_alpha_mean": alpha_summary["best_alpha_mean"],
            "best_alpha_median": alpha_summary["best_alpha_median"],
            "best_alpha_min": alpha_summary["best_alpha_min"],
            "best_alpha_max": alpha_summary["best_alpha_max"],
        }
        rows_summary.append(summary_row)

        metrics = compute_radar_metrics(predictions)
        loss_h = compute_loss_h(metrics=metrics, horizon=horizon, reference_values=reference_values)
        horizon_results.append(
            {
                "horizonte_sem": horizon,
                "target": args.target_mode,
                "variables_temporales": f"y_t + lags {list(args.lags)}",
                "variables_tematicas": ", ".join(base_feature_columns),
                "transformacion": describe_transform_mode(args.transform_mode),
                "seleccion_variables": args.feature_mode,
                "validacion": "walk-forward_expanding",
                "dataset_periodo": dataset_periodo,
                "notas_config": build_notas_config(
                    args=args,
                    model_name="ridge_tscv",
                    model_params=build_model_params(args, alpha_grid),
                    horizon=horizon,
                    summary_row=summary_row,
                ),
                "estado": "corrido",
                "comentarios": (
                    f"Ridge limpio con tuning temporal interno | "
                    f"metric={args.alpha_selection_metric} | inner_splits={args.inner_splits}"
                ),
                "loss_h": loss_h,
                **metrics,
            }
        )

    save_run_outputs(run=run, predictions_by_horizon=predictions_by_horizon, rows_summary=rows_summary)
    run.save_json(
        alpha_summary_payload,
        "alpha_tuning_horizontes.json",
        artifact_type="tuning",
        notes="Best alpha por fold externo y resumen por horizonte.",
    )

    total_radar = compute_total_radar_loss(
        horizon_results=horizon_results,
        reference_values=reference_values,
        l_coh=None,
    )
    comparison_payload = build_comparison_payload(
        workbook_path=args.workbook,
        reference_run_id=args.reference_run_id,
        clean_run_id=args.run_id,
        horizon_results=horizon_results,
        l_total_radar=float(total_radar["l_total_radar"]),
    )
    comparison_filename = f"comparacion_vs_{args.reference_run_id}.json"
    run.save_json(
        comparison_payload,
        comparison_filename,
        artifact_type="comparacion",
        notes=f"Comparacion de la corrida clean contra {args.reference_run_id}.",
    )

    run.finalize(
        horizon_results=horizon_results,
        target=args.target_mode,
        variables_temporales=f"y_t + lags {list(args.lags)}",
        variables_tematicas=", ".join(base_feature_columns),
        transformacion=describe_transform_mode(args.transform_mode),
        seleccion_variables=args.feature_mode,
        validacion="walk-forward_expanding",
        dataset_periodo=dataset_periodo,
        notas_config=json.dumps(rows_summary, ensure_ascii=False),
        estado="corrido",
        comentarios=(
            f"Ridge clean con tuning interno TimeSeriesSplit | "
            f"metric={args.alpha_selection_metric} | "
            f"L_total_Radar={total_radar['l_total_radar']:.6f}"
        ),
        l_coh=None,
    )

    print(f"Run registrado: {args.run_id}")
    print(f"Horizontes procesados: {', '.join(str(h) for h in args.horizons)}")
    print(f"L_total_Radar: {total_radar['l_total_radar']:.6f}")


if __name__ == "__main__":
    main()
